from django.http import HttpResponse
from django.shortcuts import render, redirect
from listings.models import Livraison
from .models import Livreur
from .models import Tacheafaire
from .models import Journee
from .models import Photo
import base64
from .models import Phototaches
from django.db.models import Sum
from .models import Route, Recupfrigo
from django.views.generic.list import ListView
from .models import Distances
from .models import Checklist, Recuplivreur
from django.shortcuts import get_object_or_404
from .forms import *
from .forms import LivraisonFeuilleForm
from .forms import LivraisonDragForm
from .forms import LivraisonDragFormtoday, TaskUpdateForm
from .forms import LivraisonsVentesForm, RoutedetailForm
from django.urls import reverse
import json
from django.db.models.functions import TruncDate
from django.shortcuts import render, redirect
from django.core.paginator import Paginator
from django.contrib import messages
from .models import Checklist
from .forms import ChecklistForm, DateFilterForm
from django.utils import timezone
import calendar
from django.shortcuts import render
from django.http import JsonResponse
from .models import Checklist
from .forms import DistanceForm
from tablib import Dataset
from .ressources import LivraisonResource
from django.utils.timezone import now
from datetime import datetime, timedelta, time, date
from .models import Recuperation
import googlemaps
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
from django.views import View
from datetime import datetime
from django.views.decorators.http import require_POST
from django.http import QueryDict
from django.shortcuts import render, get_object_or_404
from .models import *
from .models import ItemInv
from django.shortcuts import render
from django.contrib import messages
from .models import Item
from openpyxl import load_workbook
from tempfile import NamedTemporaryFile
from urllib.request import urlopen
from django.core.files import File
from .forms import ItemInvForm
from .forms import SearchFormInv
from .models import Product, Checklist, ChecklistItem
from .forms import RouteForm
from .forms import *
from .forms import ChecklistForm
from django.http import FileResponse, HttpResponseRedirect, HttpResponse
from .utils import add_quantity_to_checklist, remove_quantity_from_checklist
from .forms import DateFilterForm
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.views.generic.edit import UpdateView
from django.http import HttpResponseForbidden
from django.urls import reverse_lazy
from django.views.generic.edit import DeleteView
from django.utils import timezone
import random
import requests
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.db import transaction
from django.contrib import messages


# views.py
from django.shortcuts import redirect
from django.contrib.auth import authenticate, login
from django.contrib.auth.forms import PasswordChangeForm
from listings.models import UserProfile  # Replace with your app name

from django.shortcuts import render, get_object_or_404
from .forms import LivraisonDragForm
from .models import Livraison
from django.http import JsonResponse
from django.template.loader import render_to_string


from django.shortcuts import get_object_or_404
from django.template.loader import render_to_string
from django.http import JsonResponse
from .models import Livraison
from .forms import LivraisonDragForm

import pandas as pd
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Livraison

from django.shortcuts import render, redirect
from .forms import XLSXUploadForm  # Ensure you import the form

from django.http import JsonResponse
import pandas as pd
import io


from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.contrib.auth.models import User

from django.http import JsonResponse, HttpResponseNotAllowed
import json
from django.utils.translation import activate, deactivate

@login_required
@csrf_exempt
def change_submission_user(request, submission_id):
    if request.method != 'POST':
        return HttpResponseNotAllowed(['POST'])

    try:
        data = json.loads(request.body)
        new_user_id = data.get('user_id')
        if not new_user_id:
            return JsonResponse({'success': False, 'error': 'No user_id provided'})
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'})
        
    new_user = get_object_or_404(User, id=new_user_id)
    submission = get_object_or_404(Submission, id=submission_id)

    old_user = submission.user.username
    submission.user = new_user
    submission.save()
    activate('fr')
    formatted_date = formats.date_format(submission.date, "l, d F Y", use_l10n=True) 
    deactivate()
    # Notification
    message = f"{old_user} vous a attribué la soumission '{submission.company_name} pour le {formatted_date}'"
    link = request.build_absolute_uri(submission.get_absolute_url())
    Notification.objects.create(recipient=new_user, message=message, link=link)

    return JsonResponse({'success': True})

from django.views.decorators.http import require_GET, require_POST

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json

@csrf_exempt
def add_note(request, submission_id):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            note_content = data.get('note', '')
            submission = Submission.objects.get(id=submission_id)
            submission.note = note_content
            submission.save()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Invalid request'})

# views.py
from django.http import JsonResponse
from .models import Submission

def submissions_by_date(request, date):
    # Filtrer par date
    user = request.user
    submissions = Submission.objects.filter(user=user, date=date)
    data = [
        {
            'id': sub.id,
            'company_name': sub.company_name,
            'submission_type': sub.submission_type,
            'status': sub.status,
            'ordered_by': sub.ordered_by,
            'note': sub.note,
        }
        for sub in submissions
    ]
    return JsonResponse({'submissions': data})


@require_POST  # ou @require_POST selon la méthode utilisée
def mark_notification_as_read(request, notification_id):
    notification = get_object_or_404(Notification, id=notification_id, recipient=request.user)
    notification.is_read = True
    notification.save()
    return JsonResponse({'status': 'ok'})


from django.shortcuts import get_object_or_404
from django.http import JsonResponse
from .models import Submission, MenuSubmission
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST

@csrf_exempt
@require_POST
def duplicate_submission(request, submission_id):
    # Get the existing submission
    original_submission = get_object_or_404(Submission, id=submission_id)

    # Copy the submission, including all specified fields
    new_submission = Submission.objects.create(
        user=original_submission.user,
        submission_type=original_submission.submission_type,
        type_prise_de_commande=original_submission.type_prise_de_commande,
        company_name=f"COPIE DE {original_submission.company_name}",
        event_location=original_submission.event_location,
        contact_person=original_submission.contact_person,
        ordered_by=original_submission.ordered_by,
        phone=original_submission.phone,
        email=original_submission.email,
        billing_address=original_submission.billing_address,
        etage=original_submission.etage,
        dock_livraison=original_submission.dock_livraison,
        escalier=original_submission.escalier,
        ascenseur=original_submission.ascenseur,
        carte_dock=original_submission.carte_dock,
        avec_service=original_submission.avec_service,
        avec_service_md=original_submission.avec_service_md,
        language=original_submission.language,
        location_materiel=original_submission.location_materiel,
        commentaire=original_submission.commentaire,
        payment_mode=original_submission.payment_mode,
        client=original_submission.client,
        commentaire_boissons=original_submission.commentaire_boissons,
        commentaire_items=original_submission.commentaire_items,
        document=original_submission.document,
        event_postcode=original_submission.event_postcode,
        billing_postcode=original_submission.billing_postcode,
    )

    # Copy MenuSubmissions associated with the original
    menu_submissions = MenuSubmission.objects.filter(submission=original_submission)
    for menu_submission in menu_submissions:
        MenuSubmission.objects.create(
            submission=new_submission,
            menu=menu_submission.menu,
            allergies=menu_submission.allergies,
            service_count=menu_submission.service_count,
            commentaire_menu=menu_submission.commentaire_menu,
            delivery_mode=menu_submission.delivery_mode,
        )

    return HttpResponseRedirect(reverse('submission_detail', args=[new_submission.id]))



@csrf_exempt
@require_POST
def duplicate_submission_without_menus(request, submission_id):
    # Get the existing submission
    original_submission = get_object_or_404(Submission, id=submission_id)

    # Copy the submission without menus and specific comments
    new_submission = Submission.objects.create(
        user=original_submission.user,
        submission_type=original_submission.submission_type,
        type_prise_de_commande=original_submission.type_prise_de_commande,
        company_name=f"COPIE DE {original_submission.company_name}",
        event_location=original_submission.event_location,
        contact_person=original_submission.contact_person,
        ordered_by=original_submission.ordered_by,
        phone=original_submission.phone,
        email=original_submission.email,
        billing_address=original_submission.billing_address,
        etage=original_submission.etage,
        dock_livraison=original_submission.dock_livraison,
        escalier=original_submission.escalier,
        ascenseur=original_submission.ascenseur,
        carte_dock=original_submission.carte_dock,
        avec_service=original_submission.avec_service,
        avec_service_md=original_submission.avec_service_md,
        language=original_submission.language,
        location_materiel=original_submission.location_materiel,
        commentaire=original_submission.commentaire,
        payment_mode=original_submission.payment_mode,
        client=original_submission.client,
        document=original_submission.document,
        event_postcode=original_submission.event_postcode,
        billing_postcode=original_submission.billing_postcode,
        # Exclude the following:
        # commentaire_boissons=original_submission.commentaire_boissons,
        # commentaire_items=original_submission.commentaire_items,
    )

    return HttpResponseRedirect(reverse('submission_detail', args=[new_submission.id]))



@login_required
def import_xlsx(request):
    print("📂 import_xlsx view was called!")
    form = XLSXUploadForm()
    if request.method == 'POST':
        print("✅ POST request received!")
        form = XLSXUploadForm(request.POST, request.FILES)

        if form.is_valid():
            print("✔️ Form is valid!")
            file = form.cleaned_data['file']

            if not file:
                print("❌ No file received.")
                messages.error(request, "❌ No file uploaded. Please select a file.")
                return redirect('import_xlsx')

            try:
                df = pd.read_excel(file, skiprows=3)  # Skip first 3 rows
                print("📊 File read successfully!")

                print("Columns found in file:", df.columns)  # Debugging

                # Correct column renaming
                df.rename(columns={
                    "Nom de l'événement": "new_column1",
                    "Adresse": "new_column2",
                    "APP": "new_column3",
                    "Ligne 2": "new_column4",
                    "Code postal": "new_column5",
                    "Nb convives": "new_column6",
                    "Mode d'envoi": "new_column7",
                    "Heure livraison": "new_column8",
                    "# Commande": "new_column9",
                    "Nom du client commandé": "new_column10",
                    "Livraison et personne à contacter sur le site": "new_column11",
                    "Informations supplémentaires": "new_column12",
                    "Nom du conseiller": "new_column13"
                }, inplace=True)

                df.insert(0, 'id', '')  # Insert empty 'id' column
                df = df.fillna('')

                # Add heure_livraison_classement with same content as heure_livraison
                df['heure_livraison_classement'] = df['new_column8']

                print("🛠️ Transformed DataFrame columns:", df.columns)

                # Convert DataFrame to model instances
                for index, row in df.iterrows():
                    print(f"➡️ Processing row {index + 1}")
                    try:
                        Livraison.objects.create(
                            nom=row['new_column1'],
                            adress=row['new_column2'],
                            app=row['new_column3'],
                            ligne2=row['new_column4'],
                            zipcode=row['new_column5'],
                            convives=row['new_column6'],
                            mode_envoi=row['new_column7'],
                            heure_livraison=row['new_column8'],
                            heure_livraison_classement=row['heure_livraison_classement'],  # Added field
                            num_commande=row['new_column9'],
                            nom_client=row['new_column10'],
                            contact_site=row['new_column11'],
                            infodetail=row['new_column12'],
                            vendeur=row['new_column13'],
                        )
                        print(f"✅ Livraison created for row {index + 1}")
                    except Exception as e:
                        print(f"❌ Error creating row {index + 1}: {e}")

                messages.success(request, "✅ File imported successfully!")
                print("🎉 Import completed successfully!")
            except Exception as e:
                print(f"❌ Error processing file: {e}")
                messages.error(request, f"❌ Error processing file: {e}")

        return redirect('livraisons_without_date')

    return render(request, 'listings/import_xlsx.html', {'form': form})

def submission_detail(request, submission_id):
    submission = get_object_or_404(Submission, id=submission_id)
    key = settings.GOOGLE_API_KEY
    all_menus = Menu.objects.all()
    all_users = User.objects.all()
    all_payment_modes = PaymentMode.objects.all()
    conseillers = Conseiller.objects.select_related('user').all()
    # Récupérer uniquement la catégorie "SANS ALCOOL"
    category_sans_alcool = Category.objects.filter(name="SANS ALCOOL").prefetch_related('product_set').first()
    category_submission = Category.objects.filter(name="SOUMISSIONS").prefetch_related('product_set').first()

    categories = []
    if category_sans_alcool:
        categories.append(category_sans_alcool)

    categoriess = []
    if category_submission:
        categoriess.append(category_submission)

    all_delivery_modes = DeliveryMode.objects.all()
    menus = Menu.objects.all()
    # Build a dict: menu.id -> list of delivery mode ids
    menus_with_modes_ids = {}
    for menu in menus:
        modes_ids = list(menu.delivery_modes.values_list('id', flat=True))
        menus_with_modes_ids[menu.id] = modes_ids

    context = {
        'submission': submission,
        'menus': menus,
        'all_users': all_users,
        'categories': categories,
        'categoriess': categoriess,
        'menus_with_modes_ids': menus_with_modes_ids,
        'all_menus' : all_menus,
        'all_delivery_modes' : all_delivery_modes,
        'conseillers': conseillers,
        'key':key,
        'all_payment_modes':all_payment_modes,
    }
    return render(request, 'listings/submission_detail.html', context)

import qrcode
import io
import base64

def generate_qr_code_base64(url):
    qr = qrcode.QRCode(
        version=1,
        box_size=10,
        border=4,
    )
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill_color='black', back_color='white')

    buf = io.BytesIO()
    img.save(buf, format='PNG')
    encoded_string = base64.b64encode(buf.getvalue()).decode()
    return f"data:image/png;base64,{encoded_string}"

# views.py
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
from .models import Product, Checklist, ChecklistItem, Category

def checklist_products(request, checklist_id):
    checklist = get_object_or_404(Checklist, pk=checklist_id)
    checklist_documents = ChecklistDocument.objects.filter(checklist=checklist)
    recup_photos = ChecklistRecupPhoto.objects.filter(checklist=checklist)
    md_photos = ChecklistMDPhoto.objects.filter(checklist=checklist)
    existing_items = list(ChecklistItem.objects.filter(checklist=checklist).values(
    'id', 'product_id', 'quantity', 'product__name', 'commentaire'  # Ajouter 'commentaire'
    ))
    products = Product.objects.all()
    product_quantities = {product.id: product.quantity for product in products}
    special_categories = ["ALCOOL FORT", "BIERES", "SANS ALCOOL", "VINS"]
    category_icons = {
    'ALCOOL FORT': 'fa-bottle-water',
    'BIERES': 'fa-beer',
    'VINS': 'fa-wine-glass',
    'SANS ALCOOL': 'fa-glass-water',
    'ÉQUIPEMENT DE BASE': 'fa-cogs',                 # example icon
    'JETABLE': 'fa-recycle',                          # example icon
    'ACCESSOIRES DE DÉCOR': 'fa-image',               # example icon
    'ÉQUIPEMENT DE BAR': 'fa-mug-hot',                # example icon
    'ÉQUIPEMENT POUR SERVICE CAFÉ': 'fa-coffee',     # example icon
    'ITEMS DIVERS': 'fa-boxes-stacked',               # example icon
    'TABLE ET LINGE DE TABLE': 'fa-table',            # example icon
    'VERRERIE': 'fa-glass-cheers',                     # example icon
    'PORCELAINE ET COUTELLERIE': 'fa-utensils',        # example icon
    'ÉQUIPEMENT POUR MONTAGE CANAPÉS': 'fa-dolly-flatbed', # example icon
    'ÉQUIPEMENT DE CUISSON': 'fa-mortar-pestle',       # example icon
    'USTENSILES DE SERVICE': 'fa-hand-holding',       # example icon
    'CFCDN': 'fa-skull',                                # example icon
}

    breuvages = ChecklistItem.objects.filter(checklist=checklist, product__category__name__in=["ALCOOL FORT", "SANS ALCOOL", "VINS",  "BIERES"], quantity__gt=0)
    # Create the formset (always include an extra form)
    document_formset = ChecklistDocumentFormSet(
        request.POST or None,
        request.FILES or None,
        queryset=checklist_documents,  # Existing docs
        prefix='documents'
    )
    if not document_formset.forms:
        document_formset = ChecklistDocumentFormSet(queryset=ChecklistDocument.objects.none(), prefix='documents')

    if request.method == 'POST' and document_formset.is_valid():
        for form in document_formset:
            if form.cleaned_data.get('DELETE', False):
                form.instance.delete()
            else:
                document_instance = form.save(commit=False)
                document_instance.checklist = checklist
                document_instance.uploaded_by = request.user
                document_instance.save()

        # 🔹 After saving, reinitialize the formset to include at least one blank form
        return redirect('checklist-detail', checklist_id=checklist.id)


    # Checklist form instance
    formbis = ChecklistForm(request.POST or None, instance=checklist, prefix='checklist_form')
    commentairemd_form = CommentairemdForm(request.POST or None, instance=checklist, prefix='commentaire_form')
    product_form = ProductsForm
    forminfo = ChecklistInfosForm(request.POST or None, instance=checklist, prefix='checklistinfos_form')

    if request.method == 'POST':
        # Handle checklistinfos_form submission
        if 'submit_checklistinfos' in request.POST:
            forminfo = ChecklistInfosForm(request.POST, instance=checklist, prefix='checklistinfos_form')
            if forminfo.is_valid():
                forminfo.save()
                messages.success(request, 'Les informations ont été ajoutées avec succès.')
                return redirect('checklist-detail', checklist_id=checklist_id)
        # Handle commentaire_form submission separately
        elif 'commentaire_form' in request.POST:
            commentaire_form = CommentaireForm(request.POST, instance=checklist)
            if commentaire_form.is_valid():
                commentaire_form.save()
                messages.success(request, 'Commentaire ajouté avec succès.')
                return redirect('checklist-detail', checklist_id=checklist_id)

        # Handle checklist form submission
        elif 'checklist_form-name' in request.POST and formbis.is_valid():
            if checklist.statusro == 'verifié':
                    checklist.statusro = 'modifié'  # Update the status to 'modifié'
            formbis.save()
            checklist.save()
            return HttpResponseRedirect(reverse('checklist-detail', args=[checklist_id]))
    categories = Category.objects.all().order_by('name')
    products_by_category = {}
    for cat in categories:
        products_by_category[cat.name] = Product.objects.filter(category=cat)
    checklist_itemss = ChecklistItem.objects.select_related('product').filter(checklist=checklist, quantity__gt=0).prefetch_related('product__category')
    checklist_items_by_category = {}

    for item in checklist_itemss:
        product = item.product
        for category in product.category.all():  # Iterate over each category of the product
            if category.name not in checklist_items_by_category:
                checklist_items_by_category[category.name] = []
            checklist_items_by_category[category.name].append({
                'product': product,
                'quantity': item.quantity,
                'status': item.status,
                'commentaire': item.commentaire,
                'com': item.com,
                'is_completed': item.is_completed,
                'consumed_quantity': item.consumed_quantity,
                'unconsumed_quantity': item.unconsumed_quantity,
            })
    context = {
        'checklist': checklist,
        'categories': categories,
        'existing_items': existing_items,
        'special_categories': special_categories,
        'category_icons': category_icons,
        'products_by_category': products_by_category,
        'formbis': formbis,
        'product_quantities': product_quantities,
        'document_formset': document_formset,
        'commentairemd_form': commentairemd_form,
        'checklist_documents': checklist_documents,
        'recup_photos': recup_photos,
        'md_photos': md_photos,
        'product_form': product_form,
        'forminfo': forminfo,
        'breuvages': breuvages,
        'checklist_items_by_category': checklist_items_by_category,
    }
    return render(request, 'listings/checklist_detail.html', context)

from django.db import transaction


@csrf_exempt
def update_checklist_item_comment(request, item_id):
    if request.method == 'POST':
        data = json.loads(request.body)
        commentaire = data.get('commentaire', '')
        
        item = get_object_or_404(ChecklistItem, pk=item_id)
        item.commentaire = commentaire
        item.save()
        
        return JsonResponse({'status': 'ok'})
    return JsonResponse({'status': 'error'})

from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from .models import Category, Product

def get_products_by_category(request):
    category_name = request.GET.get('category', '')
    search_term = request.GET.get('search', '')
    print(f"Category: {category_name}, Search: {search_term}")  # pour déboguer
    try:
        category = Category.objects.get(name=category_name)
    except Category.DoesNotExist:
        return JsonResponse({'products': []})
    products_qs = Product.objects.filter(category=category)
    if search_term:
        products_qs = products_qs.filter(name__icontains=search_term)
    products = list(products_qs.values('id', 'name'))
    return JsonResponse({'products': products})

from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
import json

@csrf_exempt
def update_checklist_item_quantity(request, item_id):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            new_quantity = int(data.get('quantity', 0))
            
            # Récupérer l'item spécifique
            item = ChecklistItem.objects.get(pk=item_id)
            
            # Sauvegarder l'ancienne quantité
            old_quantity = item.quantity
            old_status = item.status
            
            # Ne mettre à jour que si la quantité change vraiment
            if old_quantity != new_quantity:
                item.quantity = new_quantity
                # La méthode save() gérera automatiquement le changement de statut
                item.save()
                
                return JsonResponse({
                    'status': 'ok',
                    'item_status': item.status,
                    'old_quantity': old_quantity,
                    'new_quantity': new_quantity,
                    'status_changed': old_status != item.status
                })
            else:
                return JsonResponse({
                    'status': 'ok',
                    'message': 'No change needed',
                    'quantity': old_quantity
                })
                
        except ChecklistItem.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Item not found'})
        except (ValueError, TypeError) as e:
            return JsonResponse({'status': 'error', 'message': f'Invalid quantity: {str(e)}'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

def add_products_to_checklist(request, checklist_id):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            items = data.get('items', [])
            created_items = []
            
            for item in items:
                product_id = item.get('product_id')
                quantity = item.get('quantity', 0)
                commentaire = item.get('commentaire', '')
                
                checklist_item, created = ChecklistItem.objects.update_or_create(
                    checklist_id=checklist_id,
                    product_id=product_id,
                    defaults={
                        'quantity': quantity,
                        'commentaire': commentaire
                    }
                )
                created_items.append({
                    'id': checklist_item.id,
                    'product_id': product_id,
                    'quantity': quantity
                })
            
            # Retourner les IDs des items créés/modifiés
            if created_items:
                return JsonResponse({
                    'status': 'success',
                    'items': created_items,
                    'item_id': created_items[0]['id'] if len(created_items) == 1 else None
                })
            
            return JsonResponse({'status': 'success'})
            
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request'})

from django.views.decorators.csrf import csrf_exempt
from django.db import IntegrityError
import json
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required


@login_required
@csrf_exempt
def create_journee(request):
    if request.method == "POST" and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        try:
            data = json.loads(request.body)
            print(f"Received data: {data}")  # Log the received data

            nom = data.get('name')
            date = data.get('date')

            if not nom or not date:
                return JsonResponse({"success": False, "error": "Missing required fields"}, status=400)

            # TODO: Add date validation here

            try:
                journee = Journee.objects.create(nom=nom, date=date)
                return JsonResponse({'success': True, 'message': 'Journée créée avec succès.'})
            except IntegrityError:
                return JsonResponse({'success': False, 'error': 'Une journée existe déjà pour cette date.'}, status=400)
            except Exception as e:
                return JsonResponse({'success': False, 'error': str(e)}, status=500)

        except json.JSONDecodeError:
            return JsonResponse({"success": False, "error": "Invalid JSON format"}, status=400)
    else:
        return JsonResponse({"success": False, "error": "Invalid request method"}, status=405)


from django.shortcuts import render, get_list_or_404
from .models import Livraison
@login_required
def livraisons_without_date(request):
    livraisons = Livraison.objects.filter(date__isnull=True)  # Fetch only those without a date
    journees = Journee.objects.all()

    return render(request, 'listings/livraisons_without_date.html', {'livraisons': livraisons, 'journees':journees,})

from django.shortcuts import get_object_or_404, redirect
from .models import Livraison
from django.contrib import messages
@login_required
def delete_livraison(request, livraison_id):
    livraison = get_object_or_404(Livraison, id=livraison_id)
    livraison.delete()
    messages.success(request, "🚀 Livraison deleted successfully!")
    return redirect('livraisons_without_date')  # Redirect back to the list


from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from .models import Livraison, Journee

from django.shortcuts import redirect, get_object_or_404
from django.contrib import messages
from listings.models import Livraison, Journee
@login_required
def bulk_edit_livraisons(request):
    if request.method == "POST":
        selected_ids = request.POST.getlist('selected_ids')  # Get selected Livraison IDs
        bulk_date = request.POST.get('bulk_date')  # Get bulk date
        bulk_journee = request.POST.get('bulk_journee')  # Get bulk journee ID
        action = request.POST.get("action")  # Determine action (edit or delete)

        if not selected_ids:
            messages.error(request, "❌ No livraisons selected.")
            return redirect('livraisons_without_date')

        if action == "delete":
            # 🚮 Delete selected livraisons
            Livraison.objects.filter(id__in=selected_ids).delete()
            messages.success(request, f"🗑️ {len(selected_ids)} livraisons deleted!")
            return redirect('livraisons_without_date')

        elif action == "edit":
            # 🛠 Edit selected livraisons
            for livraison_id in selected_ids:
                livraison = get_object_or_404(Livraison, id=livraison_id)

                # Update individual fields
                livraison.nom = request.POST.get(f"nom_{livraison_id}", livraison.nom)
                livraison.convives = request.POST.get(f"convives_{livraison_id}", livraison.convives)


                # Apply bulk date to both `date` and `date_livraison`
                if bulk_date:
                    livraison.date = bulk_date
                    livraison.date_livraison = bulk_date

                # Apply bulk journee
                if bulk_journee:
                    livraison.journee = get_object_or_404(Journee, id=bulk_journee)

                livraison.save()

            messages.success(request, f"✅ {len(selected_ids)} livraisons updated!")

        return redirect('livraisons_without_date')

    return redirect('livraisons_without_date')

@login_required
def edit_task_form(request, task_id):
    task = get_object_or_404(Livraison, id=task_id)

    if request.method == 'POST':
        form = LivraisonDragForm(request.POST, instance=task)
        if form.is_valid():
            form.save()
            return JsonResponse({'success': True})  # Return success response
        else:
            # Return form errors
            return JsonResponse({'success': False, 'errors': form.errors})

    else:
        form = LivraisonDragForm(instance=task)
        form_html = render_to_string('listings/partials/livraison_form.html', {'form': form})
        return JsonResponse({'form_html': form_html})


@login_required
def user_login(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)

            # Check if user needs to change their password
            try:
                user_profile = user.userprofile
                if user_profile.force_password_change:
                    return redirect(reverse('password_change'))  # Redirect to the password change view
            except UserProfile.DoesNotExist:
                pass  # Handle case if the user doesn't have a profile

            return redirect('home')  # Redirect to the home page or dashboard
    return render(request, 'listings/login.html')


# views.py
from django.contrib.auth.views import PasswordChangeView
from listings.models import UserProfile
@login_required
class CustomPasswordChangeView(PasswordChangeView):
    def form_valid(self, form):
        response = super().form_valid(form)

        # After password change, reset the 'force_password_change' flag
        user_profile = self.request.user.userprofile
        user_profile.force_password_change = False
        user_profile.save()

        return response


@login_required
def create_ordercuisine(request):
    query = request.GET.get('q')  # Get the search query from the request
    if query:
        # Filter items by name or category name (ForeignKey lookup)
        items = ItemCuisine.objects.filter(
            Q(name__icontains=query) | Q(category__name__icontains=query)
        )
    else:
        items = ItemCuisine.objects.all()  # Get all items if no search query

    if request.method == 'POST':
        for item in items:
            # Get the quantity from the form for each item
            ordered_quantity = request.POST.get(f'quantity-{item.id}')
            if ordered_quantity and int(ordered_quantity) > 0:
                # Create or update the order
                order, created = OrderCuisine.objects.get_or_create(
                    user=request.user,
                    item=item,
                    defaults={'ordered_quantity': ordered_quantity}
                )
                if not created:
                    order.ordered_quantity += int(ordered_quantity)
                    order.save()

        return redirect('order_list_cuisine')
    if not request.user.is_superuser:
        return redirect('unauthorized')

    return render(request, 'listings/create_ordercuisine.html', {'items': items})

@login_required
def user_order_list(request):
    orders = OrderCuisine.objects.filter(user=request.user)
    # Paginate the orders list
    paginator = Paginator(orders, 15)  # Show 15 orders per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    if not request.user.is_superuser:
        return redirect('unauthorized')
    return render(request, 'listings/user_order_list.html', {'orders': page_obj})

@login_required
def update_order_cuisine(request, order_id):
    order = get_object_or_404(OrderCuisine, id=order_id, user=request.user)

    if request.method == 'POST':
        ordered_quantity = request.POST.get('ordered_quantity')
        if ordered_quantity:
            order.ordered_quantity = int(ordered_quantity)
            order.save()
    if not request.user.is_superuser:
        return redirect('unauthorized')
        return redirect('user_order_list')

@login_required
def delete_order_cuisine(request, order_id):
    order = get_object_or_404(OrderCuisine, id=order_id, user=request.user)

    # Prevent deletion if order is marked as done or delivered
    if order.is_done or order.is_delivered:
        # Optionally, you can add a message here to notify the user
        return redirect('user_order_list')

    order.delete()
    if not request.user.is_superuser:
        return redirect('unauthorized')
    return redirect('user_order_list')

@login_required
def order_listcuisine(request):
    search_date = request.GET.get('search_date', None)

    if search_date:
        # Convert string to date
        search_date = datetime.strptime(search_date, '%Y-%m-%d').date()
        orders = OrderCuisine.objects.filter(user=request.user, date=search_date)
    else:
        orders = OrderCuisine.objects.filter(user=request.user)

    # Paginate the orders list
    paginator = Paginator(orders, 15)  # Show 15 orders per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    if not request.user.is_superuser:
        return redirect('unauthorized')

    return render(request, 'listings/order_listcuisine.html', {
        'orders': page_obj,
        'search_date': search_date
    })

from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.contrib import messages

@login_required
def mark_order_donecuisine(request, order_id):
    order = OrderCuisine.objects.get(id=order_id)
    order.is_done = True
    order.save()
    if not request.user.is_superuser:
        return redirect('unauthorized')
    return redirect('order_list_cuisine')
@login_required
def mark_order_deliveredcuisine(request, order_id):
    order = OrderCuisine.objects.get(id=order_id)
    order.is_delivered = True
    order.save()
    if not request.user.is_superuser:
        return redirect('unauthorized')
    return redirect('order_list_cuisine')


from django.http import JsonResponse
import googlemaps
from django.conf import settings
from .models import Livraison
@login_required
def geocode_all_livraisons(request):
    if request.method == 'GET':  # Ensure the request method is GET
        today = timezone.localdate()
        livraisons = Livraison.objects.filter(lat__isnull=True, lng__isnull=True, place_id__isnull=True,date__gte=today)

        if livraisons.exists():
            try:
                # Initialize the Google Maps Client with the API Key
                gmaps = googlemaps.Client(key=settings.GOOGLE_API_KEY)

                for livraison in livraisons:
                    if livraison.adress and livraison.country and livraison.zipcode and livraison.city:
                        # Create the address string for geocoding
                        adress_string = f"{livraison.adress}, {livraison.zipcode}, {livraison.city}, {livraison.country}"

                        # Perform geocoding request to Google Maps
                        result = gmaps.geocode(adress_string)
                        if result:
                            lat = result[0].get('geometry', {}).get('location', {}).get('lat', None)
                            lng = result[0].get('geometry', {}).get('location', {}).get('lng', None)
                            place_id = result[0].get('place_id', None)

                            # Update the livraison instance
                            livraison.lat = lat
                            livraison.lng = lng
                            livraison.place_id = place_id
                            livraison.save()

                return JsonResponse({'success': True, 'message': 'Géocode reussit'})
            except Exception as e:
                return JsonResponse({'success': False, 'error': str(e)})
        else:
            return JsonResponse({'success': False, 'message': 'Aucune livraison à géocoder'})

    # If the request method is not GET
    return JsonResponse({'success': False, 'error': 'Invalid request method. Use GET.'})


import random

TASK_NAMES = ['Nettoyer machines à café', 'Nettoyer intérieur des camions',
              'Faire boites de thé + café', 'Nettoyer dock de livraison', 'Mettre essence camions']
@login_required
def create_random_task(request):
    # List of names to choose from
    livreur_names = ["Mohamed", "Samuel", "Alex", "Jef", "Zayd"]

    # Check if the number of livreurs is less than the number of tasks
    if len(livreur_names) < len(TASK_NAMES):
        return redirect('acceuilresponsables')  # Handle error: more tasks than livreurs

    # Fetch Livreur instances corresponding to names
    livreurs = []
    for name in livreur_names:
        try:
            livreur = Livreur.objects.get(nom=name)
            livreurs.append(livreur)
        except Livreur.DoesNotExist:
            return redirect('acceuilresponsables')  # Handle case where a Livreur is missing

    # Set the date for tomorrow
    tomorrow = timezone.now().date() + timezone.timedelta(days=1)

    # Shuffle the task names to ensure randomness
    random.shuffle(TASK_NAMES)

    # Assign a random task to each livreur
    for i, livreur in enumerate(livreurs):
        if i < len(TASK_NAMES):
            random_task = TASK_NAMES[i]
            # Create the task
            Tacheafaire.objects.create(
                livreur=livreur,  # Livreur instance
                nom=random_task,
                date=tomorrow
            )
    if not request.user.is_superuser:
        return redirect('unauthorized')
    return redirect('acceuilresponsables')

@login_required
@csrf_exempt
@require_POST
def create_routes(request):
    try:
        # Calculate tomorrow's date
        tomorrow = timezone.now().date() + timedelta(days=1)

        # Create 20 routes
        routes = []
        for i in range(1, 21):
            route = Route(
                nom=f"{i}",
                date=tomorrow,
                heure_depart="08h00"  # You can set this to any default value
            )
            route.save()
            routes.append(route)

        return JsonResponse({'success': True, 'routes': [route.id for route in routes]})

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
@login_required
@csrf_exempt
@require_POST
def create_routesn(request):
    try:
        # Calculate tomorrow's date
        tomorrow = timezone.now().date() + timedelta(days=2)

        # Create 20 routes
        routes = []
        for i in range(1, 21):
            route = Route(
                nom=f"{i}",
                date=tomorrow,
                heure_depart="08h00"  # You can set this to any default value
            )
            route.save()
            routes.append(route)

        return JsonResponse({'success': True, 'routes': [route.id for route in routes]})

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
@login_required
@csrf_exempt
@require_POST
def create_routesnn(request):
    try:
        # Calculate tomorrow's date
        tomorrow = timezone.now().date() + timedelta(days=3)

        # Create 20 routes
        routes = []
        for i in range(1, 21):
            route = Route(
                nom=f"{i}",
                date=tomorrow,
                heure_depart="08h00"  # You can set this to any default value
            )
            route.save()
            routes.append(route)

        return JsonResponse({'success': True, 'routes': [route.id for route in routes]})

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
class ChecklistItemDeleteAjaxView(View):
    def post(self, request, *args, **kwargs):
        item_id = self.kwargs.get('pk')
        try:
            item = ChecklistItem.objects.get(pk=item_id)
            item.delete()
            response_data = {'status': 'success', 'message': 'Objet bien supprimé :)'}
        except ChecklistItem.DoesNotExist:
            response_data = {'status': 'error', 'message': 'Objet non trouvé'}
        return JsonResponse(response_data)

from django.http import JsonResponse
from django.views.decorators.http import require_GET
from django.utils.dateparse import parse_date
from django.db.models import F
@login_required
@require_GET
def get_checklist_items_for_date(request):
    date_str = request.GET.get('date')

    if not date_str:
        return JsonResponse({'error': 'No date provided'}, status=400)

    parsed_date = parse_date(date_str)  # Convert string to date
    if not parsed_date:
        return JsonResponse({'error': 'Invalid date format'}, status=400)

    checklist_items = ChecklistItem.objects.filter(
        checklist__date=parsed_date,
        checklist__is_active=True,
        quantity__gt=0
    ).annotate(
        product_name=F('product__name')
    ).values('product_name', 'quantity', 'status')

    return JsonResponse({'items': list(checklist_items)}, safe=False)


from django.views.decorators.http import require_GET
from django.http import JsonResponse
from .models import ChecklistItem, Checklist

# Dans votre vue Django
def checklist_items_encours_par_date(request):
    date = request.GET.get('date')
    status_filter = request.GET.get('status_filter', '')
    
    # Convertir la date
    selected_date = datetime.strptime(date, '%Y-%m-%d').date()
    
    # Requête de base
    items = ChecklistItem.objects.filter(
        checklist__date=selected_date,
        checklist__is_active=True
    )
    
    # Appliquer le filtre de statut si fourni
    if status_filter:
        status_list = status_filter.split(',')
        items = items.filter(status__in=status_list)
    
    # Grouper par produit et calculer les totaux
    items = items.values(
        'product__name', 
        'checklist__name'
    ).annotate(
        total_quantity=Sum('quantity')
    )
    
    return JsonResponse({'items': list(items)})

def format_time_without_seconds(time_str):
    if ':' in time_str:
        parts = time_str.split(':')
        return f"{parts[0]}:{parts[1]}"
    return time_str


from django.shortcuts import render
from django.http import JsonResponse
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_GET
from django.utils import timezone
from django.core.cache import cache
from django.db.models import Sum
import calendar
from datetime import date, datetime, timedelta

@login_required
@require_GET
def voir_checklist(request):
    today = date.today()

    # Préchargement des checklists actives
    checklists = Checklist.objects.filter(is_active=True)

    encours = "en_cours"
    valide = "valide"
    refuse = "refuse"
    tomorrow = timezone.now().date() + timedelta(days=1)

    # Checklists du jour
    checklists_of_the_day = Checklist.objects.filter(date=today, is_active=True)

    # Années pour le filtre
    current_year = today.year
    years = [year for year in range(current_year - 5, current_year + 1)]

    # Livraison
    livraisons = Livraison.objects.filter(recuperation=False)
    for livraison in livraisons:
        livraison.formatted_heure_livraison = format_time_without_seconds(livraison.heure_livraison)

    # Séléction du mois
    selected_day = int(request.GET.get('day', 1))
    selected_month = int(request.GET.get('month', today.month))
    _, num_days_in_month = calendar.monthrange(current_year, selected_month)
    days_in_month = [day for day in range(1, num_days_in_month + 1)]

    # Listes de noms de mois en français
    french_months = [
        "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
        "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"
    ]

    # Requêtes pour checklists du mois
    checklists = Checklist.objects.filter(date__month=selected_month)

    # Logs
    change_logs = ChecklistItemChangeLog.objects.all().order_by('-timestamp')
    change_logs_checklist = ChecklistChangeLog.objects.all().order_by('-timestamp')

    # Date sélectionnée
    selected_date_str = request.GET.get('selected_datee')
    if selected_date_str:
        selected_datee = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
    else:
        selected_datee = tomorrow

    # --- Optimisation de la requête checklist_item_totals avec cache ---
    cache_key = f'checklist_item_totals_{selected_datee}'
    checklist_item_totals = cache.get(cache_key)

    if checklist_item_totals is None:
        # Si pas dans le cache, faire la requête
        checklist_items_qs = ChecklistItem.objects.filter(
            checklist__date=selected_datee,
            checklist__is_active=True
        ).select_related('product')  # Précharger 'product' pour éviter N+1

        checklist_item_totals = (
            checklist_items_qs
            .values('product__name', 'product__category')
            .annotate(total_quantity=Sum('quantity'))
            .filter(total_quantity__gt=0)
            .order_by('product__category', 'product__name')
        )

        # Stocker dans le cache (15 min par exemple)
        cache.set(cache_key, list(checklist_item_totals), timeout=900)
    else:
        # Si en cache, convertir en liste
        checklist_item_totals = checklist_item_totals

    # Préparer les données JSON pour la requête AJAX
    data = {
        'checklist_item_totals': list(checklist_item_totals),
        'selected_datee': selected_datee,
    }

    # Si requête AJAX, renvoyer JSON
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse(data)

    # Générez les QR Codes
    for checklist in checklists:
        url = request.build_absolute_uri(reverse('checklist_en_cours', args=[checklist.id]))
        checklist.qr_code_uri = generate_qr_code_base64(url)
        checklist.formatted_heure_livraison = format_time_without_seconds(checklist.heure_livraison)

    # Contexte pour le rendu
    context = {
        'checklists': checklists,
        'encours': encours,
        'valide': valide,
        'refuse': refuse,
        'today': today,
        'change_logs': change_logs,
        'change_logs_checklist': change_logs_checklist,
                'months': list(range(1, 13)),  # Mois sous forme de liste
        'selected_month': selected_month,
        'days': days_in_month,
        'checklists_of_the_day': checklists_of_the_day,
        'months': [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12],
        'french_months': french_months,
        'selected_day': selected_day,
        'selected_date': date(current_year, selected_month, selected_day).strftime('%d %B %Y'),
        'years': years,
        'checklist_item_totals': checklist_item_totals,
        'selected_year': current_year,
        'livraisons': livraisons,
    }
    return render(request, 'listings/voir-checklist.html', context)

@login_required
def view_items_by_category(request, category):
    # Fetch checklist items for the specified category
    checklist_items = ChecklistItem.objects.filter(product__category=category)

    context = {
        'checklist_items': checklist_items,
        'category': category,
    }

    return render(request, 'listings/items_by_category.html', context)

@login_required
def delete_route(request, route_id):
    if request.method == 'DELETE':
        route = get_object_or_404(Route, id=route_id)

        # Update the status of tasks associated with this route
        livraisons = Livraison.objects.filter(statut=route)  # Assuming a ForeignKey relationship
        livraisons.update(statut_id=21)  # Change 'statut_id' to 21, according to your model field

        # After updating the tasks, delete the route
        route.delete()

        return JsonResponse({'success': True})

    return JsonResponse({'success': False}, status=400)



from django.shortcuts import get_object_or_404, redirect, render
from django.http import JsonResponse
from .models import Checklist, ChecklistItem, QuantityChangeLog

@login_required
def checklistvoir_detail(request, checklist_id):
    products = Product.objects.all()
    checklist = get_object_or_404(Checklist, pk=checklist_id)

    # Get all checklist items
    all_items = (
        ChecklistItem.objects
        .filter(checklist=checklist, quantity__gt=0)
        .prefetch_related('product__category')
        .order_by('product__category__name')
    )
    
    # Group unique checklist items by product
    checklist_items_map = {}
    for item in all_items:
        if item.product_id not in checklist_items_map:
            checklist_items_map[item.product_id] = item  # Keep the first occurrence

    checklist_items_unique = list(checklist_items_map.values())  # Get the unique checklist items

    if request.method == 'POST':
        # Handle note update
        note = request.POST.get('note')
        if note is not None:
            checklist.notechecklist = note
            checklist.save()
            return redirect('checklistvoir-detail', checklist_id=checklist.id)

        # Handle product/quantity update
        action = request.POST.get('action')
        if action == 'update_item':
            item_id = request.POST.get('item_id')
            product_id = request.POST.get('product_id')
            quantity = request.POST.get('quantity')
            
            if item_id and product_id and quantity:
                checklist_item = get_object_or_404(ChecklistItem, id=item_id, checklist=checklist)
                new_product = get_object_or_404(Product, id=product_id)
                
                # Update the item
                checklist_item.product = new_product
                checklist_item.quantity = int(quantity)
                checklist_item.save()
                
                messages.success(request, 'Produit mis à jour avec succès!')
                return redirect('checklistvoir-detail', checklist_id=checklist.id)

        # Handle status update - SIMPLIFIED VERSION
        item_id = request.POST.get('item_id')
        status = request.POST.get('status')
        
        if item_id and status:
            checklist_item = get_object_or_404(ChecklistItem, id=item_id, checklist=checklist)
            
            # Simply update the status without any inventory logic
            checklist_item.status = status
            checklist_item.save()
            
            # Update the parent checklist status
            checklist.update_status()

            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': 'Mise à jour réussie!'})

            return redirect('checklistvoir-detail', checklist_id=checklist.id)

    # Generate URL for the checklist "en_cours" page
    checklist_url = request.build_absolute_uri(reverse('checklist_en_cours', args=[checklist.id]))
    # Generate QR code image
    qr_code_data_uri = generate_qr_code_base64(checklist_url)

    # Sort items by status for display
    items_encours = [item for item in checklist_items_unique if item.status == 'en_cours']
    items_autres = [item for item in checklist_items_unique if item.status not in ('en_cours', 'valide', 'refuse')]
    items_valides_refuses = [item for item in checklist_items_unique if item.status in ('valide', 'refuse')]
    
    # Combine all items in the desired order
    all_items_ordered = items_encours + items_autres + items_valides_refuses

    context = {
        'checklist': checklist,
        'all_items': all_items_ordered,  # Use this single list in template
        'products': products,
        'qr_code_data_uri': qr_code_data_uri,
    }
    
    return render(request, 'listings/checklistevoir_detail.html', context)
@login_required
def checklist_en_cours_view(request, checklist_id):
    checklist = get_object_or_404(Checklist, pk=checklist_id)
    items_en_cours = ChecklistItem.objects.filter(checklist=checklist, status='en_cours', quantity__gt=0)
    sac_de_glace_items = ChecklistItem.objects.filter(checklist=checklist, product__name="SAC DE GLACE")
    context = {
        'checklist': checklist,
        'items_en_cours': items_en_cours,
        'sac_de_glace_items': sac_de_glace_items,
    }
    return render(request, 'listings/checklist_en_cours.html', context)

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt

@csrf_exempt  # Pour simplifier, sinon gérer le CSRF côté fetch
def remettre_en_stock(request):
    if request.method == 'POST':
        product_id = request.POST.get('product_id')
        quantity = int(request.POST.get('quantity', 0))
        try:
            product = Product.objects.get(id=product_id)
            product.adjust_quantity(quantity)  # Méthode à définir dans votre modèle
            product.save()
            return JsonResponse({'success': True})
        except Product.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Produit non trouvé'})
    return JsonResponse({'success': False, 'error': 'Methode non autorisée'})


@login_required
def product_detail(request, item_id):
    item = get_object_or_404(ItemInv, pk=item_id)

    context = {
        'item': item,
    }
    return render(request, 'listings/product_detail.html', context)


from django.shortcuts import get_object_or_404, redirect
from django.http import JsonResponse
from .models import ChecklistItem, Product
@login_required
def validate_checklist_item(request, checklist_item_id):
    checklist_item = get_object_or_404(ChecklistItem, id=checklist_item_id)
    product = checklist_item.product

    if checklist_item.status == 'pending':
        # Calculate the difference between current and previous quantity
        quantity_difference = checklist_item.quantity - checklist_item.previous_quantity
    else:
        # Directly subtract the quantity
        quantity_difference = checklist_item.quantity

    # Update product quantity
    product.quantity -= quantity_difference
    product.save()

    # Mark checklist item as validated
    checklist_item.status = 'validated'
    checklist_item.previous_quantity = checklist_item.quantity  # Update previous quantity
    checklist_item.save()

    return JsonResponse({'success': True, 'new_product_quantity': product.quantity})

@login_required
def inventory(request):
    products = Product.objects.all()
    context = {
        'products': products,
    }
    return render(request, 'listings/inventory.html', context)

@login_required
def subtract_to_checklist(request, checklist_id):

    checklist_item = get_object_or_404(Checklist, pk=checklist_id)
    item = checklist_item.item
    quantity_to_subtract = checklist_item.quantity_checked

    # Ensure there are enough items in inventory to subtract
    if item.quantity >= quantity_to_subtract:
        # Update item quantity in inventory
        item.quantity -= quantity_to_subtract
        item.save()

        # Delete checklist item after subtracting
        checklist_item.delete()

    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

import logging
from django.http import JsonResponse

logger = logging.getLogger(__name__)
@login_required
def add_to_checklist(request, checklist_id):
    if request.method == 'POST':
        checklist = get_object_or_404(Checklist, pk=checklist_id)

        # Fetch product IDs, quantities, and comments
        product_ids = request.POST.getlist('product_ids[]')
        quantities = request.POST.getlist('quantities[]')
        commentaires = request.POST.getlist('commentaires[]')

        # Initialize a flag to check if any item was processed
        item_added_or_updated = False

        for product_id, quantity, commentaire in zip(product_ids, quantities, commentaires):
            try:
                product = Product.objects.get(pk=product_id)
                quantity = int(quantity)  # Ensure quantity is an integer


                # Check the condition for quantity 0 with a comment
                if quantity >= 0:
                    # Get or create the ChecklistItem
                    checklist_item, created = ChecklistItem.objects.get_or_create(
                        checklist=checklist,
                        product=product,
                        defaults={'quantity': quantity, 'commentaire': commentaire}
                    )

                    if not created:
                        # This item already exists, so we are modifying it
                        checklist_item._changed_by = request.user
                        checklist_item.quantity = quantity

                        # Only update commentaire if it's changed
                        if commentaire.strip() and commentaire != checklist_item.commentaire:
                            checklist_item.commentaire = commentaire

                        checklist_item.save()
                    else:
                        # New item added
                        item_added_or_updated = True  # Mark that at least one item was added

            except Product.DoesNotExist:
                return JsonResponse({'success': False, 'message': f'Product ID {product_id} not found.'})

        # If any item was added or updated, change status to 'modifié'
        if item_added_or_updated:
            checklist.statusro = 'modifié'  # Set status to 'modifié'
            checklist.save()

    return redirect(request.META.get('HTTP_REFERER', '/'))

from collections import defaultdict


@login_required
def checklist_detail(request, checklist_id):
    checklist = get_object_or_404(Checklist, pk=checklist_id)
    checklist_items = ChecklistItem.objects.filter(checklist=checklist, quantity__gt=0).select_related('product')
    checklist_products = checklist_items.values_list('product__id', flat=True)
    breuvages = ChecklistItem.objects.filter(checklist=checklist, product__category__name__in=["ALCOOL FORT", "SANS ALCOOL", "VINS",  "BIERES"], quantity__gt=0)
    checklist_documents = ChecklistDocument.objects.filter(checklist=checklist)
    recup_photos = ChecklistRecupPhoto.objects.filter(checklist=checklist)
    md_photos = ChecklistMDPhoto.objects.filter(checklist=checklist)
    products = Product.objects.prefetch_related('category').all()  # Prefetch categories to avoid additional queries
    bieres_category = Category.objects.get(name="BIERES")
    vins_category = Category.objects.get(name="VINS")
    alcool_category = Category.objects.get(name="ALCOOL FORT")
    sansalcool_category = Category.objects.get(name="SANS ALCOOL")
    base_category = Category.objects.get(name="ÉQUIPEMENT DE BASE")
    jetable_category = Category.objects.get(name="JETABLE")
    decor_category = Category.objects.get(name="ACCESSOIRES DE DÉCOR")
    bar_category = Category.objects.get(name="ÉQUIPEMENT DE BAR")
    cafe_category = Category.objects.get(name="ÉQUIPEMENT POUR SERVICE CAFÉ")
    table_category = Category.objects.get(name="TABLE ET LINGE DE TABLE")
    verre_category = Category.objects.get(name="VERRERIE")
    porcelaine_category = Category.objects.get(name="PORCELAINE ET COUTELLERIE")
    canape_category = Category.objects.get(name="ÉQUIPEMENT POUR MONTAGE CANAPÉS")
    cuisson_category = Category.objects.get(name="ÉQUIPEMENT DE CUISSON")
    service_category = Category.objects.get(name="USTENSILES DE SERVICE")
    divers_category = Category.objects.get(name="ITEMS DIVERS")
    cfcdn_category = Category.objects.get(name="CFCDN")


    for product in products:
        product.is_bieres = bieres_category in product.category.all()
        product.is_vins = vins_category in product.category.all()
        product.is_alcool = alcool_category in product.category.all()
        product.is_sansalcool = sansalcool_category in product.category.all()
        product.is_base = base_category in product.category.all()
        product.is_jetable = jetable_category in product.category.all()
        product.is_decor = decor_category in product.category.all()
        product.is_bar = bar_category in product.category.all()
        product.is_cafe = cafe_category in product.category.all()
        product.is_table = table_category in product.category.all()
        product.is_verre = verre_category in product.category.all()
        product.is_porcelaine = porcelaine_category in product.category.all()
        product.is_canape = canape_category in product.category.all()
        product.is_cuisson = cuisson_category in product.category.all()
        product.is_service = service_category in product.category.all()
        product.is_divers = divers_category in product.category.all()
        product.is_cfcdn = cfcdn_category in product.category.all()

    query = request.GET.get('query')
    checklist_item_quantities = {
        item.product_id: item.quantity for item in checklist_items
    }
    checklist_itemss = ChecklistItem.objects.select_related('product').filter(checklist=checklist, quantity__gt=0).prefetch_related('product__category')
    checklist_item_comments = {item.product_id: item.commentaire for item in checklist_items}



    checklist_items_by_category = {}

    for item in checklist_itemss:
        product = item.product
        for category in product.category.all():  # Iterate over each category of the product
            if category.name not in checklist_items_by_category:
                checklist_items_by_category[category.name] = []
            checklist_items_by_category[category.name].append({
                'product': product,
                'quantity': item.quantity,
                'status': item.status,
                'commentaire': item.commentaire,
                'com': item.com,
                'is_completed': item.is_completed,
                'consumed_quantity': item.consumed_quantity,
                'unconsumed_quantity': item.unconsumed_quantity,
            })

   # Get unique product categories

    product_categories = Category.objects.filter(product__isnull=False).distinct()



    user_groups = request.user.groups.values_list('name', flat=True)
    checklist_documents = ChecklistDocument.objects.filter(checklist=checklist)

    # Create the formset (always include an extra form)
    document_formset = ChecklistDocumentFormSet(
        request.POST or None,
        request.FILES or None,
        queryset=checklist_documents,  # Existing docs
        prefix='documents'
    )
    if not document_formset.forms:
        document_formset = ChecklistDocumentFormSet(queryset=ChecklistDocument.objects.none(), prefix='documents')

    if request.method == 'POST' and document_formset.is_valid():
        for form in document_formset:
            if form.cleaned_data.get('DELETE', False):
                form.instance.delete()
            else:
                document_instance = form.save(commit=False)
                document_instance.checklist = checklist
                document_instance.uploaded_by = request.user
                document_instance.save()

        # 🔹 After saving, reinitialize the formset to include at least one blank form
        return redirect('checklist-detail', checklist_id=checklist.id)


    # Checklist form instance
    formbis = ChecklistForm(request.POST or None, instance=checklist, prefix='checklist_form')
    commentairemd_form = CommentairemdForm(request.POST or None, instance=checklist, prefix='commentaire_form')
    product_form = ProductsForm
    forminfo = ChecklistInfosForm(request.POST or None, instance=checklist, prefix='checklistinfos_form')

    if request.method == 'POST':
        # Handle checklistinfos_form submission
        if 'submit_checklistinfos' in request.POST:
            forminfo = ChecklistInfosForm(request.POST, instance=checklist, prefix='checklistinfos_form')
            if forminfo.is_valid():
                forminfo.save()
                messages.success(request, 'Les informations ont été ajoutées avec succès.')
                return redirect('checklist-detail', checklist_id=checklist_id)

        # Handle commentaire_form submission separately
        elif 'commentaire_form' in request.POST:
            commentaire_form = CommentaireForm(request.POST, instance=checklist)
            if commentaire_form.is_valid():
                commentaire_form.save()
                messages.success(request, 'Commentaire ajouté avec succès.')
                return redirect('checklist-detail', checklist_id=checklist_id)

        # Handle product_form submission separately
        elif 'product-form' in request.POST:
            product_form = ProductsForm(request.POST, user=request.user)
            if product_form.is_valid():
                product = product_form.save()
                # Log the product creation logic, etc.
                messages.success(request, 'Nouveau produit créé avec succès.')
                return redirect('checklist-detail', checklist_id=checklist_id)


        # Handle checklist form submission
        elif 'checklist_form-name' in request.POST and formbis.is_valid():
            if checklist.statusro == 'verifié':
                    checklist.statusro = 'modifié'  # Update the status to 'modifié'
            formbis.save()
            checklist.save()
            return HttpResponseRedirect(reverse('checklist-detail', args=[checklist_id]))



    # Filter products based on the search query
    if query:
        products = products.filter(name__icontains=query)

    # Group products by categories
    products_by_category = defaultdict(list)
    for product in products:
        for category in product.category.all():
            products_by_category[category.name].append(product)

    is_cfcdn = "cfcdn" in checklist.name.lower()

    context = {
        'products_by_category': dict(products_by_category),
        'breuvages': breuvages,
        'checklist': checklist,
        'checklist_products': checklist_products,
        'checklist_items': checklist_items,
        'products': products,
        'formbis': formbis,
        'commentairemd_form': commentairemd_form,
        'document_formset': document_formset,
        'checklist_documents': checklist_documents,
        'recup_photos': recup_photos,
        'md_photos': md_photos,
        'product_form': product_form,
        'forminfo': forminfo,
        'checklist_item_quantities': checklist_item_quantities,
        'checklist_item_comments': checklist_item_comments,
        'checklist_items_by_category': checklist_items_by_category,
        'product_categories': product_categories,
        'is_admin': 'admin' in user_groups,
        'is_chefcuisine': request.user.groups.filter(name='chefcuisine').exists(),
        'is_ventes': request.user.groups.filter(name='ventes').exists(),
        'is_cfcdn': is_cfcdn,
    }
    return render(request, 'listings/checklist_detail.html', context)



@login_required
@require_POST
def delete_document(request, document_id):
    document = get_object_or_404(ChecklistDocument, id=document_id)
    # Check if user has permission to delete
    if request.user == document.uploaded_by or request.user.is_superuser:
        document.delete()
        return JsonResponse({'success': True, 'message': 'Document deleted successfully.'})
    return JsonResponse({'success': False, 'message': 'You do not have permission to delete this document.'})

@login_required
@csrf_exempt
def adjust_product_quantity(request):
    if request.method == 'POST':
        form = AdjustProductQuantityForm(request.POST)
        if form.is_valid():
            product = form.cleaned_data['product']
            quantity = form.cleaned_data['quantity']
            if product.quantity >= quantity:
                product.adjust_quantity(-quantity)
                messages.success(request, f"{quantity} {product.name} ont bien été enlevé.")
            else:
                messages.error(request, f"Stock insuffisant pour {product.name}. Disponible: {product.quantity}.")
            return redirect('adjust_product_quantity')  # Replace with your desired redirect URL
    else:
        form = AdjustProductQuantityForm()

    return render(request, 'listings/adjust_quantity.html', {'form': form})
@login_required
def edit_item(request, pk):
    item = get_object_or_404(ItemInv, pk=pk)
    if request.method == 'POST':
        form = ItemInvForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            return redirect('inventaire')
        else:
            return JsonResponse({'status': 'error', 'message': 'Form is not valid.'}, status=400)
    else:
        form = ItemInvForm(instance=item)
    return render(request, 'listings/edit_item.html', {'form': form, 'item':item})


from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.shortcuts import render, get_object_or_404
from django.contrib import messages
from django.utils import timezone

@login_required
def conseiller_dashboard(request):
    encours = "en_cours"
    valide = "valide"
    refuse = "refuse"
    envoye = "envoyé"
    conseiller_instance = get_object_or_404(Conseiller, user=request.user)

    # Fetch unread notifications for the current user
    unread_notifications = request.user.notifications.filter(is_read=False)

    # Fetch checklists
    checklists = Checklist.objects.filter(conseillere=conseiller_instance).order_by("-date")
    active_checklists = checklists.filter(is_active=True)
    inactive_checklists = checklists.filter(is_active=False)

    # Paginate checklists
    checklist_paginator = Paginator(checklists, 10)
    checklist_page_number = request.GET.get('checklist_page')
    try:
        checklist_page_obj = checklist_paginator.page(checklist_page_number)
    except PageNotAnInteger:
        checklist_page_obj = checklist_paginator.page(1)
    except EmptyPage:
        checklist_page_obj = checklist_paginator.page(checklist_paginator.num_pages)

    # Pagination for user's submissions
    submissions = Submission.objects.filter(user=request.user).order_by("-created_at")

    submissions_48h_encours = submissions.filter(
        status='en_cours',
        created_at__lt=timezone.now() - timedelta(hours=48),
        submission_type__icontains='soumission'
    )
    commande_48h_encours = submissions.filter(
        status='en_cours',
        created_at__lt=timezone.now() - timedelta(hours=48),
        submission_type__icontains='commande'
    )
    submissions_48h_envoye = submissions.filter(
        status='envoyé',
        created_at__lt=timezone.now() - timedelta(hours=48),
        submission_type__icontains='soumission'
    )
    commande_48h_envoye = submissions.filter(
        status='envoyé',
        created_at__lt=timezone.now() - timedelta(hours=48),
        submission_type__icontains='commande'
    )

    # Paginate submissions
    submission_paginator = Paginator(submissions, 10)
    submission_page_number = request.GET.get('submission_page')
    try:
        submission_page_obj = submission_paginator.page(submission_page_number)
    except PageNotAnInteger:
        submission_page_obj = submission_paginator.page(1)
    except EmptyPage:
        submission_page_obj = submission_paginator.page(submission_paginator.num_pages)

    # Initialize forms
    checklist_form = ChecklistForm(initial={'conseillere': conseiller_instance})
    product_form = ProductsForm()

    if request.method == 'POST':
        if 'create_checklist' in request.POST:
            checklist_form = ChecklistForm(request.POST)
            if checklist_form.is_valid():
                checklist = checklist_form.save(commit=False)
                checklist.conseillere = conseiller_instance
                checklist.save()
                messages.success(request, 'Nouvelle checklist créée.')
                return redirect('conseiller_dashboard')
            else:
                messages.error(request, 'Veuillez corriger les erreurs dans le formulaire de création.')

        elif 'toggle_checklist' in request.POST:
            checklist_id = request.POST.get('checklist_id')
            try:
                checklist = Checklist.objects.get(id=checklist_id, conseillere=conseiller_instance)
                checklist.is_active = not checklist.is_active
                checklist.save()
                messages.success(request, 'Checklist activée.' if checklist.is_active else 'Checklist désactivée.')
            except Checklist.DoesNotExist:
                messages.error(request, 'Checklist non trouvée.')

        elif 'product-form' in request.POST:
            product_form = ProductsForm(request.POST)
            if product_form.is_valid():
                product = product_form.save(commit=False)
                product._current_user = request.user
                product.save()
                messages.success(request, 'Nouveau produit créé avec succès.')
                return redirect('creerchecklist')

    # Calculate counts for each status
    total_submissions = submissions.count()
    count_encours = submissions.filter(status='en_cours').count()
    count_valide = submissions.filter(status='valide').count()
    count_refuse = submissions.filter(status='refuse').count()
    count_envoye = submissions.filter(status='envoyé').count()

    context = {
        'checklist_form': checklist_form,
        'conseiller_instance': conseiller_instance,
        'active_checklists': active_checklists,
        'inactive_checklists': inactive_checklists,
        'encours': encours,
        'valide': valide,
        'total_submissions': total_submissions,
        'count_encours': count_encours,
        'count_valide': count_valide,
        'count_refuse': count_refuse,
        'count_envoye': count_envoye,
        'product_form': product_form,
        'refuse': refuse,
        'checklists': checklists,
        'checklist_page_obj': checklist_page_obj,
        'submission_page_obj': submission_page_obj,
        'today': timezone.now().date(),
        'submissions_48h_envoye': submissions_48h_envoye,
        'submissions_48h_encours': submissions_48h_encours,
        'commande_48h_encours': commande_48h_encours,
        'commande_48h_envoye': commande_48h_envoye,
        'envoye': envoye,
        # Ajouter les notifications non lues pour la page
        'unread_notifications': unread_notifications,
    }

    return render(request, 'listings/conseiller_dashboard.html', context)


from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json

@csrf_exempt  # Use only for testing; implement CSRF protection in production
def update_submission_status_dashboard(request, submission_id):
    if request.method == 'POST':
        try:
            submission = Submission.objects.get(id=submission_id)
            data = json.loads(request.body)  # Parse the JSON payload

            # Get the new status from the request
            status = data.get('status')
            submission.status = status

            # If status is 'refuse', save the comment as well
            if status == 'refuse':
                comment = data.get('comment', '')
                submission.refusal_comment = comment  # Assign the comment to the model

            submission.save()

            return JsonResponse({'success': True})

        except Submission.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Submission not found.'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})  # General error handling

    return JsonResponse({'success': False, 'error': 'Invalid request.'})


from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
@login_required
@csrf_exempt  # Only use this if you're handling CSRF tokens manually; otherwise, set properly in your templates
def update_checklist_status(request, checklist_id):
    if request.method == 'POST':
        try:
            checklist = Checklist.objects.get(id=checklist_id)
            checklist.statusro = 'verifié'  # Update status to 'verifié'
            checklist.save()
            return JsonResponse({'status': 'success'})
        except Checklist.DoesNotExist:
            return JsonResponse({'error': 'Checklist not found'}, status=404)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=400)
@login_required
@require_POST
def update_checklist_item(request):
    item_id = request.POST.get('item_id')
    product_id = request.POST.get('product_id')

    # Safely convert quantity to integer
    quantity = int(request.POST.get('quantity'))  # Convert to integer

    # Get the checklist item
    checklist_item = get_object_or_404(ChecklistItem, id=item_id)

    # Store original quantity for logging or processing
    original_quantity = checklist_item.quantity

    # Update the product and quantity
    checklist_item.product_id = product_id  # Assuming product ID is passed
    checklist_item.quantity = quantity

    # Save the updated item
    checklist_item.save()

    # Here you may want to log the quantity change if needed
    QuantityChangeLog.objects.create(
        checklist_item=checklist_item,
        previous_quantity=original_quantity,
        new_quantity=quantity,
        quantity_change=quantity - original_quantity,  # Ensure this operates on integers
        changed_by=request.user
    )

    return JsonResponse({'success': True, 'message': 'Checklist item updated successfully.'})

from django.db.models import Max


@login_required
def creerchecklist(request):
    # Retrieve all checklists and define status labels
    checklists = Checklist.objects.all().order_by('date')
    encours = "en_cours"
    valide = "valide"
    refuse = "refuse"
    today = date.today()

    current_year = date.today().year
    years = [year for year in range(current_year - 5, current_year + 1)]

    selected_day = int(request.GET.get('day', 1))  # Default to the first day of the month if none selected
    current_year = date.today().year
    months = [(month, calendar.month_name[month]) for month in range(1, 13)]
    selected_month = int(request.GET.get('month', today.month))
    # Get the number of days in the selected month
    _, num_days_in_month = calendar.monthrange(current_year, selected_month)
    days_in_month = [day for day in range(1, num_days_in_month + 1)]  # Adjust days in month

    # Get the starting day of the month (0 = Monday, 1 = Tuesday, ..., 6 = Sunday)
    first_day_of_month = date(current_year, selected_month, 1).weekday()  # Returns 0-6, Monday-Sunday
    first_day_of_month = (first_day_of_month + 1) % 7  # Adjust to make Sunday = 0

    # Pad the beginning of the month with empty days
    padded_days = [''] * first_day_of_month + days_in_month

        # Limit to a maximum of 7 rows (7 days x 6 rows if overflow, note that normally only one row will be needed to display)
    # Calculate how many rows we need (84 slots for a month max, but we are only showing 7 rows)
    num_rows = (len(padded_days) + 6) // 7  # To fill the calendar

    french_months = [
        "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
        "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"
    ]

    # Fetch checklists for the selected month
    checklists = Checklist.objects.filter(date__month=int(selected_month), is_active=True)


    form2 = ChecklistForm()
    selected_date = date(current_year, selected_month, selected_day)
    # Assume created and modified status are marked by a boolean or similar field
    created_checklists = Checklist.objects.filter(date=selected_date, created=True, is_active=True)
    modified_checklists = Checklist.objects.filter(date=selected_date, modified=True, is_active=True)

    created_count = created_checklists.count()
    modified_count = modified_checklists.count()

    days_counts = {}
    for day in days_in_month:
        day_date = date(current_year, selected_month, day)
        created_count = Checklist.objects.filter(date=day_date, statusro='nouveau', is_active=True).count()
        modified_count = Checklist.objects.filter(date=day_date, statusro='modifié', is_active=True).count()

        # Fetch the checklist status for the day (assuming one or more checklists)
        status = Checklist.objects.filter(date=day_date, is_active=True).aggregate(status=Max('status'))  # Adjust accordingly
        days_counts[day] = {
            'created_count': created_count,
            'modified_count': modified_count,
            'status': status['status'] if status['status'] else 'none',  # Use 'none' for fallback
        }



    product_form = ProductsForm

    if request.method == 'POST':
        if 'product-form' in request.POST:
            product_form = ProductsForm(request.POST, user=request.user)
            if product_form.is_valid():
                product = product_form.save()

                # Log the product creation
                ProductLog.objects.create(product=product, created_by=request.user)

                messages.success(request, 'Nouveau produit créé avec succès.')

                return redirect('creerchecklist')
        else:
            # Handle checklist form
            form2 = ChecklistForm(request.POST)
            if form2.is_valid():
                form2.save()
                messages.success(request, 'Nouvelle Checklist')
                return redirect('creerchecklist')
    else:
        product_form = ProductsForm()
        form2 = ChecklistForm()




    context = {
        'checklists': checklists,
        'form2': form2,
        'encours': encours,
        'valide': valide,
        'refuse': refuse,
        'today': today,
        'days_counts': days_counts,
        'created_count': created_count,
        'modified_count': modified_count,
        'product_form': product_form,
        'months': [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12],  # Months as numbers
        'selected_month': selected_month,
        'days': padded_days,
        'months': months,
        'french_months': french_months,
        'selected_day': selected_day,
        'selected_date': date(current_year, selected_month, selected_day).strftime('%d %B %Y'),
        'years': years,
        'selected_year': current_year,
        'is_chefcuisine': request.user.groups.filter(name='chefcuisine').exists(),
        'is_ventes': request.user.groups.filter(name='ventes').exists(),
        'is_admin': request.user.groups.filter(name='admin').exists(),
        'categories': Category.objects.all(),
    }

    return render(request, 'listings/checklistcreate.html', context)
@login_required
@require_POST
def upload_document(request, checklist_id):
    checklist = get_object_or_404(Checklist, id=checklist_id)
    if request.FILES.get('document'):
        document = ChecklistDocument(
            checklist=checklist,
            document=request.FILES['document'],
            uploaded_by=request.user
        )
        document.save()
        return JsonResponse({'success': True, 'message': 'Document uploaded successfully.'})
    return JsonResponse({'success': False, 'message': 'No document uploaded.'})

from django.http import JsonResponse
from django.utils.dateparse import parse_date
from .models import Checklist
import calendar
from django.utils import formats
@login_required
def get_checklists_for_day(request, day):

    # Get the selected month and year
    selected_month = int(request.GET.get('month', date.today().month))
    selected_year = int(request.GET.get('year', date.today().year))

    # Ensure day is within the range of the selected month
    _, num_days_in_month = calendar.monthrange(selected_year, selected_month)
    day = min(int(day), num_days_in_month)  # Prevent out-of-range days

    # Construct the date object for filtering checklists
    selected_date = date(selected_year, selected_month, day)

    selected_date_formatted = formats.date_format(selected_date, "j F Y", use_l10n=True)

    # Retrieve only active checklists for the selected date
    checklists = Checklist.objects.filter(date=selected_date, is_active=True).select_related('conseillere').values(
        'id', 'name', 'date', 'heure_livraison', 'nb_convive', 'status',
        'conseillere__user__username', 'added_on', 'statusro', 'previous_statusro'
    )

    created_count = Checklist.objects.filter(date=selected_date, created=True).count()
    modified_count = Checklist.objects.filter(date=selected_date, modified=True).count()
    # Prepare data for JSON response
    checklists_list = list(checklists)

    # Format the date objects for JSON serialization
    for checklist in checklists_list:
        checklist['date'] = checklist['date'].strftime('%Y-%m-%d')
        checklist['conseillere'] = checklist.get('conseillere__user__username', 'N/A')  # Get the username or set to 'N/A'

    response_data = {
        'checklists': checklists_list,
        'createdCount': created_count,
        'modifiedCount': modified_count,
        'selected_date': selected_date_formatted,
    }

    return JsonResponse(response_data)

from django.core.paginator import Paginator
from django.http import JsonResponse
from django.utils import formats
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required
import calendar
from datetime import date
from .models import Submission

@login_required
def get_submissions_created_at(request, day):
    # Get the selected month and year
    selected_month = int(request.GET.get('month', date.today().month))
    selected_year = int(request.GET.get('year', date.today().year))

    # Ensure day is within the range of the selected month
    _, num_days_in_month = calendar.monthrange(selected_year, selected_month)
    day = min(int(day), num_days_in_month)  # Prevent out-of-range days

    # Construct the date object for filtering
    selected_date = date(selected_year, selected_month, day)

    # Retrieve submissions created on that date
    submissions_created = Submission.objects.filter(created_at__date=selected_date).values(
        'id', 'company_name', 'submission_type', 'user__username', 'status', 'created_at', 'date'
    )

    submission_list = list(submissions_created)  # Convert the QuerySet to a list directly

    # Format date for JSON serialization
    for submission in submission_list:
        submission['created_at'] = submission['created_at'].strftime('%Y-%m-%d')

    response_data = {
        'submission': submission_list,
        'selected_date': formats.date_format(selected_date, "j F Y", use_l10n=True),
    }

    return JsonResponse(response_data)


@login_required
def get_submission_for_day(request, day):
    # Get the selected month and year
    selected_month = int(request.GET.get('month', date.today().month))
    selected_year = int(request.GET.get('year', date.today().year))

    # Ensure day is within the range of the selected month
    _, num_days_in_month = calendar.monthrange(selected_year, selected_month)
    day = min(int(day), num_days_in_month)  # Prevent out-of-range days

    # Construct the date object for filtering checklists
    selected_date = date(selected_year, selected_month, day)

    selected_date_formatted = formats.date_format(selected_date, "j F Y", use_l10n=True)

    # Retrieve only active checklists for the selected date
    submissions = Submission.objects.filter(date=selected_date).values(
        'id', 'company_name', 'date', 'submission_type', 'user__username', 'status', 'created_at'
    )



    submission_list = list(submissions)

    # Format the date objects for JSON serialization
    for submission in submission_list:
        submission['date'] = submission['date'].strftime('%Y-%m-%d')

    response_data = {
        'submission': submission_list,
        'selected_date': selected_date_formatted,

    }

    return JsonResponse(response_data)

@login_required
def import_items(request):
    if request.method == 'POST' and request.FILES['myfile']:
        myfile = request.FILES['myfile']
        temp_file = NamedTemporaryFile(delete=True, suffix='.xlsx')

        # Write the uploaded file to a temporary file
        for chunk in myfile.chunks():
            temp_file.write(chunk)
        temp_file.flush()

        # Load the Excel workbook
        wb = load_workbook(temp_file.name)
        ws = wb.active

        # Process each row in the Excel sheet
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = row[0]
            description = row[1]
            quantity = row[2]
            photo_url = row[3]  # Assuming the photo URL is in the fourth column

            item = ItemInv(name=name, description=description, quantity=quantity)

            # Download the photo and save it as ImageField
            if photo_url:
                img_temp = NamedTemporaryFile(delete=True)
                img_temp.write(urlopen(photo_url).read())
                img_temp.flush()
                item.photo.save(f'{name}.jpg', File(img_temp))

            item.save()

        messages.success(request, 'Les données ont été importées avec succès.')
        temp_file.close()
    else:
        messages.error(request, 'Veuillez fournir un fichier.')

    return render(request, 'listings/import.html')


from django.http import HttpResponse
from django.urls import reverse
import qrcode
from django.conf import settings
@login_required
def generate_qr_code(request, product_id):
    # Get product details
    product = Product.objects.get(id=product_id)
    quantity = product.quantity  # You can include any relevant information for the QR code

    # URL that will be encoded in the QR code
    product_url = reverse('product_quantity_update', args=[product_id])

    # Generate the full URL including the domain
    full_url = request.build_absolute_uri(product_url)

    # Create the QR code
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(full_url)  # Use the full URL with domain
    qr.make(fit=True)

    # Create an image of the QR code
    img = qr.make_image(fill='black', back_color='white')

    # Create an HTTP response with the image
    response = HttpResponse(content_type="image/png")
    img.save(response, "PNG")
    return response


from django.shortcuts import render, redirect
from .models import Product, QuantityChangeLog
from django.contrib.auth.models import User

@login_required
def update_product_quantity(request, product_id):
    product = Product.objects.get(id=product_id)

    if request.method == 'POST':
        increment = request.POST.get('quantity')

        # Ensure increment is a valid number
        if increment and increment.isdigit():
            # Get the previous quantity
            previous_quantity = product.quantity

            # Add the increment to the current quantity
            product.quantity += int(increment)
            product.save()  # Save the updated product

            # Log the change
            QuantityProductChangeLog.objects.create(
                product=product,
                previous_quantity=previous_quantity,
                new_quantity=product.quantity,
                user=request.user,  # The user who made the change
            )

            return redirect('product_list')  # Redirect to the product list or another page

        else:
            # Handle invalid input (non-numeric or empty value)
            return render(request, 'listings/update_quantity.html', {'product': product, 'error': 'Veuillez entrer une quantité valide.'})

    return render(request, 'listings/update_quantity.html', {'product': product})

from django.shortcuts import render
from .models import QuantityChangeLog

from datetime import timedelta
@login_required
def view_quantity_change_logs(request, product_id):
    product = Product.objects.get(id=product_id)
    logs = QuantityProductChangeLog.objects.filter(product=product).order_by('-timestamp')
    product_logs = ProductLog.objects.filter(product=product).order_by('-created_at')
    # Calculate the difference for each log entry and adjust timestamp
    for log in logs:
        log.difference = log.new_quantity - log.previous_quantity
        log.adjusted_timestamp = log.timestamp - timedelta(hours=5)  # Subtract 5 hours

    adjusted_logs = []
    for log in product_logs:
        log.adjusted_created_at = log.created_at - timedelta(hours=5)
        adjusted_logs.append(log)

    context = {
        'product': product,
        'logs': logs,
        'product_logs': adjusted_logs,
    }

    return render(request, 'listings/view_logs.html', context)



from django.shortcuts import render, get_object_or_404, redirect
from .models import Checklist, Livraison
@login_required
def associate_livraison(request, checklist_id):
    checklist = get_object_or_404(Checklist, id=checklist_id)

    # Find Livraison that has the same num_commande as the checklist
    livraison = Livraison.objects.filter(num_commande=checklist.num_contrat).first()

    if livraison:
        checklist.livraison = livraison
        checklist.save()
        messages.success(request, "Livraison associée avec succès")
    else:
        messages.error(request, "Pas de livraison trouvée")

    return redirect('checklist-detail', checklist_id=checklist.id)
from django.utils.timezone import now
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import redirect
from .models import Checklist, Livraison

from django.http import JsonResponse

@login_required
def associate_all_livraisons(request):
    today = now().date()
    checklists = Checklist.objects.all()
    associated_count = 0

    for checklist in checklists:
        livraison = Livraison.objects.filter(
            num_commande=checklist.num_contrat,
            date__gte=today
        ).first()
        if livraison:
            checklist.livraison = livraison
            checklist.save()
            associated_count += 1

    if associated_count > 0:
        message = f"{associated_count} checklists ont été associées avec succès."
    else:
        message = "Aucune checklist n'a pu être associée."

    # Return JSON response
    return JsonResponse({
        'associated_count': associated_count,
        'message': message
    })


from django.http import JsonResponse
import json
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Q

@csrf_exempt
def group_livraisons(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            ids = data.get('ids', [])
            # Récupérer toutes les livraisons sélectionnées
            livraisons = Livraison.objects.filter(id__in=ids)

            if not livraisons.exists():
                return JsonResponse({'status': 'error', 'message': 'Aucune livraison sélectionnée.'})

            # Grouper par heure de livraison
            # Ensuite, regrouper par nom similaire
            groups = {}
            for liv in livraisons:
                key = (liv.heure_livraison, get_similar_name(liv.nom))
                if key not in groups:
                    groups[key] = []
                groups[key].append(liv)

            # Regrouper chaque groupe
            count_regroupements = 0
            for key, group_list in groups.items():
                # Prioriser livraison avec checklist si possible
                checklist_liv = None
                if hasattr(group_list[0], 'checklist'):
                    checklist_liv = next((l for l in group_list if hasattr(l, 'checklist') and l.checklist.exists()), None)
                main_liv = checklist_liv or group_list[0]

                # Construire le nom du groupe
                noms = [l.nom for l in group_list]
                nouveau_nom = '+'.join(set(noms))
                # Additionner convives
                total_convives = sum(int(l.convives or 0) for l in group_list if l.convives.isdigit())
                main_liv.nom = nouveau_nom
                main_liv.convives = str(total_convives)
                main_liv.save()

                # Supprimer ou réassigner autres livraisons
                for l in group_list:
                    if l.id != main_liv.id:
                        l.delete()

                count_regroupements += 1

            return JsonResponse({'status': 'success', 'message': f'{count_regroupements} regroupements effectués.'})

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

def get_similar_name(name):
    # Si nom vide ouNone, renvoyer une chaîne vide
    if not name:
        return ''
    # Supposer que le nom peut contenir une version ou une partie commune séparée par un espace ou un point
    # On veut regrouper par le prefix avant un point ou un espace
    # Exemple: "SAP 1" et "SAP 1.1" => "SAP 1"
    # On peut utiliser une expression régulière ou une logique simple
    
    # Importation de re si nécessaire
    import re
    
    # Chercher un motif : tout jusqu'à un premier point ou espace
    match = re.match(r'^(.+?)(?:[\.\s].*)?$', name)
    if match:
        return match.group(1)
    return name  # Si rien d'autre, retourner le nom complet




@login_required
def tacheslist(request):

    today = now().date()
    tachestoday = Tacheafaire.objects.filter(date = today)
    tachesok = Tacheafaire.objects.filter(status=True)
    tachesko = Tacheafaire.objects.filter(status=False)
    taches = Tacheafaire.objects.all()



    context = {
        'tachestoday':tachestoday,
        'tachesok':tachesok,
        'tachesko':tachesko,
        'taches':taches,
                }
    if not request.user.is_superuser:
        return redirect('unauthorized')
    return render(request, 'listings/tacheslist.html', context)
@login_required
def inventory_list(request):
    items = ItemInv.objects.all()
    query = request.GET.get('query')

    if query:
        items = items.filter(name__icontains=query)

    context = {
        'items': items,
        'query': query,
        'form': SearchFormInv()
    }
    return render(request, 'listings/inventory_list.html', context)
@login_required
@csrf_exempt
def save_positions(request):
    if request.method == 'POST':
        positions = json.loads(request.POST['positions'])
        for position, livraison_id in enumerate(positions):
            Livraison.objects.filter(id=livraison_id).update(position=position)
        return JsonResponse({'success': True})
    return JsonResponse({'success': False})

from django.http import JsonResponse
from .models import Product

@login_required
def search_productsbase(request, checklist_id):
    query = request.GET.get('query', '')
    checklist_items = ChecklistItem.objects.filter(checklist_id=checklist_id)
    products = Product.objects.filter(name__icontains=query, category = 1) if query else Product.objects.filter( category = 1)

    product_data = []
    for product in products:
        # Get the quantity for each product from the checklist
        checklist_item = checklist_items.filter(product=product).first()
        quantity = checklist_item.quantity if checklist_item else 0

        product_data.append({
            'id': product.id,
            'name': product.name,
            'quantity': product.quantity,
            'checklist_quantity': quantity,  # Add this to send the checklist quantity
            'commentaire': checklist_item.commentaire if checklist_item else '',  # Ensure commentaire is included
        })

    return JsonResponse({'products': product_data})
@login_required
def search_productscfcdn(request, checklist_id):
    query = request.GET.get('query', '')
    checklist_items = ChecklistItem.objects.filter(checklist_id=checklist_id)
    products = Product.objects.filter(name__icontains=query, category = 17) if query else Product.objects.filter( category = 17)

    product_data = []
    for product in products:
        # Get the quantity for each product from the checklist
        checklist_item = checklist_items.filter(product=product).first()
        quantity = checklist_item.quantity if checklist_item else 0

        product_data.append({
            'id': product.id,
            'name': product.name,
            'quantity': product.quantity,
            'checklist_quantity': quantity,  # Add this to send the checklist quantity
            'commentaire': checklist_item.commentaire if checklist_item else '',  # Ensure commentaire is included
        })
    return JsonResponse({'products': product_data})

@login_required
def search_productsjetable(request, checklist_id):
    query = request.GET.get('query', '')
    checklist_items = ChecklistItem.objects.filter(checklist_id=checklist_id)
    products = Product.objects.filter(name__icontains=query, category = 2) if query else Product.objects.filter( category = 2)

    product_data = []
    for product in products:
        # Get the quantity for each product from the checklist
        checklist_item = checklist_items.filter(product=product).first()
        quantity = checklist_item.quantity if checklist_item else 0

        product_data.append({
            'id': product.id,
            'name': product.name,
            'quantity': product.quantity,
            'checklist_quantity': quantity,  # Add this to send the checklist quantity
            'commentaire': checklist_item.commentaire if checklist_item else '',  # Ensure commentaire is included
        })
    return JsonResponse({'products': product_data})
@login_required
def search_productsdecor(request, checklist_id):
    query = request.GET.get('query', '')
    checklist_items = ChecklistItem.objects.filter(checklist_id=checklist_id)
    products = Product.objects.filter(name__icontains=query, category = 3) if query else Product.objects.filter( category = 3)

    product_data = []
    for product in products:
        # Get the quantity for each product from the checklist
        checklist_item = checklist_items.filter(product=product).first()
        quantity = checklist_item.quantity if checklist_item else 0

        product_data.append({
            'id': product.id,
            'name': product.name,
            'quantity': product.quantity,
            'checklist_quantity': quantity,  # Add this to send the checklist quantity
            'commentaire': checklist_item.commentaire if checklist_item else '',  # Ensure commentaire is included
        })

    return JsonResponse({'products': product_data})
@login_required
def search_productsbar(request, checklist_id):
    query = request.GET.get('query', '')
    checklist_items = ChecklistItem.objects.filter(checklist_id=checklist_id)
    products = Product.objects.filter(name__icontains=query, category = 4) if query else Product.objects.filter( category = 4)

    product_data = []
    for product in products:
        # Get the quantity for each product from the checklist
        checklist_item = checklist_items.filter(product=product).first()
        quantity = checklist_item.quantity if checklist_item else 0

        product_data.append({
            'id': product.id,
            'name': product.name,
            'quantity': product.quantity,
            'checklist_quantity': quantity,  # Add this to send the checklist quantity
            'commentaire': checklist_item.commentaire if checklist_item else '',  # Ensure commentaire is included
        })

    return JsonResponse({'products': product_data})
@login_required
def search_productscafe(request, checklist_id):
    query = request.GET.get('query', '')
    checklist_items = ChecklistItem.objects.filter(checklist_id=checklist_id)
    products = Product.objects.filter(name__icontains=query, category = 5) if query else Product.objects.filter( category = 5)

    product_data = []
    for product in products:
        # Get the quantity for each product from the checklist
        checklist_item = checklist_items.filter(product=product).first()
        quantity = checklist_item.quantity if checklist_item else 0

        product_data.append({
            'id': product.id,
            'name': product.name,
            'quantity': product.quantity,
            'checklist_quantity': quantity,  # Add this to send the checklist quantity
            'commentaire': checklist_item.commentaire if checklist_item else '',  # Ensure commentaire is included
        })

    return JsonResponse({'products': product_data})
@login_required
def get_breuvages_products(request, checklist_id):
    breuvages_items = ChecklistItem.objects.filter(checklist_id=checklist_id, product__category="BREUVAGE")

    product_data = [
        {
            'id': item.product.id,
            'name': item.product.name,
            'total_quantity': item.total_quantity,
            'consumed_quantity': item.consumed_quantity,
            'unconsumed_quantity': item.unconsumed_quantity
        }
        for item in breuvages_items
    ]

    return JsonResponse({'products': product_data})
@login_required
def search_productstable(request, checklist_id):
    query = request.GET.get('query', '')
    checklist_items = ChecklistItem.objects.filter(checklist_id=checklist_id)
    products = Product.objects.filter(name__icontains=query, category = 7) if query else Product.objects.filter( category = 7)

    product_data = []
    for product in products:
        # Get the quantity for each product from the checklist
        checklist_item = checklist_items.filter(product=product).first()
        quantity = checklist_item.quantity if checklist_item else 0

        product_data.append({
            'id': product.id,
            'name': product.name,
            'quantity': product.quantity,
            'checklist_quantity': quantity,  # Add this to send the checklist quantity
            'commentaire': checklist_item.commentaire if checklist_item else '',  # Ensure commentaire is included
        })

    return JsonResponse({'products': product_data})
@login_required
def search_productsverre(request, checklist_id):
    query = request.GET.get('query', '')
    checklist_items = ChecklistItem.objects.filter(checklist_id=checklist_id)
    products = Product.objects.filter(name__icontains=query, category = 8) if query else Product.objects.filter( category = 8)

    product_data = []
    for product in products:
        # Get the quantity for each product from the checklist
        checklist_item = checklist_items.filter(product=product).first()
        quantity = checklist_item.quantity if checklist_item else 0

        product_data.append({
            'id': product.id,
            'name': product.name,
            'quantity': product.quantity,
            'checklist_quantity': quantity,  # Add this to send the checklist quantity
            'commentaire': checklist_item.commentaire if checklist_item else '',  # Ensure commentaire is included
        })

    return JsonResponse({'products': product_data})
@login_required
def search_productsporcelaine(request, checklist_id):
    query = request.GET.get('query', '')
    checklist_items = ChecklistItem.objects.filter(checklist_id=checklist_id)
    products = Product.objects.filter(name__icontains=query, category = 9) if query else Product.objects.filter( category = 9)

    product_data = []
    for product in products:
        # Get the quantity for each product from the checklist
        checklist_item = checklist_items.filter(product=product).first()
        quantity = checklist_item.quantity if checklist_item else 0

        product_data.append({
            'id': product.id,
            'name': product.name,
            'quantity': product.quantity,
            'checklist_quantity': quantity,  # Add this to send the checklist quantity
            'commentaire': checklist_item.commentaire if checklist_item else '',  # Ensure commentaire is included
        })

    return JsonResponse({'products': product_data})
@login_required
def search_productscanape(request, checklist_id):
    query = request.GET.get('query', '')
    checklist_items = ChecklistItem.objects.filter(checklist_id=checklist_id)
    products = Product.objects.filter(name__icontains=query, category = 10) if query else Product.objects.filter( category = 10)

    product_data = []
    for product in products:
        # Get the quantity for each product from the checklist
        checklist_item = checklist_items.filter(product=product).first()
        quantity = checklist_item.quantity if checklist_item else 0

        product_data.append({
            'id': product.id,
            'name': product.name,
            'quantity': product.quantity,
            'checklist_quantity': quantity,  # Add this to send the checklist quantity
            'commentaire': checklist_item.commentaire if checklist_item else '',  # Ensure commentaire is included
        })

    return JsonResponse({'products': product_data})
@login_required
def search_productscuisson(request, checklist_id):
    query = request.GET.get('query', '')
    checklist_items = ChecklistItem.objects.filter(checklist_id=checklist_id)
    products = Product.objects.filter(name__icontains=query, category = 11) if query else Product.objects.filter( category = 11)

    product_data = []
    for product in products:
        # Get the quantity for each product from the checklist
        checklist_item = checklist_items.filter(product=product).first()
        quantity = checklist_item.quantity if checklist_item else 0

        product_data.append({
            'id': product.id,
            'name': product.name,
            'quantity': product.quantity,
            'checklist_quantity': quantity,  # Add this to send the checklist quantity
            'commentaire': checklist_item.commentaire if checklist_item else '',  # Ensure commentaire is included
        })

    return JsonResponse({'products': product_data})
@login_required
def search_productsservice(request, checklist_id):
    query = request.GET.get('query', '')
    checklist_items = ChecklistItem.objects.filter(checklist_id=checklist_id)
    products = Product.objects.filter(name__icontains=query, category = 12) if query else Product.objects.filter( category = 12)

    product_data = []
    for product in products:
        # Get the quantity for each product from the checklist
        checklist_item = checklist_items.filter(product=product).first()
        quantity = checklist_item.quantity if checklist_item else 0

        product_data.append({
            'id': product.id,
            'name': product.name,
            'quantity': product.quantity,
            'checklist_quantity': quantity,  # Add this to send the checklist quantity
            'commentaire': checklist_item.commentaire if checklist_item else '',  # Ensure commentaire is included
        })

    return JsonResponse({'products': product_data})
@login_required
def search_productsdivers(request, checklist_id):
    query = request.GET.get('query', '')
    checklist_items = ChecklistItem.objects.filter(checklist_id=checklist_id)
    products = Product.objects.filter(name__icontains=query, category = 6) if query else Product.objects.filter( category = 6)

    product_data = []
    for product in products:
        # Get the quantity for each product from the checklist
        checklist_item = checklist_items.filter(product=product).first()
        quantity = checklist_item.quantity if checklist_item else 0

        product_data.append({
            'id': product.id,
            'name': product.name,
            'quantity': product.quantity,
            'checklist_quantity': quantity,  # Add this to send the checklist quantity
            'commentaire': checklist_item.commentaire if checklist_item else '',  # Ensure commentaire is included
        })

    return JsonResponse({'products': product_data})
@login_required
def search_productsalcoolfort(request, checklist_id):
    query = request.GET.get('query', '')
    checklist_items = ChecklistItem.objects.filter(checklist_id=checklist_id)
    products = Product.objects.filter(name__icontains=query, category = 13) if query else Product.objects.filter( category = 13)

    product_data = []
    for product in products:
        # Get the quantity for each product from the checklist
        checklist_item = checklist_items.filter(product=product).first()
        quantity = checklist_item.quantity if checklist_item else 0

        product_data.append({
            'id': product.id,
            'name': product.name,
            'quantity': product.quantity,
            'checklist_quantity': quantity,  # Add this to send the checklist quantity
            'commentaire': checklist_item.commentaire if checklist_item else '',  # Ensure commentaire is included
        })

    return JsonResponse({'products': product_data})
@login_required
def search_productsbieres(request, checklist_id):
    query = request.GET.get('query', '')
    checklist_items = ChecklistItem.objects.filter(checklist_id=checklist_id)
    products = Product.objects.filter(name__icontains=query, category = 14) if query else Product.objects.filter( category = 14)

    product_data = []
    for product in products:
        # Get the quantity for each product from the checklist
        checklist_item = checklist_items.filter(product=product).first()
        quantity = checklist_item.quantity if checklist_item else 0

        product_data.append({
            'id': product.id,
            'name': product.name,
            'quantity': product.quantity,
            'checklist_quantity': quantity,  # Add this to send the checklist quantity
            'commentaire': checklist_item.commentaire if checklist_item else '',  # Ensure commentaire is included
        })

    return JsonResponse({'products': product_data})
@login_required
def search_productsvins(request, checklist_id):
    query = request.GET.get('query', '')
    checklist_items = ChecklistItem.objects.filter(checklist_id=checklist_id)
    products = Product.objects.filter(name__icontains=query, category = 15) if query else Product.objects.filter( category = 15)

    product_data = []
    for product in products:
        # Get the quantity for each product from the checklist
        checklist_item = checklist_items.filter(product=product).first()
        quantity = checklist_item.quantity if checklist_item else 0

        product_data.append({
            'id': product.id,
            'name': product.name,
            'quantity': product.quantity,
            'checklist_quantity': quantity,  # Add this to send the checklist quantity
            'commentaire': checklist_item.commentaire if checklist_item else '',  # Ensure commentaire is included
        })

    return JsonResponse({'products': product_data})
@login_required
def search_productssansalcool(request, checklist_id):
    query = request.GET.get('query', '')
    checklist_items = ChecklistItem.objects.filter(checklist_id=checklist_id)
    products = Product.objects.filter(name__icontains=query, category = 16) if query else Product.objects.filter( category = 16)

    product_data = []
    for product in products:
        # Get the quantity for each product from the checklist
        checklist_item = checklist_items.filter(product=product).first()
        quantity = checklist_item.quantity if checklist_item else 0

        product_data.append({
            'id': product.id,
            'name': product.name,
            'quantity': product.quantity,
            'checklist_quantity': quantity,  # Add this to send the checklist quantity
            'commentaire': checklist_item.commentaire if checklist_item else '',  # Ensure commentaire is included
        })

    return JsonResponse({'products': product_data})
@login_required
def save_breuvages_report(request, checklist_id):
    if request.method == "POST":
        for item_id, consumed in request.POST.items():
            if item_id.startswith("consumed_"):
                item_id = int(item_id.split("_")[1])
                checklist_item = ChecklistItem.objects.get(id=item_id)

                consumed_quantity = int(consumed)
                unconsumed_quantity = int(request.POST.get(f'unconsumed_{item_id}', 0))

                checklist_item.consumed_quantity = consumed_quantity
                checklist_item.unconsumed_quantity = unconsumed_quantity
                checklist_item.save()

        return JsonResponse({"success": True})
    return JsonResponse({"success": False})

# views.py
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from listings.models import Livraison, Livreur, UserProfile  # Adjust 'your_app' to your app name

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from listings.models import Livraison, Livreur, UserProfile

@login_required
def home(request):
    user = request.user

    # Ensure the user has a UserProfile
    if hasattr(user, 'userprofile'):
        user_profile = user.userprofile

        # If `force_password_change` is True, redirect to the password change page
        if user_profile.force_password_change:
            return redirect('password_change')
    else:
        # Optionally create a UserProfile if missing (for testing purposes)
        # UserProfile.objects.create(user=user)

        return redirect('error_page')  # Or handle the error gracefully

    livraisons = Livraison.objects.all()
    livreurs = Livreur.objects.all()

    return render(request, 'listings/home.html', {
        'livraisons': livraisons,
        'livreurs': livreurs,
    })



@login_required
def livraisons_list(request):

    livraisons  = Livraison.objects.all()
    livreurs = Livreur.objects.all()
    journees = Journee.objects.all()
    return render(request, 'listings/livraisons_list.html', context={'livraisons': livraisons,
                                                              'livreurs': livreurs,
                                                              'journees' : journees})
@login_required
def responsable_list(request):
    today = now().date()
    livraisons  = Livraison.objects.order_by('statut', 'position').filter(date = today)
    livraisonsok  = Livraison.objects.filter(date = today, recuperation=False)
    livraisonsrecup  = Livraison.objects.filter(date = today, recuperation=True)
    journees = Journee.objects.all().order_by('-date')


    if request.method == 'GET' and 'date' in request.GET:
        form = DateFilterForm(request.GET)
        if form.is_valid():
            date = form.cleaned_data['date']
            journees = journees.filter(date=date)
    else:
        form = DateFilterForm()

    paginator = Paginator(journees, 7)  # Show 10 events per page

    page = request.GET.get('page')
    try:
        journees = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        journees = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g., 9999), deliver last page of results.
        journees = paginator.page(paginator.num_pages)

    livreurs = Livreur.objects.all()

    ventes = "Ventes"
    cuisine = "Cuisine"
    if not request.user.is_superuser:
        return redirect('unauthorized')
    return render(request, 'listings/responsableslist.html', context={'livraisons': livraisons,
                                                              'livreurs': livreurs,
                                                              'journees' : journees,
                                                              'ventes':ventes,
                                                              'today':today,
                                                              'livraisonsok':livraisonsok,
                                                              'livraisonsrecup':livraisonsrecup,
                                                              'cuisine':cuisine,
                                                              'form':form})

# View to handle Recupfrigo and its items
@login_required
def create_recupfrigo(request):
    # Define the acceptable values for mode_envoi
    valid_modes_envoi = [
        "Porcelaine",
        "Chaud et porcelaine",
        "Porcelaine et bois",
        "Plateau de bois",
        "Froid et bois",
        "Chaud et jetable"
    ]

    # Get today's date
    today = date.today()

    # Filter livraisons by date and mode_envoi
    livraisons = Livraison.objects.filter(
        date_livraison=today,
        mode_envoi__in=valid_modes_envoi
    )
    print(livraisons)

    if request.method == 'POST':
        recupfrigo_form = RecupfrigoForm(request.POST)
        formset = RecupfrigoItemFormset(request.POST)

        if recupfrigo_form.is_valid() and formset.is_valid():
            # Save the recupfrigo
            recupfrigo = recupfrigo_form.save(commit=False)
            recupfrigo.save()

            # Save the formset
            items = formset.save(commit=False)
            for item in items:
                item.recupfrigo = recupfrigo  # Set the Recupfrigo for each item
                item.save()

            return redirect(reverse('create_recupfrigo'))

    else:
        recupfrigo_form = RecupfrigoForm()
        formset = RecupfrigoItemFormset()

    return render(request, 'listings/create_recupfrigo.html', {
        'recupfrigo_form': recupfrigo_form,
        'formset': formset,
        'livraisons': livraisons
    })
# View to handle Recuplivreur and its items
@login_required
def create_recuplivreur(request, livraison_id):
    # Fetch the Livraison instance
    livraison = get_object_or_404(Livraison, id=livraison_id)

    if request.method == 'POST':
        recuplivreur_form = RecuplivreurForm(request.POST)
        formset = RecuplivreurItemFormset(request.POST)

        if recuplivreur_form.is_valid() and formset.is_valid():
            # Create the recuplivreur object but don't save to DB yet
            recuplivreur = recuplivreur_form.save(commit=False)

            # Ensure that livraison and date are set properly
            recuplivreur.livraison = livraison
            recuplivreur.date = livraison.date_livraison - timedelta(days=1) # Set the date from livraison
            recuplivreur.filled_by = request.user  # Track the user
            recuplivreur.filled_at = timezone.now()  # Track the timestamp

            # Now save the recuplivreur object to the database
            recuplivreur.save()

            # Save the formset with recuplivreur linked
            items = formset.save(commit=False)
            for item in items:
                item.recuplivreur = recuplivreur  # Set the recuplivreur for each item
                item.save()

            # Redirect after successful save
            return redirect(reverse('create_recuplivreur', args=[livraison_id]))

    else:
        recuplivreur_form = RecuplivreurForm()
        formset = RecuplivreurItemFormset()

    return render(request, 'listings/create_recuplivreur.html', {
        'recuplivreur_form': recuplivreur_form,
        'formset': formset,
        'livraison': livraison
    })

@login_required
def journeerecupdetail(request, id):
    journee = get_object_or_404(Journee, id=id)
    recupfrigo = Recupfrigo.objects.filter(date=journee.date)
    recuplivreur = Recuplivreur.objects.filter(date=journee.date)

    # Build a comparison dictionary by item name
    comparison_data = []
    frigo_dict = {}
    livreur_dict = {}

    # Group RecupFrigo items by name
    for frigo in recupfrigo:
        for item in frigo.items_frigo.all():
            if item.item_name not in frigo_dict:
                frigo_dict[item.item_name] = []
            frigo_dict[item.item_name].append({
                'livraison_nom': frigo.livraison.nom,
                'frigo_quantity': item.quantity,
                'frigo_filled_by': frigo.filled_by,   # User who filled the form
                'frigo_filled_at': frigo.filled_at    # Timestamp when form was filled
            })

    # Group RecupLivreur items by name
    for livreur in recuplivreur:
        for item in livreur.items_livreur.all():
            if item.item_name not in livreur_dict:
                livreur_dict[item.item_name] = []
            livreur_dict[item.item_name].append({
                'livraison_nom': livreur.livraison.nom,
                'livreur_quantity': item.quantity,
                'livreur_filled_by': livreur.filled_by,   # User who filled the form
                'livreur_filled_at': livreur.filled_at,    # Timestamp when form was filled
                'livreur_nom': livreur.filled_by.username  # Access the username
            })

    # Build a comparison list
    all_items = set(list(frigo_dict.keys()) + list(livreur_dict.keys()))
    for item_name in all_items:
        livraisons = []
        frigo_data = frigo_dict.get(item_name, [])
        livreur_data = livreur_dict.get(item_name, [])

        # Combine livraisons by name for comparison
        livraison_names = set([data['livraison_nom'] for data in frigo_data] + [data['livraison_nom'] for data in livreur_data])
        for livraison_name in livraison_names:
            frigo_quantity = next((data['frigo_quantity'] for data in frigo_data if data['livraison_nom'] == livraison_name), 0)
            livreur_quantity = next((data['livreur_quantity'] for data in livreur_data if data['livraison_nom'] == livraison_name), 0)
            difference = livreur_quantity - frigo_quantity

            # Get user and timestamp info for each livraison
            frigo_filled_by = next((data['frigo_filled_by'] for data in frigo_data if data['livraison_nom'] == livraison_name), None)
            livreur_filled_by = next((data['livreur_filled_by'] for data in livreur_data if data['livraison_nom'] == livraison_name), None)
            frigo_filled_at = next((data['frigo_filled_at'] for data in frigo_data if data['livraison_nom'] == livraison_name), None)
            livreur_filled_at = next((data['livreur_filled_at'] for data in livreur_data if data['livraison_nom'] == livraison_name), None)
            livreur_nom = next((data['livreur_nom'] for data in livreur_data if data['livraison_nom'] == livraison_name), None)

            livraisons.append({
                'nom': livraison_name,
                'frigo_quantity': frigo_quantity,
                'livreur_quantity': livreur_quantity,
                'difference': difference,
                'frigo_filled_by': frigo_filled_by.username if frigo_filled_by else 'N/A',
                'livreur_filled_by': livreur_filled_by.username if livreur_filled_by else 'N/A',
                'frigo_filled_at': frigo_filled_at,
                'livreur_filled_at': livreur_filled_at,
                'livreur_nom': livreur_nom,
            })

        comparison_data.append({
            'item_name': item_name,
            'frigo_quantity': sum([data['frigo_quantity'] for data in frigo_data]),
            'livreur_quantity': sum([data['livreur_quantity'] for data in livreur_data]),
            'quantity_difference': sum([data['difference'] for data in livraisons]),
            'livraisons': livraisons,  # Include livraisons for dropdown
            'mismatch': any(data['difference'] != 0 for data in livraisons)
        })

    context = {
        'journee': journee,
        'comparison_data': comparison_data,
    }

    return render(request, 'listings/journeerecupdetail.html', context)


@login_required
def journeedetailvente(request, id):
    journees = get_object_or_404(Journee, id=id)
    livreurs = Livreur.objects.all()
    shifts = Shift.objects.filter(date=journees.date)
    livraisonsroute = Livraison.objects.order_by('statut', 'position')
    today = now().date()

    livraisonss = Livraison.objects.prefetch_related('statut__livreur') \
    .filter(date=journees.date) \
    .order_by('statut', 'position')

    # Convert to JSON-like structure with error handling for 'livreur'
    livraisons_data = [
        {
            'id': livraison.id,
            'nom': livraison.nom,
            'infodetail': livraison.infodetail if livraison.infodetail else "N/A",
            'heure_livraison': livraison.heure_livraison,
            'heure_livraison_classement': livraison.heure_livraison_classement,
            'livreurs': [livreur.nom for livreur in livraison.statut.livreur.all()] if livraison.statut else ["Aucun livreur"],  # Handle ManyToManyField
            'heure_depart': livraison.statut.heure_depart if livraison.statut else "Non défini",
            'recuperation': livraison.recuperation,
            'status': livraison.status,
            'statut.heure_depart': livraison.statut.heure_depart,
        }
        for livraison in livraisonss
    ]

    livraisons_json = json.dumps(livraisons_data)
    livraisonsok = Livraison.objects.filter(recuperation=False, date=journees.date)
    recuperations = Livraison.objects.filter(recuperation=True, date=journees.date)
    recuperationes = Livraison.objects.filter(recuperation=True, date=journees.date)

    retourtraiteur = "oui"
    retourtraiteurno = "non"
    recuperation = "oui"
    recuperationo = "non"
    loic = "Loic"
    maxime = "Maxime"
    rien = "."

    return render(request,
                  'listings/journeedetailvente.html',
                  context={
                      'journees': journees,
                      'livraisonsroute': livraisonsroute,
                      'livreurs': livreurs,
                      'recuperations': recuperations,
                      'retourtraiteur': retourtraiteur,
                      'recuperation': recuperation,
                      'retourtraiteurno': retourtraiteurno,
                      'livraisonss': livraisons_json,
                      'recuperationo': recuperationo,
                      'loic': loic,
                      'maxime': maxime,
                      'rien': rien,
                      'recuperations': recuperation,
                      'livraisons_data': livraisons_data,
                      'livraisonsok': livraisonsok,
                      'recuperationes': recuperationes,
                      'today': today,
                      'shifts': shifts,
                  })  # nous passons l'id au modèle

@login_required
def journees_list(request):
    today = now().date()
    livraisons  = Livraison.objects.order_by('statut', 'position').filter(date = today)
    livraisonsok  = Livraison.objects.filter(date = today, recuperation=False)
    livraisonsrecup  = Livraison.objects.filter(date = today, recuperation=True)
    journees = Journee.objects.all().order_by('-date')
    selected_date = request.GET.get('journeeSelect')
    max = "Maxime"
    loic = "Loic"
    jef = "Jef"
    md = "md"
    livraisonschaud = Livraison.objects.filter(
    nom__icontains="part chaud",
    statut__livreur__isnull=False,  # Ensure a livreur is linked
    statut__heure_depart__isnull=False,
    recuperation=False,
    date=selected_date
    ).prefetch_related('statut__livreur').distinct()  # Ensure distinct results




    if request.method == 'GET' and 'date' in request.GET:
        form = DateFilterForm(request.GET)
        if form.is_valid():
            date = form.cleaned_data['date']
            journees = journees.filter(date=date)
    else:
        form = DateFilterForm()

    paginator = Paginator(journees, 7)  # Show 10 events per page

    page = request.GET.get('page')
    try:
        journees = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        journees = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g., 9999), deliver last page of results.
        journees = paginator.page(paginator.num_pages)

    livreurs = Livreur.objects.all()

    ventes = "Ventes"
    cuisine = "Cuisine"
    return render(request, 'listings/journees_list.html', context={'livraisons': livraisons,
                                                              'livreurs': livreurs,
                                                              'journees' : journees,
                                                              'ventes':ventes,
                                                              'today':today,
                                                              'livraisonsok':livraisonsok,
                                                              'livraisonsrecup':livraisonsrecup,
                                                              'cuisine':cuisine,
                                                              'livraisonschaud':livraisonschaud,
                                                              'form':form,
                                                              'max':max,
                                                              'is_md': request.user.groups.filter(name='md').exists(),
                                                              'is_checklist': request.user.groups.filter(name='checklist').exists(),
                                                              'is_ventes': request.user.groups.filter(name='ventes').exists(),
                                                              'is_livreur': request.user.groups.filter(name='livreurs').exists(),
                                                              'is_cuisine': request.user.groups.filter(name='cuisine').exists(),
                                                              'is_chaud': request.user.groups.filter(name='chaud').exists(),
                                                              'is_chefcuisine': request.user.groups.filter(name='chefcuisine').exists(),
                                                              'jef':jef,
                                                              'md':md,
                                                              'loic':loic})

@login_required
def jeux(request):
    top_scores = Score1.objects.order_by('-score')[:5]
    return render(request, 'listings/jeux.html', {
        'top_scores': top_scores,
    })

def top_scores(request):
    scores = Score1.objects.order_by('-score')[:5]
    data = {
        'scores': [
            {'user': s.user.username, 'score': s.score}
            for s in scores
        ]
    }
    return JsonResponse(data)

@login_required
def save_score1(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            score_value = int(data.get('score', 0))
            Score1.objects.create(user=request.user, score=score_value)
            return JsonResponse({'status': 'ok'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)

# RecupFrigo detail view
@login_required
def recupfrigo_detail(request, id):
    recupfrigo = Recupfrigo.objects.get(id=id)
    items = recupfrigo.itemsfrigo.all()

    return render(request, 'listings/recupfrigo_detail.html', {
        'recupfrigo': recupfrigo,
        'items': items,
    })


# Recuplivreur detail view
@login_required
def recuplivreur_detail(request, id):
    recuplivreur = Recuplivreur.objects.get(id=id)
    items = recuplivreur.items.all()

    return render(request, 'listings/recuplivreur_detail.html', {
        'recuplivreur': recuplivreur,
        'items': items,
    })


@login_required
def recupslist(request):
    if request.user.is_authenticated:
        today = now().date()
        journees = Journee.objects.all().order_by('-date')


        # Initialize the form first
        form = DateFilterForm(request.GET or None)

        # Process the form if it's a GET request and has a 'date' field
        if request.method == 'GET' and 'date' in request.GET:
            if form.is_valid():
                date = form.cleaned_data['date']
                journees = journees.filter(date=date)

        paginator = Paginator(journees, 7)  # Show 7 items per page
        page = request.GET.get('page')

        try:
            journees = paginator.page(page)
        except PageNotAnInteger:
            journees = paginator.page(1)  # Deliver first page if page is not an integer
        except EmptyPage:
            journees = paginator.page(paginator.num_pages)  # Deliver last page if out of range

        return render(request, 'listings/recups-list.html', context={
            'journees': journees,
            'form': form,
            'today':today,
        })
    else:
        return redirect('home')
@login_required
def routedetail(request, id):  # notez le paramètre id supplémentaire

   route = Route.objects.get(id=id)
   today = datetime.now().date()
   tomorrow = today + timedelta(1)
   livraison = Livraison.objects.filter(date = tomorrow)
   form = RoutedetailForm(request.POST or None, instance = route)
   if form.is_valid():
       form.save()
       return redirect('my_map_view')

   return render(request,
          'listings/routedetail.html',
          context={'route': route, 'form': form, 'livraison':livraison}) # nous passons l'id au modèle

from django.http import JsonResponse
from django.views.generic.edit import UpdateView
from .models import Route
from .forms import RoutedetailForm

class RouteUpdateView(UpdateView):
    model = Route
    form_class = RoutedetailForm
    template_name = 'partial_template/route_form.html'

    def form_valid(self, form):
        route = form.save(commit=False)  # Create the route instance without saving
        route.save()  # Save the route instance
        form.save_m2m()  # This saves the many-to-many relationships (livreurs)
        return HttpResponseRedirect(self.request.META.get('HTTP_REFERER', '/'))  # Redirect back


@login_required
def journee_detail(request, id):  # notez le paramètre id supplémentaire
    journees = get_object_or_404(Journee, id=id)
    livreurs = Livreur.objects.all()
    shifts = Shift.objects.filter(date=journees.date)
    livraisonsroute = Livraison.objects.order_by('statut', 'position')
    today = now().date()

    livraisonss = Livraison.objects.prefetch_related('statut__livreur') \
    .filter(date=journees.date) \
    .order_by('statut', 'position')

    # Convert to JSON-like structure with error handling for 'livreur'
    livraisons_data = [
        {
            'id': livraison.id,
            'nom': livraison.nom,
            'infodetail': livraison.infodetail if livraison.infodetail else "N/A",
            'heure_livraison': livraison.heure_livraison,
            'heure_livraison_classement': livraison.heure_livraison_classement,
            'livreurs': [livreur.nom for livreur in livraison.statut.livreur.all()] if livraison.statut else ["Aucun livreur"],  # Handle ManyToManyField
            'heure_depart': livraison.statut.heure_depart if livraison.statut else "Non défini",
            'recuperation': livraison.recuperation,
            'status': livraison.status,
            'statut.heure_depart': livraison.statut.heure_depart,
        }
        for livraison in livraisonss
    ]

    livraisons_json = json.dumps(livraisons_data)
    livraisonsok = Livraison.objects.filter(recuperation=False, date=journees.date)
    recuperations = Livraison.objects.filter(recuperation=True, date=journees.date)
    recuperationes = Livraison.objects.filter(recuperation=True, date=journees.date)

    retourtraiteur = "oui"
    retourtraiteurno = "non"
    recuperation = "oui"
    recuperationo = "non"
    loic = "Loic"
    maxime = "Maxime"
    rien = "."

    return render(request,
                  'listings/journee_detail.html',
                  context={
                      'journees': journees,
                      'livraisonsroute': livraisonsroute,
                      'livreurs': livreurs,
                      'recuperations': recuperations,
                      'retourtraiteur': retourtraiteur,
                      'recuperation': recuperation,
                      'retourtraiteurno': retourtraiteurno,
                      'livraisonss': livraisons_json,
                      'recuperationo': recuperationo,
                      'loic': loic,
                      'maxime': maxime,
                      'rien': rien,
                      'recuperations': recuperation,
                      'livraisons_data': livraisons_data,
                      'livraisonsok': livraisonsok,
                      'recuperationes': recuperationes,
                      'today': today,
                      'shifts': shifts,
                  })  # nous passons l'id au modèle


@login_required
def livreur_list(request):
    if request.user.is_authenticated:
        livreurs = Livreur.objects.exclude(user=request.user)
        return render(request, 'listings/livreur_list.html', context={'livreurs': livreurs
                                                                      })
    else:
        return redirect('home')
@login_required
def validate_livraison(request, livraison_id):
    livraison = get_object_or_404(Livraison, id=livraison_id)

    if request.method == 'POST':
        form = LivraisonForm(request.POST, instance=livraison)
        if form.is_valid():
            livraison = form.save(commit=False)
            livraison.status = True  # Set status to True.
            livraison.save()

            # If it's an AJAX request, return JSON response
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'redirect_url': reverse('livraison-detail', args=[livraison.id]),  # Ensure you pass the correct argument
                })
            return redirect('livraison-detail', ip=livraison.id)  # Use `ip` or `livraison_id` based on your URL config

    return redirect('livraison-detail', ip=livraison.id)  # Again, use `ip` or `livraison_id` accordingly


@login_required
def livreur_detail(request, pk):  # notez le paramètre id supplémentaire
    if request.user.is_authenticated :
        livreur = Livreur.objects.get(user_id= pk)
        livraisons  = Livraison.objects.all()
        taches = Tacheafaire.objects.all()
        recuperations = Recuperation.objects.all()
        journee = Journee.objects.all()
        return render(request, "listings/livreur_detail.html", context={'livreur':livreur,
                                                                'livraisons' : livraisons,
                                                                'taches' : taches,
                                                                'recuperations' : recuperations,
                                                                'journee' : journee})
    else:
        return redirect('home')
@login_required
def taskdetail(request, id):
    task = get_object_or_404(Tacheafaire, id=id)  # Get the task instance
    photoss = task.photo

    if request.method == 'POST':
        form = TaskUpdateForm(request.POST, instance=task)
        formbis = PhotoTachesFormSet(request.POST, request.FILES, queryset=Phototaches.objects.filter(tache=task))

        if form.is_valid():
            # Save the task instance first
            task = form.save()

        if formbis.is_valid():
            # Process and save the associated photos
            photos = formbis.save(commit=False)
            for photo in photos:
                photo.tache = task  # Associate each photo with the task
                photo.save()

            # If there are many-to-many fields, save them as well
            formbis.save_m2m()

            return redirect(request.META.get('HTTP_REFERER', 'task_list'))  # Redirect after save


    else:
        form = TaskUpdateForm(instance=task)
        formbis = PhotoTachesFormSet(queryset=Phototaches.objects.filter(tache=task))  # Initialize formset with existing photos for the task

    return render(request, 'listings/taskdetail.html', {'task': task, 'form': form, 'formbis': formbis, 'photoss':photoss,})
@login_required
def view_shifts_by_date(request):
    # Get date from the request or use today’s date as default
    selected_date = request.GET.get('date', now().date())

    # Fetch all shifts for that date
    shifts = Shift.objects.filter(date=selected_date).select_related('livreur')
    if not request.user.is_superuser:
        return redirect('unauthorized')

    return render(request, 'listings/view_shifts_by_date.html', {
        'shifts': shifts,
        'selected_date': selected_date
    })

from django.contrib.auth.views import PasswordChangeView

class CustomPasswordChangeView(PasswordChangeView):
    template_name = 'registration/password_change_form.html'
    success_url = reverse_lazy('home')  # Redirect to 'home' after success

    def form_valid(self, form):
        # Debugging: Print user information
        print(f"User: {self.request.user.username}")

        # Update the `force_password_change` flag
        if hasattr(self.request.user, 'userprofile'):
            user_profile = self.request.user.userprofile
            print(f"Before Update - force_password_change: {user_profile.force_password_change}")

            user_profile.force_password_change = False
            user_profile.save()

            print(f"After Update - force_password_change: {user_profile.force_password_change}")
        else:
            print("UserProfile does not exist for this user.")

        # Call the base form_valid method and redirect to success_url
        return super().form_valid(form)
@login_required
def dashboard(request, pk, id):  # notez le paramètre id supplémentaire
    if request.user.is_authenticated :
        journee = Journee.objects.get(id=id)
        livreur = Livreur.objects.get(user_id= pk)
        key = settings.GOOGLE_API_KEY
        userid = livreur.id
        today = now().date()
        livraisonss = Livraison.objects.filter( statut__livreur = userid, date = journee.date) \
        .order_by('statut', 'position')
        livraisons  = Livraison.objects.order_by('statut', 'position').filter( statut__livreur = userid, date = journee.date)
        livraisonstatusok = Livraison.objects.filter(status=True, recuperation=False, statut__livreur = userid, date = journee.date)
        livraisonstatusko = Livraison.objects.filter(status=False, recuperation=False, statut__livreur = userid, date = journee.date)
        recuperation = Livraison.objects.filter(recuperation=True,  statut__livreur = userid, date = journee.date)
        recuperationok = Livraison.objects.filter(recuperation=True, status=True,  statut__livreur = userid, date = journee.date)
        recuperationko = Livraison.objects.filter(status=False,recuperation=True, statut__livreur = userid, date = journee.date)
        livraison = Livraison.objects.filter(recuperation=False, statut__livreur = userid, date = journee.date)
        tacheok = Tacheafaire.objects.filter(livreur = userid, status=True, date = today )
        tacheko =  Tacheafaire.objects.filter(livreur = userid, status=False, date = today)
        recuperation = "oui"
        recuperationo = "non"
        retourtraiteur = "oui"
        taches = Tacheafaire.objects.filter(livreur = userid)
        routes = Livraison.objects.order_by('statut', 'position')
        routess = Route.objects.filter(date = journee.date, livreur = livreur)
        vehicles = Vehicle.objects.filter(routes__in=routess).distinct()  # Use distinct() to prevent duplicates
        routes_with_livraisons = routess.prefetch_related('livreur', 'livraisons').order_by('heure_depart')
        for route in routes_with_livraisons:
            # This will give you the livraisons associated with the current route
            ordered_livraisons = route.livraisons.all().order_by('statut', 'position')

            # Store these ordered livraisons in a convenient way, e.g., in the route object itself
            route.ordered_livraisons = ordered_livraisons

        if request.method == 'POST':
            if 'route_vehicule' in request.POST:
                route_id = request.POST.get('route_vehicule')
                route_instance = Route.objects.filter(id=route_id).first()
                
                if route_instance:
                    # Récupérer ou créer le véhicule
                    vehicle_name = request.POST.get('vehicle_name')
                    vehicle_commentaire = request.POST.get('vehicle_commentaire', '')
                    
                    # Créer un nouveau véhicule ou récupérer un existant
                    vehicle, created = Vehicle.objects.get_or_create(
                        name=vehicle_name,
                        defaults={'commentaire': vehicle_commentaire}
                    )
                    
                    # Si le véhicule existe déjà, mettre à jour le commentaire si nécessaire
                    if not created and vehicle_commentaire:
                        vehicle.commentaire = vehicle_commentaire
                        vehicle.save()
                    
                    # Ajouter le véhicule à la route
                    route_instance.vehicles.add(vehicle)
                    
                    # Récupérer la légende pour les photos/vidéos
                    caption = request.POST.get('photo_caption', '')
                    
                    # Gérer les photos
                    for photo in request.FILES.getlist('vehicle_photos'):
                        PhotoVehicle.objects.create(
                            vehicle=vehicle,
                            image=photo,
                            caption=caption
                        )
                    
                    # Gérer les vidéos
                    for video in request.FILES.getlist('vehicle_videos'):
                        PhotoVehicle.objects.create(
                            vehicle=vehicle,
                            video=video,
                            caption=caption
                        )
                
                return redirect('dashboard', pk=pk, id=id)


        livraisons_data = [
        {
            'address': livraison.adress,
            'latitude': float(livraison.lat),
            'longitude': float(livraison.lng),
            'nom': livraison.nom,
            'mode_envoi': livraison.mode_envoi,
            'convive': livraison.convives,
            'heure_livraison': livraison.heure_livraison,
        }
        for livraison in livraisonss

    ]





        return render(request, "listings/dashboard.html", context={'livreur':livreur,
                                                                   'livraisons':livraisons,
                                                                'livraisonss' : livraisons_data,
                                                                'livraisonstatusok':livraisonstatusok,
                                                                'livraisonstatusko':livraisonstatusko,
                                                                'recuperation' : recuperation,
                                                                'journee' : journee,
                                                                'recuperationok':recuperationok,
                                                                'recuperationko':recuperationko,
                                                                'livraison':livraison,
                                                                'recuperation':recuperation,
                                                                'key':key,
                                                                'userid':userid,
                                                                'recuperationo':recuperationo,
                                                                'retourtraiteur':retourtraiteur,
                                                                'routes':routes,
                                                                'today':today,
                                                                'routess':routess,
                                                                'routes_with_livraisons': routes_with_livraisons,
                                                                'taches':taches,
                                                                'tacheok':tacheok,
                                                                'tacheko':tacheko,
                                                                'vehicles': vehicles,




                                                                })
    else:
        return redirect('home')



@login_required
def delete_vehicle(request, vehicle_id):
    vehicle = get_object_or_404(Vehicle, id=vehicle_id)  # Retrieve vehicle
    vehicle.delete()  # Delete the vehicle instance
    messages.success(request, "Vehicle deleted successfully.")  # Set success message

    # Redirect back to the page that made the request
    referer = request.META.get('HTTP_REFERER', '/')  # Get the Referer header or fallback to home
    return HttpResponseRedirect(referer)  # Redirect to the referer


@login_required
def update_task(request, pk):

    task = get_object_or_404(Tacheafaire, pk=pk)

    if request.method == 'POST':
        form = TaskUpdateForm(request.POST, request.FILES)
        if form.is_valid():
            # Update the task status
            task.status = form.cleaned_data['status']
            task.save()


    return render(request, 'listings/update_task.html', {'task': task})


from django.http import JsonResponse
from django.db.models import Q
from django.contrib.auth.decorators import login_required
from datetime import datetime

@login_required
def api_checklist_items_by_date(request):
    """
    API pour récupérer TOUS les ChecklistItems d'une date donnée
    avec le détail par checklist
    """
    date_str = request.GET.get('date')
    
    if not date_str:
        return JsonResponse({'error': 'Date requise'}, status=400)
    
    try:
        selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'error': 'Format de date invalide'}, status=400)
    
    # Récupérer tous les ChecklistItems pour cette date
    checklist_items = ChecklistItem.objects.filter(
        checklist__date=selected_date,
        checklist__is_active=True
    ).select_related('product', 'checklist').order_by('checklist__name', 'product__name')
    
    items_data = []
    for item in checklist_items:
        items_data.append({
            'checklist_id': item.checklist.id,
            'checklist_name': item.checklist.name,
            'product_id': item.product.id if item.product else None,
            'product_name': item.product.name if item.product else 'Produit inconnu',
            'quantity': item.quantity,
            'status': item.status,
        })
    
    return JsonResponse({
        'items': items_data,
        'count': len(items_data),
        'date': selected_date.isoformat()
    })


@login_required
def api_checklist_items_encours(request):
    """
    API pour récupérer uniquement les ChecklistItems 
    avec status en_cours ou pending pour une date donnée
    """
    date_str = request.GET.get('date')
    status_filter = request.GET.get('status', 'en_cours,pending')
    
    if not date_str:
        return JsonResponse({'error': 'Date requise'}, status=400)
    
    try:
        selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'error': 'Format de date invalide'}, status=400)
    
    # Parser les statuts
    statuses = [s.strip() for s in status_filter.split(',')]
    
    # Construire la requête avec filtre sur les statuts
    query = Q()
    for status in statuses:
        query |= Q(status=status)
    
    # Récupérer uniquement les items en_cours ou pending
    checklist_items = ChecklistItem.objects.filter(
        checklist__date=selected_date,
        checklist__is_active=True
    ).filter(query).select_related('product', 'checklist').order_by('checklist__name', 'product__name')
    
    items_data = []
    for item in checklist_items:
        items_data.append({
            'checklist_id': item.checklist.id,
            'checklist_name': item.checklist.name,
            'product_id': item.product.id if item.product else None,
            'product_name': item.product.name if item.product else 'Produit inconnu',
            'quantity': item.quantity,
            'status': item.status,
        })
    
    return JsonResponse({
        'items': items_data,
        'count': len(items_data),
        'date': selected_date.isoformat(),
        'filtered_statuses': statuses
    })

from django.http import JsonResponse
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_http_methods
from .models import Product, Category
from decimal import Decimal
import json

@login_required
def create_product(request):
    if request.method == 'POST':
        try:
            name = request.POST.get('name')
            quantity = request.POST.get('quantity', '0.00')
            category_ids = request.POST.getlist('category')
            
            # Create the product
            product = Product.objects.create(
                name=name,
                quantity=Decimal(quantity),
                created_by=request.user
            )
            
            # Add categories
            if category_ids:
                product.category.set(category_ids)
            
            messages.success(request, 'Produit créé avec succès!')
            return redirect('creerchecklist')  # Rediriger vers votre page actuelle
                
        except Exception as e:
            messages.error(request, f'Erreur: {str(e)}')
            return redirect('creerchecklist')
    
    return redirect('creerchecklist')

@login_required
def update_photo_task(request, pk):
    task = get_object_or_404(Tacheafaire, pk=pk)

    if request.method == 'POST':
       photo_formset = PhotoTachesFormSet(request.POST, request.FILES)
       if photo_formset.is_valid():
        photos = photo_formset.save(commit=False)
        for photo in photos:
                photo.task = task  # Set the related Livraison for each photo
                photo.save()
    else:
        form = PhotoTachesForm()

    return render(request, 'listings/update_photo_task.html', {'form': form, 'task': task})

# Admin view for creating shifts


from django.shortcuts import render
from django.db.models import Sum, Count
from django.utils import timezone
from .models import Livraison, Checklist, ChecklistItem
from django.shortcuts import render
from django.db.models import Count, Sum, Q
from django.utils import timezone
from .models import Livraison, Checklist, ChecklistItem, Route




@login_required
def create_shift(request):
    liste_livreur = Livreur.objects.all()

    if request.method == 'POST':
        # Si c’est une soumission pour créer un nouveau planning
        if 'create_planning' in request.POST:
            global_shift_date = request.POST.get("global_shift_date")
            if not global_shift_date:
                return render(request, 'listings/create_shift.html', {
                    'liste_livreur': liste_livreur,
                    'error': 'La date est requise.',
                })
            for livreur in liste_livreur:
                is_repos = f"repos_{livreur.id}" in request.POST
                selected_start_time = request.POST.get(f"shift_start_{livreur.id}")
                custom_start_time = request.POST.get(f"custom_start_time_{livreur.id}")
                shift_start = custom_start_time or selected_start_time or None

                if shift_start:
                    try:
                        datetime.strptime(shift_start, '%H:%M')
                    except ValueError:
                        return render(request, 'listings/create_shift.html', {
                            'liste_livreur': liste_livreur,
                            'error': 'Format d\'heure invalide.',
                        })

                if is_repos:
                    Shift.objects.create(
                        livreur=livreur,
                        date=global_shift_date,
                        start_time=None,
                        notes="Repos"
                    )
                else:
                    Shift.objects.create(
                        livreur=livreur,
                        date=global_shift_date,
                        start_time=shift_start,
                        notes=""
                    )
            return redirect('create-shift')

        # Si c’est une requête de modification d’un shift existant
        elif 'edit_shift' in request.POST:
            shift_id = request.POST.get('shift_id')
            shift = get_object_or_404(Shift, id=shift_id)
            start_time = request.POST.get('start_time')
            notes = request.POST.get('notes', '')

            if start_time:
                try:
                    datetime.strptime(start_time, '%H:%M')
                    shift.start_time = start_time
                except ValueError:
                    pass  # ou gérer l'erreur
            shift.notes = notes
            shift.save()

            return redirect(request.path + '?date=' + str(shift.date))
    
    # Si GET, on affiche la page avec la liste des shifts
    selected_date = request.GET.get('date', str(date.today()))
    shifts = Shift.objects.filter(date=selected_date)
    if not request.user.is_superuser:
        return redirect('unauthorized')
    return render(request, 'listings/create_shift.html', {
        'liste_livreur': liste_livreur,
        'shifts': shifts,
        'date': selected_date,
        'error': None,
    })



@login_required
def responsableschoixjournee(request):

    if request.method == 'POST':
        livraison_resource = LivraisonResource()
        dataset = Dataset()
        new_livraisons = request.FILES['livraisons_file']
        imported_data = dataset.load(new_livraisons.read(), format='xlsx')
        for data in imported_data:
            value = Livraison(
                data[0],
                data[1],
                data[2],
                data[3],
                data[4],
                data[5],
                data[6],
                data[7],
                data[8],
                data[9],
                data[10],
                data[11],
                data[12],
                data[13],
                data[14],
                data[15],
                data[16],
            )
            value .save()

    livraisons  = Livraison.objects.all()
    livreurs = Livreur.objects.all()
    journees = Journee.objects.all()
    if not request.user.is_superuser:
        return redirect('unauthorized')
    return render(request, 'listings/responsableschoixjournee.html', context={'livraisons': livraisons,
                                                              'livreurs': livreurs,
                                                              'journees' : journees,
                                                              })
@login_required
def responsables(request, id):
    today = datetime.now().date()
    journee = Journee.objects.get(id=id)
    tomorrow = today + timedelta(1)
    livraisons = Livraison.objects.order_by('statut', 'position').select_related('statut')
    livraisonstatusok = Livraison.objects.filter(status=True,recuperation=False, date=journee.date)
    livraisonstatusko = Livraison.objects.filter(status=False,recuperation=False, date=journee.date)
    recuperation = Livraison.objects.filter(recuperation=True, date=journee.date)
    recuperationok = Livraison.objects.filter(recuperation=True, status=True, date=journee.date)
    recuperationko = Livraison.objects.filter(status=False,recuperation=True, date=journee.date)
    livraison = Livraison.objects.filter(recuperation=False, date=journee.date)
    livreurs = Livreur.objects.all()
    routess = Route.objects.filter(date = journee.date)
    key = settings.API_KEY_JLT
    vehicles = Vehicle.objects.filter(routes__in=routess).distinct()  # Use distinct() to prevent duplicates
    recuperations = "oui"
    if not request.user.is_superuser:
        return redirect('unauthorized')


    return render(request, 'listings/responsables.html', context={
                                                              'livraisons': livraisons,
                                                              'livreurs': livreurs,
                                                              'journee' : journee,
                                                              'recuperations' : recuperations,
                                                              'livraisonstatusok':livraisonstatusok,
                                                              'livraisonstatusko':livraisonstatusko,
                                                              'livraison':livraison,
                                                              'recuperationok':recuperationok,
                                                              'recuperationko':recuperationko,
                                                              'recuperation' : recuperation,
                                                              'vehicles' : vehicles,
                                                              'key':key,
                                                              })

@login_required
def unauthorized_view(request):
    return render(request, 'listings/pasauthorise.html', status=403)

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
from .models import Product

@csrf_exempt
def product_inline_update(request):
    if request.method == "POST":
        data = json.loads(request.body)
        product_id = data.get("id")
        quantity = data.get("quantity")

        try:
            product = Product.objects.get(id=product_id)
            product.quantity = quantity
            product.save()
            return JsonResponse({"success": True})
        except Product.DoesNotExist:
            return JsonResponse({"success": False, "error": "Produit introuvable"})
    return JsonResponse({"success": False, "error": "Methode non autorisée"})


from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
@login_required
def product_list(request):
    # Get the search query (if any)
    query = request.GET.get('query', '').strip()

    # Filter products based on the search query
    products = Product.objects.filter(name__icontains=query) if query else Product.objects.all()

    # Paginate the filtered products (10 products per page)
    paginator = Paginator(products, 12)
    page_number = request.GET.get('page')

    try:
        page_obj = paginator.get_page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.get_page(1)
    except EmptyPage:
        page_obj = paginator.get_page(paginator.num_pages)

    # Check user group memberships
    user_groups = request.user.groups.values_list('name', flat=True)
    product_form = ProductsForm

    if request.method == 'POST':
        if 'product-form' in request.POST:
            product_form = ProductsForm(request.POST, user=request.user)
            if product_form.is_valid():
                product = product_form.save()

                # Log the product creation
                ProductLog.objects.create(product=product, created_by=request.user)

                messages.success(request, 'Nouveau produit créé avec succès.')
                return HttpResponseRedirect(reverse('product_list'))
    context = {
        'page_obj': page_obj,
        'query': query,
        'is_ventes': 'ventes' in user_groups,
        'is_checklist': 'checklist' in user_groups,
        'is_admin': 'admin' in user_groups,
        'is_paginated': paginator.num_pages > 1,
        'product_form': product_form,
    }

    return render(request, 'listings/product_list.html', context)




class DistanceView(View):

    def get(self, request):
        form = DistanceForm
        distances = Distances.objects.all()
        context={'form': form,
                 'distances':distances,
                                                                    }
        return render(request, 'listings/distances.html', context)

    def post(self, request):
        form = DistanceForm(request.POST)
        if form.is_valid():
            from_location = form.cleaned_data['from_location']
            from_location_info = Livraison.objects.get(nom=from_location)
            from_adress_string = str(from_location_info.adress)+", "+str(from_location_info.zipcode)+", "+str(from_location_info.city)+", "+str(from_location_info.country)

            to_location = form.cleaned_data['to_location']
            to_location_info = Livraison.objects.get(nom=to_location)
            to_adress_string = str(to_location_info.adress)+", "+str(to_location_info.zipcode)+", "+str(to_location_info.city)+", "+str(to_location_info.country)

            mode = form.cleaned_data['mode']
            now = datetime.now()

            gmaps = googlemaps.Client(key= settings.GOOGLE_API_KEY)
            calculate = gmaps.distance_matrix(
                from_adress_string,
                to_adress_string,
                mode = mode,
                departure_time = now
            )
            print(calculate)

            duration_secons = calculate['rows'][0]['elements'][0]['duration']['value']
            duration_minutes = duration_secons/60

            distance_meters = calculate['rows'][0]['elements'][0]['distance']['value']
            distance_km = distance_meters/1000

            if 'duration_in_traffic' in calculate['rows'][0]['elements'][0]:
                duration_in_traffic_seconds = calculate['rows'][0]['elements'][0]['duration_in_traffic']['value']
                duration_in_traffic_minutes = duration_in_traffic_seconds/60
            else:
                duration_in_traffic_minutes = None

            obj = Distances(
                from_location = Livraison.objects.get(nom=from_location),
                to_location = Livraison.objects.get(nom=to_location),
                mode = mode,
                distance_km = distance_km,
                distance_mins = duration_minutes,
                distance_traffic_mins = duration_in_traffic_minutes
            )

            obj.save()

        return redirect('my_distance_view')


from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
@login_required
@csrf_exempt  # Remove this in production; use CSRF token instead
def update_livraison_route(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)  # Try parsing JSON data

            livraison_id = data.get("livraison_id")
            route_id = data.get("route_id")

            if not livraison_id or not route_id:
                return JsonResponse({"success": False, "error": "Missing data"}, status=400)

            # Assuming you have a Livraison model
            livraison = Livraison.objects.get(id=livraison_id)
            livraison.route_id = route_id  # Update route
            livraison.save()

            return JsonResponse({"success": True})
        except json.JSONDecodeError:
            return JsonResponse({"success": False, "error": "Invalid JSON"}, status=400)
        except Livraison.DoesNotExist:
            return JsonResponse({"success": False, "error": "Livraison not found"}, status=404)
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)}, status=500)

    return JsonResponse({"success": False, "error": "Invalid request"}, status=405)


@login_required
@csrf_exempt
def update_livraison(request):
    if request.method == 'POST':
        livraison_id = request.POST.get('livraison_id')
        new_route_id = request.POST.get('new_route_id')

        try:
            livraison = Livraison.objects.get(id=livraison_id)
            new_route = Route.objects.get(id=new_route_id)

            # Update route and possibly other fields
            livraison.statut = new_route

            # Optionally update journee if needed
            livraison.journee = new_route.journee  # Assuming you want to update journee based on the new route

            livraison.save()

            return JsonResponse({'success': True})
        except Livraison.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Livraison not found'})
        except Route.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Route not found'})
    return JsonResponse({'success': False, 'error': 'Invalid request'})

@login_required
def update_status(request):
    if request.method == 'POST' and request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        task_id = request.POST.get('livraison_id')
        new_status = request.POST.get('new_statut_id')
        task = Livraison.objects.get(id=task_id)
        task.statut_id = new_status
        task.save()
        return JsonResponse({'message': 'Route mise à jour'})
    else:
        return JsonResponse({'message': 'Echec'})


from django.http import JsonResponse
from .models import Livraison  # Adjust the import based on your models
@login_required
def get_route_livraisons(request, route_id):
    livraisons = Livraison.objects.filter(statut_id=route_id).order_by("position")
    data = [{"id": l.id, "nom": l.nom, "heure_livraison": l.heure_livraison, "mode_envoi": l.mode_envoi} for l in livraisons]
    return JsonResponse({"livraisons": data})


import json
from django.http import JsonResponse
from .models import Livraison
from django.views.decorators.csrf import csrf_exempt
@login_required
@csrf_exempt
def update_livraison_positions(request):
    if request.method == "POST":
        data = json.loads(request.body)
        positions = data.get("positions", [])

        for item in positions:
            Livraison.objects.filter(id=item["id"]).update(position=item["position"])

        return JsonResponse({"success": True})

    return JsonResponse({"error": "Invalid request"}, status=400)



class MapAujourView(View):

    def post(self, request):
        form = RoutedetailForm(request.POST)
        if form.is_valid():
            route_instance = form.save(commit=False)  # Create without saving to the database
            route_instance.save()  # Save the instance
            form.save_m2m()  # Save many-to-many relationships
            messages.success(request, 'Route detail saved successfully!')
            return redirect('your_redirect_view_name')  # Redirect appropriately
        else:
            messages.error(request, 'Please correct the errors below.')
            return self.get(request)  # Re-render the form in case of errors

    def get(self, request):
        key = settings.GOOGLE_API_KEY
        date_str = request.GET.get('date', datetime.now().date().strftime('%Y-%m-%d'))
        selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()

        matin = ['05h00', '05h15', '05h30', '05h45', '06h00', '06h15', '06h30', '06h45',
                 '07h00', '07h15', '07h30', '07h45', '08h00', '08h15', '08h30', '08h45',
                 '09h00', '09h15', '09h30', '09h45', 'recup']

        form = RoutedetailForm()

        todo_livraison = Livraison.objects.filter(date=selected_date, heure_livraison__in=matin, place_id__isnull=False, statut__id=21)
        routes21 = Route.objects.filter(id=21)
        routes = Route.objects.filter(date=selected_date)
        route1 = Livraison.objects.filter(date=selected_date, heure_livraison__in=matin).order_by('statut', 'position')
        eligable_locations = Livraison.objects.order_by('statut', 'position').filter(place_id__isnull=False, heure_livraison__in=matin, date=selected_date)

        livraisons = []

        for a in eligable_locations:
            data = {
                'lat': float(a.lat),
                'lng': float(a.lng),
                'place_id': a.place_id,
                'nom': a.nom,
                'heure_livraison': a.heure_livraison,
                'adress': a.adress,
                'convives': a.convives,
                'mode_envoi': a.mode_envoi,
            }
            livraisons.append(data)

        context = {
            'key': key,
            'livraisons': livraisons,
            'selected_date': selected_date.strftime('%Y-%m-%d'),
            'routes': routes,
            'todo_livraison': todo_livraison,
            'routes21': routes21,
            'route1': route1,
            'routedetail_form': form,
        }

        return render(request, 'listings/mapmatinaujour.html', context)

class MapApremAujourView(View):
    def get(self, request):
        key = settings.GOOGLE_API_KEY
        date_str = request.GET.get('date', datetime.now().date().strftime('%Y-%m-%d'))
        selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        routes = Route.objects.filter(date=selected_date)


        aprem = ['13h00', '13h15', '13h30', '13h45', '14h00', '14h15', '14h30', '14h45', '15h00', '15h15', '15h30', '15h45', '16h00', '16h15', '16h30', '16h45', '17h00', '17h15', '17h30', '17h45', '18h00', '18h15', '18h30', '18h45', '19h00', 'recup']

        midi = ['05h00', '05h15', '05h30', '05h45', '06h00', '06h15', '06h30', '06h45', '07h00', '07h15', '07h30', '07h45', '08h00', '08h15', '08h30', '08h45', '09h00', '09h15', '09h30', '09h45', '10h00', '10h15', '10h30', '10h45', '11h00', '11h15', '11h30', '11h45', '12h00', '12h15', '12h30', '12h45']
        routes_to_hide = Route.objects.filter(
                livraisons__heure_livraison__in=midi
        )

        # Exclude routes to hide
        filtered_routes = routes.exclude(id__in=routes_to_hide)

        form = RoutedetailForm()
        todo_livraison = Livraison.objects.filter(date=selected_date, heure_livraison__in = aprem, place_id__isnull=False, statut__id= 21)
        routes21 = Route.objects.filter(id=21)
        route1 = Livraison.objects.filter(date=selected_date, heure_livraison__in = aprem).order_by('statut', 'position')
        eligable_locations = Livraison.objects.order_by('statut', 'position').filter(place_id__isnull=False, heure_livraison__in = aprem, date=selected_date)
        livraisons =[]

        for a in eligable_locations:
            data = {
                'lat': float(a.lat),
                'lng': float(a.lng),
                'place_id': a.place_id,
                'nom': a.nom,
                'heure_livraison': a.heure_livraison,
                'adress' : a.adress,
                'convives': a.convives,
                'mode_envoi': a.mode_envoi,



            }

            livraisons.append(data)

        context = {
            'key': key,
            'livraisons': livraisons,
            'selected_date': selected_date.strftime('%Y-%m-%d'),
            'routes': filtered_routes,  # Pass routes to the context
            'todo_livraison':todo_livraison,
            'routes21':routes21,
            'route1': route1,
            'routedetail_form': form,


        }
        return render(request, 'listings/mapapremaujour.html', context)


class MapMidiAujourView(View):
    def get(self, request):
        key = settings.GOOGLE_API_KEY
        date_str = request.GET.get('date', datetime.now().date().strftime('%Y-%m-%d'))
        selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        routes = Route.objects.filter(date=selected_date)


        midi = ['10h00', '10h15', '10h30', '10h45', '11h00', '11h15', '11h30', '11h45', '12h00', '12h15', '12h30', '12h45', 'recup']

        matin = ['05h00', '05h15', '05h30', '05h45', '06h00', '06h15', '06h30', '06h45', '07h00', '07h15', '07h30', '07h45', '08h00', '08h15', '08h30', '08h45', '09h00', '09h15', '09h30', '09h45']
        routes_to_hide = Route.objects.filter(
                livraisons__heure_livraison__in=matin
        )

        # Exclude routes to hide
        filtered_routes = routes.exclude(id__in=routes_to_hide)
        form = RoutedetailForm()
        todo_livraison = Livraison.objects.filter(date=selected_date, heure_livraison__in = midi, place_id__isnull=False, statut__id= 21)
        routes21 = Route.objects.filter(id=21)
        route1 = Livraison.objects.filter(date=selected_date, heure_livraison__in = midi).order_by('statut', 'position')
        eligable_locations = Livraison.objects.filter(place_id__isnull=False, heure_livraison__in = midi, date=selected_date)
        livraisons =[]

        for a in eligable_locations:
            data = {
                'lat': float(a.lat),
                'lng': float(a.lng),
                'place_id': a.place_id,
                'nom': a.nom,
                'heure_livraison': a.heure_livraison,
                'adress' : a.adress,
                'convives': a.convives,
                'mode_envoi': a.mode_envoi,



            }

            livraisons.append(data)

        context = {
            'key': key,
            'livraisons': livraisons,
            'selected_date': selected_date.strftime('%Y-%m-%d'),
            'routes': filtered_routes,  # Pass routes to the context
            'todo_livraison':todo_livraison,
            'routes21':routes21,
            'route1': route1,
            'routedetail_form': form,

        }
        return render(request, 'listings/mapmidiaujour.html', context)

from django.db.models.functions import Cast
from django.db.models import IntegerField

class CreateRouteAujourView(View):
    def post(self, request):
        # Get the selected date from the POST request, defaults to today
        selected_date_str = request.POST.get('selected_date', datetime.now().date().strftime('%Y-%m-%d'))
        selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()

        # Get the highest `nom` for routes already created on that date
        existing_routes = Route.objects.filter(date=selected_date).annotate(
            nom_as_int=Cast('nom', IntegerField())
        ).order_by('-nom_as_int')

        if existing_routes.exists():
            last_nom = int(existing_routes.first().nom)
        else:
            last_nom = 0  # Start with '1' if no routes exist for the date

        # Increment `nom`
        new_nom = str(last_nom + 1)

        # Create the new route with the incremented `nom` and the selected date
        Route.objects.create(date=selected_date, nom=new_nom)

        # Redirect back to the referring page
        return redirect(request.META.get('HTTP_REFERER', '/'))

class CreateRouteView(View):
    def post(self, request):
        # Get today's date
        today = datetime.now().date()
        tomorrow = today + timedelta(1)

        # Get the highest `nom` for routes created today
        existing_routes = Route.objects.filter(date=tomorrow).annotate(
            nom_as_int=Cast('nom', IntegerField())  # Cast `nom` to an integer
        ).order_by('-nom_as_int')

        if existing_routes.exists():
            last_nom = int(existing_routes.first().nom)
        else:
            last_nom = 0  # Start with '1' if no routes exist for today

        # Increment `nom`
        new_nom = str(last_nom + 1)

        # Create the new route with incremented `nom`
        Route.objects.create(date=tomorrow, nom=new_nom)

        # Redirect back to the referring page
        return redirect(request.META.get('HTTP_REFERER', '/'))
class CreateRouteTodayView(View):
    def post(self, request):
        # Get today's date
        today = datetime.now().date()
        tomorrow = today + timedelta(2)

        # Get the highest `nom` for routes created today
        existing_routes = Route.objects.filter(date=tomorrow).annotate(
            nom_as_int=Cast('nom', IntegerField())  # Cast `nom` to an integer
        ).order_by('-nom_as_int')

        if existing_routes.exists():
            last_nom = int(existing_routes.first().nom)
        else:
            last_nom = 0  # Start with '1' if no routes exist for today

        # Increment `nom`
        new_nom = str(last_nom + 1)

        # Create the new route with incremented `nom`
        Route.objects.create(date=tomorrow, nom=new_nom)

        # Redirect back to the referring page
        return redirect(request.META.get('HTTP_REFERER', '/'))
class CreateRouteDimView(View):
    def post(self, request):
        # Get today's date
        today = datetime.now().date()
        tomorrow = today + timedelta(3)

        # Get the highest `nom` for routes created today
        existing_routes = Route.objects.filter(date=tomorrow).annotate(
            nom_as_int=Cast('nom', IntegerField())  # Cast `nom` to an integer
        ).order_by('-nom_as_int')

        if existing_routes.exists():
            last_nom = int(existing_routes.first().nom)
        else:
            last_nom = 0  # Start with '1' if no routes exist for today

        # Increment `nom`
        new_nom = str(last_nom + 1)

        # Create the new route with incremented `nom`
        Route.objects.create(date=tomorrow, nom=new_nom)

        # Redirect back to the referring page
        return redirect(request.META.get('HTTP_REFERER', '/'))


class MapView(View):
    def get(self, request):
        key = settings.GOOGLE_API_KEY
        form = DistanceForm
        today = datetime.now().date()
        tomorrow = today + timedelta(1)
        matin = ['05h00', '05h15', '05h30', '05h45', '06h00', '06h15', '06h30', '06h45', '07h00', '07h15', '07h30', '07h45', '08h00','08h15', '08h30', '08h45', '09h00', '09h15', '09h30', '09h45', 'recup']
        distances = Distances.objects.all()
        todo_livraison = Livraison.objects.filter(date=tomorrow, heure_livraison__in = matin, place_id__isnull=False, statut__id= 21)
        route1 = Livraison.objects.filter(date=tomorrow, heure_livraison__in = matin)
        recups = Livraison.objects.filter(date=tomorrow, recuperation = True)
        route2 = Livraison.objects.filter(statut__id= 2, date=tomorrow, heure_livraison__in = matin)
        route3 = Livraison.objects.filter(statut__id= 3, date=tomorrow, heure_livraison__in = matin)
        route4 = Livraison.objects.filter(statut__id= 4, date=tomorrow, heure_livraison__in = matin)
        routes21 = Route.objects.filter(id=21)
        routes = Route.objects.filter(date=tomorrow)

        eligable_locations = Livraison.objects.order_by('statut', 'position').filter(place_id__isnull=False, heure_livraison__in = matin, date=tomorrow)
        livraisons =[]

        for a in eligable_locations:
            data = {
                'lat': float(a.lat),
                'lng': float(a.lng),
                'place_id': a.place_id,
                'nom': a.nom,
                'heure_livraison': a.heure_livraison,
                'adress' : a.adress,
                'convives': a.convives,
                'mode_envoi': a.mode_envoi,

            }

            livraisons.append(data)

        context = {'key': key,
                   'livraisons':livraisons,
                   'form': form,
                   'distances':distances,
                   'routes':routes,
                   'route1':route1,
                   'route2':route2,
                   'route3':route3,
                   'route4':route4,
                   'recups': recups,
                   'todo_livraison':todo_livraison,
                   'routes21':routes21,
                   'tomorrow':tomorrow,

        }
        return render(request, 'listings/map.html', context)

    def post(self, request):
        form = DistanceForm(request.POST)
        if form.is_valid():
            from_location = form.cleaned_data['from_location']
            from_location_info = Livraison.objects.get(nom=from_location)
            from_adress_string = str(from_location_info.adress)+", "+str(from_location_info.zipcode)+", "+str(from_location_info.city)+", "+str(from_location_info.country)

            to_location = form.cleaned_data['to_location']
            to_location_info = Livraison.objects.get(nom=to_location)
            to_adress_string = str(to_location_info.adress)+", "+str(to_location_info.zipcode)+", "+str(to_location_info.city)+", "+str(to_location_info.country)

            mode = form.cleaned_data['mode']
            now = datetime.now()

            gmaps = googlemaps.Client(key= settings.GOOGLE_API_KEY)
            calculate = gmaps.distance_matrix(
                from_adress_string,
                to_adress_string,
                mode = mode,
                departure_time = now
            )
            print(calculate)

            duration_secons = calculate['rows'][0]['elements'][0]['duration']['value']
            duration_minutes = duration_secons/60

            distance_meters = calculate['rows'][0]['elements'][0]['distance']['value']
            distance_km = distance_meters/1000

            if 'duration_in_traffic' in calculate['rows'][0]['elements'][0]:
                duration_in_traffic_seconds = calculate['rows'][0]['elements'][0]['duration_in_traffic']['value']
                duration_in_traffic_minutes = duration_in_traffic_seconds/60
            else:
                duration_in_traffic_minutes = None

            obj = Distances(
                from_location = Livraison.objects.get(nom=from_location),
                to_location = Livraison.objects.get(nom=to_location),
                mode = mode,
                distance_km = distance_km,
                distance_mins = duration_minutes,
                distance_traffic_mins = duration_in_traffic_minutes
            )

            obj.save()

        return redirect('my_map_view')
class MapApremView(View):
    def get(self, request):
        key = settings.GOOGLE_API_KEY
        form = DistanceForm
        today = datetime.now().date()
        tomorrow = today + timedelta(1)
        distances = Distances.objects.all()
        aprem = ['13h00', '13h15', '13h30', '13h45', '14h00', '14h15', '14h30', '14h45', '15h00', '15h15', '15h30', '15h45', '16h00', '16h15', '16h30', '16h45', '17h00', '17h15', '17h30', '17h45', '18h00', '18h15', '18h30', '18h45', '19h00', 'recup']
        midi = ['05h00', '05h15', '05h30', '05h45', '06h00', '06h15', '06h30', '06h45', '07h00', '07h15', '07h30', '07h45', '08h00', '08h15', '08h30', '08h45', '09h00', '09h15', '09h30', '09h45', '10h00', '10h15', '10h30', '10h45', '11h00', '11h15', '11h30', '11h45', '12h00', '12h15', '12h30', '12h45']
        routes_to_hide = Route.objects.filter(
                livraisons__heure_livraison__in=midi
        )

        # Exclude routes to hide
        route1 = Livraison.objects.filter(date=tomorrow, heure_livraison__in = aprem)
        route8 = Livraison.objects.filter(date=tomorrow, heure_livraison__in = aprem)
        routes21 = Route.objects.filter(id=21)

        eligable_locations = Livraison.objects.filter(place_id__isnull=False, heure_livraison__in = aprem, date=tomorrow )
        livraisons =[]
        for a in eligable_locations:
            data = {
                'lat': float(a.lat),
                'lng': float(a.lng),
                'place_id': a.place_id,
                'nom': a.nom,
                'heure_livraison': a.heure_livraison,
                'adress' : a.adress,
                'convives': a.convives,
                'mode_envoi': a.mode_envoi,
            }

            livraisons.append(data)

        context = {'key': key,
                   'livraisons':livraisons,
                   'form': form,
                   'distances':distances,
                   'route1':route1,
                   'route8':route8,
                   'routes21': routes21,


        }
        return render(request, 'listings/mapaprem.html', context)

    def post(self, request):
        form = DistanceForm(request.POST)
        if form.is_valid():
            from_location = form.cleaned_data['from_location']
            from_location_info = Livraison.objects.get(nom=from_location)
            from_adress_string = str(from_location_info.adress)+", "+str(from_location_info.zipcode)+", "+str(from_location_info.city)+", "+str(from_location_info.country)

            to_location = form.cleaned_data['to_location']
            to_location_info = Livraison.objects.get(nom=to_location)
            to_adress_string = str(to_location_info.adress)+", "+str(to_location_info.zipcode)+", "+str(to_location_info.city)+", "+str(to_location_info.country)

            mode = form.cleaned_data['mode']
            now = datetime.now()

            gmaps = googlemaps.Client(key= settings.GOOGLE_API_KEY)
            calculate = gmaps.distance_matrix(
                from_adress_string,
                to_adress_string,
                mode = mode,
                departure_time = now
            )
            print(calculate)

            duration_secons = calculate['rows'][0]['elements'][0]['duration']['value']
            duration_minutes = duration_secons/60

            distance_meters = calculate['rows'][0]['elements'][0]['distance']['value']
            distance_km = distance_meters/1000

            if 'duration_in_traffic' in calculate['rows'][0]['elements'][0]:
                duration_in_traffic_seconds = calculate['rows'][0]['elements'][0]['duration_in_traffic']['value']
                duration_in_traffic_minutes = duration_in_traffic_seconds/60
            else:
                duration_in_traffic_minutes = None

            obj = Distances(
                from_location = Livraison.objects.get(nom=from_location),
                to_location = Livraison.objects.get(nom=to_location),
                mode = mode,
                distance_km = distance_km,
                distance_mins = duration_minutes,
                distance_traffic_mins = duration_in_traffic_minutes
            )

            obj.save()

        return redirect('my_mapaprem_view')
class MapMidiView(View):
    def get(self, request):
        key = settings.GOOGLE_API_KEY
        form = DistanceForm
        today = datetime.now().date()
        tomorrow = today + timedelta(1)
        distances = Distances.objects.all()
        midi = ['10h00', '10h15', '10h30', '10h45', '11h00', '11h15', '11h30', '11h45', '12h00', '12h15', '12h30', '12h45', 'recup']
        matin = ['05h00', '05h15', '05h30', '05h45', '06h00', '06h15', '06h30', '06h45', '07h00', '07h15', '07h30', '07h45', '08h00', '08h15', '08h30', '08h45', '09h00', '09h15', '09h30', '09h45']
        routes_to_hide = Route.objects.filter(
                livraisons__heure_livraison__in=matin
        )

        # Exclude routes to hide
        route1 = Livraison.objects.filter(date=tomorrow, heure_livraison__in=midi)
        route2 = Livraison.objects.filter(date=tomorrow, heure_livraison__in = midi)
        route3 = Livraison.objects.filter(statut='3', date=tomorrow, heure_livraison__in = midi)
        route4 = Livraison.objects.filter(statut='4', date=tomorrow, heure_livraison__in = midi)
        route5 = Livraison.objects.filter(statut='5', date=tomorrow, heure_livraison__in = midi)
        route6 = Livraison.objects.filter(statut='6', date=tomorrow, heure_livraison__in = midi)
        route7 = Livraison.objects.filter(statut='7', date=tomorrow, heure_livraison__in = midi)
        route8 = Livraison.objects.filter(statut='8', date=tomorrow, heure_livraison__in = midi)
        route9 = Livraison.objects.filter(statut='9', date=tomorrow, heure_livraison__in = midi)

        eligable_locations = Livraison.objects.filter(place_id__isnull=False, heure_livraison__in = midi, date=tomorrow)
        livraisons =[]
        for a in eligable_locations:
            data = {
                'lat': float(a.lat),
                'lng': float(a.lng),
                'place_id': a.place_id,
                'nom': a.nom,
                'heure_livraison': a.heure_livraison,
                'adress' : a.adress,
                'convives': a.convives,
                'mode_envoi': a.mode_envoi,
            }

            livraisons.append(data)

        context = {'key': key,
                   'livraisons':livraisons,
                   'form': form,
                   'distances':distances,
                   'route2':route2,
                   'route3':route3,
                   'route4':route4,
                   'route5':route5,
                   'route6':route6,
                   'route7':route7,
                   'route8':route8,
                   'route9':route9,
                   'route1':route1,

        }
        return render(request, 'listings/mapmidi.html', context)

    def post(self, request):
        form = DistanceForm(request.POST)
        if form.is_valid():
            from_location = form.cleaned_data['from_location']
            from_location_info = Livraison.objects.get(nom=from_location)
            from_adress_string = str(from_location_info.adress)+", "+str(from_location_info.zipcode)+", "+str(from_location_info.city)+", "+str(from_location_info.country)

            to_location = form.cleaned_data['to_location']
            to_location_info = Livraison.objects.get(nom=to_location)
            to_adress_string = str(to_location_info.adress)+", "+str(to_location_info.zipcode)+", "+str(to_location_info.city)+", "+str(to_location_info.country)

            mode = form.cleaned_data['mode']
            now = datetime.now()

            gmaps = googlemaps.Client(key= settings.GOOGLE_API_KEY)
            calculate = gmaps.distance_matrix(
                from_adress_string,
                to_adress_string,
                mode = mode,
                departure_time = now
            )
            print(calculate)

            duration_secons = calculate['rows'][0]['elements'][0]['duration']['value']
            duration_minutes = duration_secons/60

            distance_meters = calculate['rows'][0]['elements'][0]['distance']['value']
            distance_km = distance_meters/1000

            if 'duration_in_traffic' in calculate['rows'][0]['elements'][0]:
                duration_in_traffic_seconds = calculate['rows'][0]['elements'][0]['duration_in_traffic']['value']
                duration_in_traffic_minutes = duration_in_traffic_seconds/60
            else:
                duration_in_traffic_minutes = None

            obj = Distances(
                from_location = Livraison.objects.get(nom=from_location),
                to_location = Livraison.objects.get(nom=to_location),
                mode = mode,
                distance_km = distance_km,
                distance_mins = duration_minutes,
                distance_traffic_mins = duration_in_traffic_minutes
            )

            obj.save()

        return redirect('my_mapmidi_view')

class MapApremTodayView(View):
    def get(self, request):
        key = settings.GOOGLE_API_KEY
        form = DistanceForm
        today = datetime.now().date()
        tomorrow = today + timedelta(2)
        distances = Distances.objects.all()
        aprem = ['13h00', '13h15', '13h30', '13h45', '14h00', '14h15', '14h30', '14h45', '15h00', '15h15', '15h30', '15h45', '16h00', '16h15', '16h30', '16h45', '17h00', '17h15', '17h30', '17h45', '18h00', '18h15', '18h30', '18h45', '19h00', 'recup']
        midi = ['05h00', '05h15', '05h30', '05h45', '06h00', '06h15', '06h30', '06h45', '07h00', '07h15', '07h30', '07h45', '08h00', '08h15', '08h30', '08h45', '09h00', '09h15', '09h30', '09h45', '10h00', '10h15', '10h30', '10h45', '11h00', '11h15', '11h30', '11h45', '12h00', '12h15', '12h30', '12h45']
        todo_livraison = Livraison.objects.filter(statut=21, date=tomorrow, heure_livraison__in = aprem, place_id__isnull=False)
        routes = Route.objects.filter(date=tomorrow)
        routes_to_hide = Route.objects.filter(
                livraisons__heure_livraison__in=midi
        )

        # Exclude routes to hide
        filtered_routes = routes.exclude(id__in=routes_to_hide)
        route1 = Livraison.objects.filter(date=tomorrow, heure_livraison__in = aprem)
        route8 = Livraison.objects.filter(date=tomorrow, heure_livraison__in = aprem)
        routes21 = Route.objects.filter(id=21)
        eligable_locations = Livraison.objects.filter(place_id__isnull=False, heure_livraison__in = aprem, date=tomorrow )
        livraisons =[]
        for a in eligable_locations:
            data = {
                'lat': float(a.lat),
                'lng': float(a.lng),
                'place_id': a.place_id,
                'nom': a.nom,
                'heure_livraison': a.heure_livraison,
                'adress' : a.adress,
                'convives': a.convives,
                'mode_envoi': a.mode_envoi,
            }

            livraisons.append(data)

        context = {'key': key,
                   'livraisons':livraisons,
                   'form': form,
                   'distances':distances,
                   'route1':route1,
                   'route8':route8,
                   'routes':filtered_routes,
                   'todo_livraison':todo_livraison,
                   'routes21': routes21,


        }
        return render(request, 'listings/maptodayaprem.html', context)

    def post(self, request):
        form = DistanceForm(request.POST)
        if form.is_valid():
            from_location = form.cleaned_data['from_location']
            from_location_info = Livraison.objects.get(nom=from_location)
            from_adress_string = str(from_location_info.adress)+", "+str(from_location_info.zipcode)+", "+str(from_location_info.city)+", "+str(from_location_info.country)

            to_location = form.cleaned_data['to_location']
            to_location_info = Livraison.objects.get(nom=to_location)
            to_adress_string = str(to_location_info.adress)+", "+str(to_location_info.zipcode)+", "+str(to_location_info.city)+", "+str(to_location_info.country)

            mode = form.cleaned_data['mode']
            now = datetime.now()

            gmaps = googlemaps.Client(key= settings.GOOGLE_API_KEY)
            calculate = gmaps.distance_matrix(
                from_adress_string,
                to_adress_string,
                mode = mode,
                departure_time = now
            )
            print(calculate)

            duration_secons = calculate['rows'][0]['elements'][0]['duration']['value']
            duration_minutes = duration_secons/60

            distance_meters = calculate['rows'][0]['elements'][0]['distance']['value']
            distance_km = distance_meters/1000

            if 'duration_in_traffic' in calculate['rows'][0]['elements'][0]:
                duration_in_traffic_seconds = calculate['rows'][0]['elements'][0]['duration_in_traffic']['value']
                duration_in_traffic_minutes = duration_in_traffic_seconds/60
            else:
                duration_in_traffic_minutes = None

            obj = Distances(
                from_location = Livraison.objects.get(nom=from_location),
                to_location = Livraison.objects.get(nom=to_location),
                mode = mode,
                distance_km = distance_km,
                distance_mins = duration_minutes,
                distance_traffic_mins = duration_in_traffic_minutes
            )

            obj.save()

        return redirect('my_maptodayaprem_view')
class MapMidiTodayView(View):
    def get(self, request):
        key = settings.GOOGLE_API_KEY
        form = DistanceForm
        today = datetime.now().date()
        aftertomorrow = today + timedelta(2)
        distances = Distances.objects.all()
        midi = ['10h00', '10h15', '10h30', '10h45', '11h00', '11h15', '11h30', '11h45', '12h00', '12h15', '12h30', '12h45', 'recup']
        matin = ['05h00', '05h15', '05h30', '05h45', '06h00', '06h15', '06h30', '06h45', '07h00', '07h15', '07h30', '07h45', '08h00', '08h15', '08h30', '08h45', '09h00', '09h15', '09h30', '09h45']
        todo_livraison = Livraison.objects.filter(date=aftertomorrow, heure_livraison__in = midi, place_id__isnull=False, statut__id= 21)
        routes21 = Route.objects.filter(id=21)
        routes = Route.objects.filter(date=aftertomorrow)
        routes_to_hide = Route.objects.filter(
                livraisons__heure_livraison__in=matin
        )

        # Exclude routes to hide
        filtered_routes = routes.exclude(id__in=routes_to_hide)
        route1 = Livraison.objects.filter(date=aftertomorrow)
        eligable_locations = Livraison.objects.filter(place_id__isnull=False, heure_livraison__in = midi, date=aftertomorrow )
        livraisons =[]
        for a in eligable_locations:
            data = {
                'lat': float(a.lat),
                'lng': float(a.lng),
                'place_id': a.place_id,
                'nom': a.nom,
                'heure_livraison': a.heure_livraison,
                'convives' : a.convives,
                'mode_envoi' : a.mode_envoi,
            }

            livraisons.append(data)

        context = {'key': key,
                   'livraisons':livraisons,
                   'form': form,
                   'distances':distances,
                   'routes21':routes21,
                   'routes':filtered_routes,
                   'todo_livraison':todo_livraison,
                   'route1':route1,

        }
        return render(request, 'listings/maptodaymidi.html', context)

    def post(self, request):
        form = DistanceForm(request.POST)
        if form.is_valid():
            from_location = form.cleaned_data['from_location']
            from_location_info = Livraison.objects.get(nom=from_location)
            from_adress_string = str(from_location_info.adress)+", "+str(from_location_info.zipcode)+", "+str(from_location_info.city)+", "+str(from_location_info.country)

            to_location = form.cleaned_data['to_location']
            to_location_info = Livraison.objects.get(nom=to_location)
            to_adress_string = str(to_location_info.adress)+", "+str(to_location_info.zipcode)+", "+str(to_location_info.city)+", "+str(to_location_info.country)

            mode = form.cleaned_data['mode']
            now = datetime.now()

            gmaps = googlemaps.Client(key= settings.GOOGLE_API_KEY)
            calculate = gmaps.distance_matrix(
                from_adress_string,
                to_adress_string,
                mode = mode,
                departure_time = now
            )
            print(calculate)

            duration_secons = calculate['rows'][0]['elements'][0]['duration']['value']
            duration_minutes = duration_secons/60

            distance_meters = calculate['rows'][0]['elements'][0]['distance']['value']
            distance_km = distance_meters/1000

            if 'duration_in_traffic' in calculate['rows'][0]['elements'][0]:
                duration_in_traffic_seconds = calculate['rows'][0]['elements'][0]['duration_in_traffic']['value']
                duration_in_traffic_minutes = duration_in_traffic_seconds/60
            else:
                duration_in_traffic_minutes = None

            obj = Distances(
                from_location = Livraison.objects.get(nom=from_location),
                to_location = Livraison.objects.get(nom=to_location),
                mode = mode,
                distance_km = distance_km,
                distance_mins = duration_minutes,
                distance_traffic_mins = duration_in_traffic_minutes
            )

            obj.save()


        return render(request, 'listings/maptodaymidi.html')
class MapTodayView(View):
    def get(self, request):
        key = settings.GOOGLE_API_KEY
        form = DistanceForm
        today = datetime.now().date()
        aftertomorrow = today + timedelta(2)
        distances = Distances.objects.all()
        matin = ['05h00', '05h15', '05h30', '05h45', '06h00', '06h15', '06h30', '06h45', '07h00', '07h15', '07h30', '07h45', '08h00','08h15', '08h30', '08h45', '09h00', '09h15', '09h30', '09h45', 'recup']
        todo_livraison = Livraison.objects.filter(date=aftertomorrow, heure_livraison__in = matin, place_id__isnull=False, statut__id= 21)
        routes21 = Route.objects.filter(id=21)
        routes = Route.objects.filter(date=aftertomorrow)

        route1 = Livraison.objects.filter(date=aftertomorrow)
        eligable_locations = Livraison.objects.filter(place_id__isnull=False, heure_livraison__in = matin, date=aftertomorrow )
        livraisons =[]
        for a in eligable_locations:
            data = {
                'lat': float(a.lat),
                'lng': float(a.lng),
                'place_id': a.place_id,
                'nom': a.nom,
                'heure_livraison': a.heure_livraison,
                'convives' : a.convives,
                'mode_envoi' : a.mode_envoi,
            }

            livraisons.append(data)

        context = {'key': key,
                   'livraisons':livraisons,
                   'form': form,
                   'distances':distances,
                   'routes21':routes21,
                   'routes':routes,
                   'todo_livraison':todo_livraison,
                   'route1':route1,

        }
        return render(request, 'listings/maptoday.html', context)

    def post(self, request):
        form = DistanceForm(request.POST)
        if form.is_valid():
            from_location = form.cleaned_data['from_location']
            from_location_info = Livraison.objects.get(nom=from_location)
            from_adress_string = str(from_location_info.adress)+", "+str(from_location_info.zipcode)+", "+str(from_location_info.city)+", "+str(from_location_info.country)

            to_location = form.cleaned_data['to_location']
            to_location_info = Livraison.objects.get(nom=to_location)
            to_adress_string = str(to_location_info.adress)+", "+str(to_location_info.zipcode)+", "+str(to_location_info.city)+", "+str(to_location_info.country)

            mode = form.cleaned_data['mode']
            now = datetime.now()

            gmaps = googlemaps.Client(key= settings.GOOGLE_API_KEY)
            calculate = gmaps.distance_matrix(
                from_adress_string,
                to_adress_string,
                mode = mode,
                departure_time = now
            )
            print(calculate)

            duration_secons = calculate['rows'][0]['elements'][0]['duration']['value']
            duration_minutes = duration_secons/60

            distance_meters = calculate['rows'][0]['elements'][0]['distance']['value']
            distance_km = distance_meters/1000

            if 'duration_in_traffic' in calculate['rows'][0]['elements'][0]:
                duration_in_traffic_seconds = calculate['rows'][0]['elements'][0]['duration_in_traffic']['value']
                duration_in_traffic_minutes = duration_in_traffic_seconds/60
            else:
                duration_in_traffic_minutes = None

            obj = Distances(
                from_location = Livraison.objects.get(nom=from_location),
                to_location = Livraison.objects.get(nom=to_location),
                mode = mode,
                distance_km = distance_km,
                distance_mins = duration_minutes,
                distance_traffic_mins = duration_in_traffic_minutes
            )

            obj.save()

        return redirect('my_maptoday_view')


class MapApremDimView(View):
    def get(self, request):
        key = settings.GOOGLE_API_KEY
        form = DistanceForm
        today = datetime.now().date()
        tomorrow = today + timedelta(3)
        distances = Distances.objects.all()
        aprem = ['13h00', '13h15', '13h30', '13h45', '14h00', '14h15', '14h30', '14h45', '15h00', '15h15', '15h30', '15h45', '16h00', '16h15', '16h30', '16h45', '17h00', '17h15', '17h30', '17h45', '18h00', '18h15', '18h30', '18h45', '19h00', 'recup']
        midi = ['05h00', '05h15', '05h30', '05h45', '06h00', '06h15', '06h30', '06h45', '07h00', '07h15', '07h30', '07h45', '08h00', '08h15', '08h30', '08h45', '09h00', '09h15', '09h30', '09h45', '10h00', '10h15', '10h30', '10h45', '11h00', '11h15', '11h30', '11h45', '12h00', '12h15', '12h30', '12h45']
        todo_livraison = Livraison.objects.filter(statut=21, date=tomorrow, heure_livraison__in = aprem, place_id__isnull=False)
        routes = Route.objects.filter(date=tomorrow)
        routes_to_hide = Route.objects.filter(
                livraisons__heure_livraison__in=midi
        )

        # Exclude routes to hide
        filtered_routes = routes.exclude(id__in=routes_to_hide)
        route1 = Livraison.objects.filter(date=tomorrow, heure_livraison__in = aprem)
        route8 = Livraison.objects.filter(date=tomorrow, heure_livraison__in = aprem)
        routes21 = Route.objects.filter(id=21)
        eligable_locations = Livraison.objects.filter(place_id__isnull=False, heure_livraison__in = aprem, date=tomorrow )
        livraisons =[]
        for a in eligable_locations:
            data = {
                'lat': float(a.lat),
                'lng': float(a.lng),
                'place_id': a.place_id,
                'nom': a.nom,
                'heure_livraison': a.heure_livraison,
                'adress' : a.adress,
                'convives': a.convives,
                'mode_envoi': a.mode_envoi,
            }

            livraisons.append(data)

        context = {'key': key,
                   'livraisons':livraisons,
                   'form': form,
                   'distances':distances,
                   'route1':route1,
                   'route8':route8,
                   'routes':filtered_routes,
                   'todo_livraison':todo_livraison,
                   'routes21': routes21,


        }
        return render(request, 'listings/mapapremdim.html', context)

    def post(self, request):
        form = DistanceForm(request.POST)
        if form.is_valid():
            from_location = form.cleaned_data['from_location']
            from_location_info = Livraison.objects.get(nom=from_location)
            from_adress_string = str(from_location_info.adress)+", "+str(from_location_info.zipcode)+", "+str(from_location_info.city)+", "+str(from_location_info.country)

            to_location = form.cleaned_data['to_location']
            to_location_info = Livraison.objects.get(nom=to_location)
            to_adress_string = str(to_location_info.adress)+", "+str(to_location_info.zipcode)+", "+str(to_location_info.city)+", "+str(to_location_info.country)

            mode = form.cleaned_data['mode']
            now = datetime.now()

            gmaps = googlemaps.Client(key= settings.GOOGLE_API_KEY)
            calculate = gmaps.distance_matrix(
                from_adress_string,
                to_adress_string,
                mode = mode,
                departure_time = now
            )
            print(calculate)

            duration_secons = calculate['rows'][0]['elements'][0]['duration']['value']
            duration_minutes = duration_secons/60

            distance_meters = calculate['rows'][0]['elements'][0]['distance']['value']
            distance_km = distance_meters/1000

            if 'duration_in_traffic' in calculate['rows'][0]['elements'][0]:
                duration_in_traffic_seconds = calculate['rows'][0]['elements'][0]['duration_in_traffic']['value']
                duration_in_traffic_minutes = duration_in_traffic_seconds/60
            else:
                duration_in_traffic_minutes = None

            obj = Distances(
                from_location = Livraison.objects.get(nom=from_location),
                to_location = Livraison.objects.get(nom=to_location),
                mode = mode,
                distance_km = distance_km,
                distance_mins = duration_minutes,
                distance_traffic_mins = duration_in_traffic_minutes
            )

            obj.save()

            return render(request, 'listings/mapdimaprem.html')
class MapMidiDimView(View):
    def get(self, request):
        key = settings.GOOGLE_API_KEY
        form = DistanceForm
        today = datetime.now().date()
        aftertomorrow = today + timedelta(3)
        distances = Distances.objects.all()
        midi = ['10h00', '10h15', '10h30', '10h45', '11h00', '11h15', '11h30', '11h45', '12h00', '12h15', '12h30', '12h45', 'recup']
        todo_livraison = Livraison.objects.filter(date=aftertomorrow, heure_livraison__in = midi, place_id__isnull=False, statut__id= 21)
        routesmidi = ['2','3','4','5','6','7','8','9','10','11','12']
        routes21 = Route.objects.filter(id=21)
        routes = Route.objects.filter(nom__in=routesmidi, date=aftertomorrow)
        route1 = Livraison.objects.filter(date=aftertomorrow)
        eligable_locations = Livraison.objects.filter(place_id__isnull=False, heure_livraison__in = midi, date=aftertomorrow )
        livraisons =[]
        for a in eligable_locations:
            data = {
                'lat': float(a.lat),
                'lng': float(a.lng),
                'place_id': a.place_id,
                'nom': a.nom,
                'heure_livraison': a.heure_livraison,
                'convives' : a.convives,
                'mode_envoi' : a.mode_envoi,
            }

            livraisons.append(data)

        context = {'key': key,
                   'livraisons':livraisons,
                   'form': form,
                   'distances':distances,
                   'routes21':routes21,
                   'routes':routes,
                   'todo_livraison':todo_livraison,
                   'route1':route1,

        }
        return render(request, 'listings/mapmididim.html', context)

    def post(self, request):
        form = DistanceForm(request.POST)
        if form.is_valid():
            from_location = form.cleaned_data['from_location']
            from_location_info = Livraison.objects.get(nom=from_location)
            from_adress_string = str(from_location_info.adress)+", "+str(from_location_info.zipcode)+", "+str(from_location_info.city)+", "+str(from_location_info.country)

            to_location = form.cleaned_data['to_location']
            to_location_info = Livraison.objects.get(nom=to_location)
            to_adress_string = str(to_location_info.adress)+", "+str(to_location_info.zipcode)+", "+str(to_location_info.city)+", "+str(to_location_info.country)

            mode = form.cleaned_data['mode']
            now = datetime.now()

            gmaps = googlemaps.Client(key= settings.GOOGLE_API_KEY)
            calculate = gmaps.distance_matrix(
                from_adress_string,
                to_adress_string,
                mode = mode,
                departure_time = now
            )
            print(calculate)

            duration_secons = calculate['rows'][0]['elements'][0]['duration']['value']
            duration_minutes = duration_secons/60

            distance_meters = calculate['rows'][0]['elements'][0]['distance']['value']
            distance_km = distance_meters/1000

            if 'duration_in_traffic' in calculate['rows'][0]['elements'][0]:
                duration_in_traffic_seconds = calculate['rows'][0]['elements'][0]['duration_in_traffic']['value']
                duration_in_traffic_minutes = duration_in_traffic_seconds/60
            else:
                duration_in_traffic_minutes = None

            obj = Distances(
                from_location = Livraison.objects.get(nom=from_location),
                to_location = Livraison.objects.get(nom=to_location),
                mode = mode,
                distance_km = distance_km,
                distance_mins = duration_minutes,
                distance_traffic_mins = duration_in_traffic_minutes
            )

            obj.save()



        return render(request, 'listings/mapdimmidi.html')
class MapDimView(View):
    def get(self, request):
        key = settings.GOOGLE_API_KEY
        form = DistanceForm
        today = datetime.now().date()
        aftertomorrow = today + timedelta(3)
        distances = Distances.objects.all()
        matin = ['05h00', '05h15', '05h30', '05h45', '06h00', '06h15', '06h30', '06h45', '07h00', '07h15', '07h30', '07h45', '08h00','08h15', '08h30', '08h45', '09h00', '09h15', '09h30', '09h45', 'recup']
        matin = ['05h00', '05h15', '05h30', '05h45', '06h00', '06h15', '06h30', '06h45', '07h00', '07h15', '07h30', '07h45', '08h00', '08h15', '08h30', '08h45', '09h00', '09h15', '09h30', '09h45']
        todo_livraison = Livraison.objects.filter(date=aftertomorrow, heure_livraison__in = matin, place_id__isnull=False, statut__id= 21)
        routes21 = Route.objects.filter(id=21)
        routes = Route.objects.filter(date=aftertomorrow)
        routes_to_hide = Route.objects.filter(
                livraisons__heure_livraison__in=matin
        )

        # Exclude routes to hide
        filtered_routes = routes.exclude(id__in=routes_to_hide)
        route1 = Livraison.objects.filter(date=aftertomorrow)
        eligable_locations = Livraison.objects.filter(place_id__isnull=False, heure_livraison__in = matin, date=aftertomorrow )
        livraisons =[]
        for a in eligable_locations:
            data = {
                'lat': float(a.lat),
                'lng': float(a.lng),
                'place_id': a.place_id,
                'nom': a.nom,
                'heure_livraison': a.heure_livraison,
                'convives' : a.convives,
                'mode_envoi' : a.mode_envoi,
            }

            livraisons.append(data)

        context = {'key': key,
                   'livraisons':livraisons,
                   'form': form,
                   'distances':distances,
                   'routes21':routes21,
                   'routes':filtered_routes,
                   'todo_livraison':todo_livraison,
                   'route1':route1,

        }
        return render(request, 'listings/mapdim.html', context)

    def post(self, request):
        form = DistanceForm(request.POST)
        if form.is_valid():
            from_location = form.cleaned_data['from_location']
            from_location_info = Livraison.objects.get(nom=from_location)
            from_adress_string = str(from_location_info.adress)+", "+str(from_location_info.zipcode)+", "+str(from_location_info.city)+", "+str(from_location_info.country)

            to_location = form.cleaned_data['to_location']
            to_location_info = Livraison.objects.get(nom=to_location)
            to_adress_string = str(to_location_info.adress)+", "+str(to_location_info.zipcode)+", "+str(to_location_info.city)+", "+str(to_location_info.country)

            mode = form.cleaned_data['mode']
            now = datetime.now()

            gmaps = googlemaps.Client(key= settings.GOOGLE_API_KEY)
            calculate = gmaps.distance_matrix(
                from_adress_string,
                to_adress_string,
                mode = mode,
                departure_time = now
            )
            print(calculate)

            duration_secons = calculate['rows'][0]['elements'][0]['duration']['value']
            duration_minutes = duration_secons/60

            distance_meters = calculate['rows'][0]['elements'][0]['distance']['value']
            distance_km = distance_meters/1000

            if 'duration_in_traffic' in calculate['rows'][0]['elements'][0]:
                duration_in_traffic_seconds = calculate['rows'][0]['elements'][0]['duration_in_traffic']['value']
                duration_in_traffic_minutes = duration_in_traffic_seconds/60
            else:
                duration_in_traffic_minutes = None

            obj = Distances(
                from_location = Livraison.objects.get(nom=from_location),
                to_location = Livraison.objects.get(nom=to_location),
                mode = mode,
                distance_km = distance_km,
                distance_mins = duration_minutes,
                distance_traffic_mins = duration_in_traffic_minutes
            )

            obj.save()

        return redirect('my_mapdim_view')
@login_required
def deleteDistance(request, pk):
    distance = Distances.objects.get(id= pk)
    if request.method == 'POST':
        distance.delete()
        return redirect('my_map_view')

    context = {'distance':distance,}
    return render(request, 'listings/deletedistance.html', context)

class GeocodeAllLivraisonsView(View):
    def get(self, request):
        # Get all livraisons that need geocoding (filter for those without lat/lng or place_id)
        livraisons = Livraison.objects.filter(lat__isnull=True, lng__isnull=True, place_id__isnull=True)

        gmaps = googlemaps.Client(key=settings.GOOGLE_API_KEY)

        for livraison in livraisons:
            if livraison.adress and livraison.country and livraison.zipcode and livraison.city:
                # Construct the address string
                adress_string = f"{livraison.adress}, {livraison.zipcode}, {livraison.city}, {livraison.country}"

                # Geocode the address
                result = gmaps.geocode(adress_string)

                if result:
                    # Extract the lat, lng, and place_id
                    lat = result[0].get('geometry', {}).get('location', {}).get('lat', {})
                    lng = result[0].get('geometry', {}).get('location', {}).get('lng', {})
                    place_id = result[0].get('place_id', {})

                    # Update the Livraison instance
                    livraison.lat = lat
                    livraison.lng = lng
                    livraison.place_id = place_id
                    livraison.save()

        # Redirect to the livraisonstomorrow page
        return redirect('livraisonstomorrow')

class GeocodingTodayView(View):
    def get(self, request, pk):
        today = datetime.now().date()
        tomorrow = today + timedelta(1)
        livraison = Livraison.objects.get(pk = pk)

        if livraison.lng and livraison.lat and livraison.place_id != None:

            lat = livraison.lat
            lng = livraison.lng
            place_id = livraison.place_id




        elif livraison.adress and livraison.country and livraison.zipcode and livraison.city != None:

            adress_string = str(livraison.adress)+", "+str(livraison.zipcode)+", "+str(livraison.city)+", "+str(livraison.country)

            gmaps = googlemaps.Client(key= settings.GOOGLE_API_KEY)
            result =  gmaps.geocode(adress_string)[0]

            lat = result.get('geometry', {}).get('location', {}).get('lat', {})
            lng = result.get('geometry', {}).get('location', {}).get('lng', {})
            place_id = result.get('place_id', {})



            livraison.lat = lat
            livraison.lng = lng
            livraison.place_id = place_id

            livraison.save()
            return redirect('livraisonstoday')

        else:

            result = ""
            lat = ""
            lng = ""
            place_id = ""

        context = {'livraison': livraison,
                   'lat':lat,
                   'lng':lng,
                   'place_id':place_id,

        }
        return render(request, 'listings/geocodingtoday.html', context)



from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt  # Only use if necessary for non-POST actions
from .models import LoadingDock, Livraison

from django.utils import timezone

def associate_livraisons_with_docks(request):
    if request.method == 'POST':
        today = timezone.localdate()  # ou timezone.now().date() si vous utilisez datetime
        # Get all loading docks
        loading_docks = LoadingDock.objects.all()

        associations_made = 0

        # Loop through each loading dock to find matching livraisons
        for dock in loading_docks:
            # Find livraisons where place_id matches the dock's place_id
            livraisons = Livraison.objects.filter(
                place_id=dock.place_id,
                date__gte=today  # Ajout du filtre pour aujourd'hui et futur
            )

            # For each matching livraison, perform the association
            for livraison in livraisons:
                livraison.loading_dock = dock
                livraison.save()
                associations_made += 1

        return JsonResponse({'status': 'success', 'message': f'{associations_made} livraisons associées avec succès.'})
    return JsonResponse({'status': 'error', 'message': 'Méthode non supportée.'}, status=400)


from django.shortcuts import render
from django.contrib import messages
from django.http import JsonResponse
from .models import Livraison, LoadingDock

def associer_toutes_livraisons_docks(request):
    docks = list(LoadingDock.objects.all())
    if not docks:
        return JsonResponse({'error': 'Aucun dock disponible pour l\'association.'}, status=400)

    num_associations = 0
    non_associated = []

    for livraison in Livraison.objects.filter(loading_docks__isnull=True):
        if livraison.nom:
            mots_livraison = livraison.nom.lower().split()
            docks_correspondants = [dock for dock in docks if any(mot in dock.name.lower() for mot in mots_livraison)]
            if docks_correspondants:
                livraison.loading_docks.set(docks_correspondants)
                num_associations += 1
            else:
                non_associated.append(livraison.nom)

    if non_associated:
        messages.warning(request, f"Pas de dock correspondant trouvé pour les livraisons: {', '.join(non_associated)}.")
    
    messages.success(request, f"{num_associations} livraisons ont été associées aux docks correspondants.")
    return JsonResponse({'message': f"{num_associations} livraisons associées", 'non_associated': non_associated})


from django.http import HttpResponse
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
import math

def generate_etiquette_pdf(request, plats):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="etiquettes.pdf"'

    c = canvas.Canvas(response, pagesize=A4)

    label_width_cm = 9  # largeur en cm
    label_width = label_width_cm * 28.35  # conversion en points (~255 pts)

    label_height_cm = 10.8
    label_height = label_height_cm * 28.35  # (~306 pts)

    margin_left = 20
    margin_top = 20

    cols = 2
    rows = 3

    h_space = (A4[0] - 2 * margin_left - cols * label_width) / (cols - 1) if cols > 1 else 0
    v_space = (A4[1] - 2 * margin_top - rows * label_height) / (rows - 1) if rows > 1 else 0

    x_positions = [margin_left + i * (label_width + h_space) for i in range(cols)]
    y_positions = [A4[1] - margin_top - i * (label_height + v_space) - label_height for i in range(rows)]

    def split_text_to_fit_width(text, max_width, font_name="Helvetica-Bold", font_size=14):
        words = text.split()
        lines = []
        current_line = words[0]
        for word in words[1:]:
            test_line = current_line + ' ' + word
            if c.stringWidth(test_line, font_name, font_size) <= max_width:
                current_line = test_line
            else:
                lines.append(current_line)
                current_line = word
        lines.append(current_line)
        return lines

    def draw_centered_multiline(x, y, lines, font_name, font_size):
        total_height = len(lines) * font_size * 1.2

        # Définir la marge en haut en fonction du nombre de lignes
        if len(lines) == 1:
            top_margin = 18  # 1 cm en points
            start_y = y + (label_height - total_height) / 2 - top_margin
        else:
            start_y = y + (label_height - total_height) / 2

        # Définir la marge en bas pour 4 lignes
        bottom_margin = 0
        if len(lines) == 4:
            bottom_margin = 28.35  # 1 cm en points
            # Ajuster start_y pour inclure cette marge en bas
            start_y += bottom_margin
        
        if len(lines) == 5:
            bottom_margin = 37  # 1 cm en points
            # Ajuster start_y pour inclure cette marge en bas
            start_y += bottom_margin

        for i, line in enumerate(lines):
            line_y = start_y - i * font_size * 1.2
            c.setFont(font_name, font_size)
            c.drawCentredString(x + label_width / 2, line_y, line)



    count = 0
    for plat in plats:
        index = count % (cols * rows)
        col_idx = index % cols
        row_idx = index // cols

        x = x_positions[col_idx]
        y = y_positions[row_idx]

        # Si c'est la troisième rangée, descendre le texte de 1cm (28.35 points)
        if row_idx == 2:
            y -= 28.35

        max_width = label_width - 10  # marge à gauche/droite dans l’étiquette
        font_size = 14

        # Couper le texte en lignes pour tenir dans la largeur de 9cm
        lines = split_text_to_fit_width(plat, max_width, font_name="Helvetica-Bold", font_size=font_size)

        # Si le nombre de lignes est trop grand, réduire la taille
        while len(lines) * font_size * 1.2 > (label_height - 10) and font_size > 4:
            font_size -= 1
            lines = split_text_to_fit_width(plat, max_width, font_name="Helvetica-Bold", font_size=font_size)

        # Dessiner le texte centré horizontalement et verticalement
        draw_centered_multiline(x, y, lines, "Helvetica-Bold", font_size)

        count += 1
        if count % (cols * rows) == 0 and count != 0:
            c.showPage()


    c.showPage()
    c.save()
    return response


def generate_multiple_pdfs(request):
    plats = request.GET.getlist('plats')
    return generate_etiquette_pdf(request, plats)


def generate_multiple_pdfs_en(request):
    plats = request.GET.getlist('plats_en')
    return generate_etiquette_pdf(request, plats)

def etiquette_tente(request):
    plats = Plat.objects.all()

    if request.method == 'POST':
        Plat.objects.create(nom=request.POST['nom'])
        return redirect('etiquette-tente')
    return render(request, 'listings/etiquettes-tentes.html', {
        'plats': plats,
    })


@login_required

def livraison_detail(request, ip):
    livraison = get_object_or_404(Livraison, id=ip)
    adresse = livraison.adress
    livreur = Livreur.objects.all()
    journee = Journee.objects.all()
    recuperation = "oui"
    loic = "Loic"
    maxime = "Maxime"
    # If already having place_id, try matching it.
    docks = livraison.loading_docks.all()

    # Google Maps geocode to fetch place_id and coordinate information



    # Initialize forms
    form = LivraisonForm(instance=livraison)
    photo_form = PhotoUploadForm()

    if request.method == "POST":
        # Handling the LivraisonForm submission
        if 'livraison_form' in request.POST:
            form = LivraisonForm(request.POST, instance=livraison)
            current_signature = livraison.signature

            if form.is_valid():
                livraison = form.save(commit=False)
                livraison.status = True  # Mark as validated
                livraison.signature = current_signature  # Preserve signature
                livraison.save()

                return redirect("livraison-detail", ip=livraison.id)

        # Handling the PhotoUploadForm submission for 'imagerecup'
        elif 'photo_form' in request.POST:
            imagesrecup = request.FILES.getlist("imagerecup")
            images = request.FILES.getlist("image") # Fetching multiple uploaded files
            if imagesrecup:  # Ensure at least one file is uploaded
                for img in imagesrecup:
                    PhotoRecup.objects.create(livraison=livraison, image=img)
            elif images:
                for img in images:
                    Photo.objects.create(livraison=livraison, image=img)

            return redirect("livraison-detail", ip=livraison.id)




    # Use Google Maps API to geocode address
    checklist = Checklist.objects.filter(livraison=livraison)
    for item in checklist:
        item.filtered_checklist_items = item.checklistitem_set.filter(quantity__gt=0)


    return render(request, 'listings/livraison_detail.html', {
        'livraison': livraison,
        'livreur': livreur,
        'recuperation': recuperation,
        'form': form,
        'journee': journee,
        'adresse': adresse,
        'loic': loic,
        'maxime': maxime,
        'photo_form': photo_form,
        'checklist': checklist,
        'docks': docks,
    })


from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from .models import LoadingDock

def modify_dock_view(request):
    if request.method == 'POST':
        dock_id = request.POST.get('dockId')  # This should be an integer
        address = request.POST.get('address')
        description = request.POST.get('description')
        place_id = request.POST.get('place_id')
        photo = request.POST.get('photo')

        # Ensure dock_id is converted to an integer
        loading_dock = get_object_or_404(LoadingDock, id=int(dock_id))  # Cast dock_id to integer
        loading_dock.address = address
        loading_dock.description = description
        loading_dock.place_id = place_id
        loading_dock.photo = photo if photo else loading_dock.photo  # Update only if a new photo is provided
        loading_dock.save()

        return JsonResponse({'status': 'success'})

    return JsonResponse({'status': 'error'}, status=400)




from django.contrib.auth.decorators import user_passes_test


@login_required
def update_photo(request, pk):
    livraison = get_object_or_404(Livraison, pk=pk)
    PhotoFormSet = modelformset_factory(Photo, fields=('image',), extra=1, can_delete=True)

    if request.method == 'POST':
        photo_formset = PhotoFormSet(request.POST, request.FILES, queryset=Photo.objects.filter(livraison=livraison))
        if photo_formset.is_valid():
            photos = photo_formset.save(commit=False)
            for photo in photos:
                photo.livraison = livraison
                photo.save()
            for photo in photo_formset.deleted_objects:
                photo.delete()

            # Redirect to livraison_detail view after successful update
            return redirect('livraison-detail', ip=livraison.id)
    else:
        photo_formset = PhotoFormSet(queryset=Photo.objects.filter(livraison=livraison))

    return render(request, 'listings/update_photo.html', {
        'formbis': photo_formset,
        'livraison': livraison,
    })


from django.shortcuts import render
from .models import LoadingDock, Livraison
from django.http import JsonResponse
from datetime import datetime

def loading_docks_view(request):
    if request.method == 'POST':
        selected_date = request.POST.get('selected_date')
        deliveries = Livraison.objects.filter(date=datetime.strptime(selected_date, '%Y-%m-%d'))

        # Use the field name correctly based on your model definition
        docks = {dock.place_id: dock for dock in LoadingDock.objects.filter(
            id__in=deliveries.values_list('loading_dock_id', flat=True)  # Use _id if it's a foreign key.
        )}

        # Prepare data for AJAX response
        docks_data = {dock.place_id: {
            'address': dock.address,
            'description': dock.description,
            'photo': dock.photo.url if dock.photo else None,
            'deliveries': list(deliveries.filter(loading_dock=dock.id).values('id', 'nom'))  # Modify as necessary
        } for dock in docks.values()}

        return JsonResponse(docks_data)

    loading_docks = LoadingDock.objects.all()
    return render(request, 'listings/loading_docks.html', {'loading_docks': loading_docks})


@login_required
def create_loading_dock(request):

    key = settings.GOOGLE_API_KEY
    if request.method == 'POST':
        form = LoadingDockForm(request.POST, request.FILES)
        if form.is_valid():
            loading_dock = form.save(commit=False)
            loading_dock.place_id = request.POST.get('place_id')  # Retrieve place_id from POST data
            loading_dock.save()
            return redirect('create_loading_dock')
    else:
        form = LoadingDockForm()

    return render(request, 'listings/create_loading_dock.html', {'form': form, 'key': key,})


from datetime import datetime, timedelta
from django.shortcuts import render, redirect
from .models import Livraison
@login_required
def livraisonstomorrow(request):
    date_str = request.GET.get('date', datetime.now().date().strftime('%Y-%m-%d'))
    selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    form = XLSXUploadForm()

    matin = ['05h00', '05h15', '05h30', '05h45', '06h00', '06h15', '06h30', '06h45',
             '07h00', '07h15', '07h30', '07h45', '08h00', '08h15', '08h30', '08h45',
             '09h00', '09h15', '09h30', '09h45', 'recup']

    midi = ['10h00', '10h15', '10h30', '10h45', '11h00', '11h15', '11h30', '11h45',
            '12h00', '12h15', '12h30', '12h45', 'recup']

    apresmidi = ['13h00', '13h15', '13h30', '13h45', '14h00', '14h15', '14h30', '14h45',
                 '15h00', '15h15', '15h30', '15h45', '16h00', '16h15', '16h30', '16h45',
                 '17h00', '17h15', '17h30', '17h45', '18h00', '18h15', '18h30', '18h45', '19h00', 'recup']

    livraisonsmatin = Livraison.objects.filter(heure_livraison__in=matin, date=selected_date).order_by('statut', 'position')
    livraisonsmidi = Livraison.objects.filter(heure_livraison__in=midi, date=selected_date).order_by('statut', 'position')
    livraisonsapresmidi = Livraison.objects.filter(heure_livraison__in=apresmidi, date=selected_date).order_by('statut', 'position')
    if not request.user.is_superuser:
        return redirect('unauthorized')
    context = {
        'selected_date': selected_date.strftime('%Y-%m-%d'),
        'form':form,
        'livraisons_par_periode': [
            ('MATIN', livraisonsmatin, 'my_mapaujour_view'),
            ('MIDI', livraisonsmidi, 'my_mapmidiaujour_view'),
            ('APRES MIDI', livraisonsapresmidi, 'my_mapapremaujour_view'),
        ]
    }

    return render(request, 'listings/livraisonstomorrow.html', context)
@login_required
def success_page(request):
    return render(request, 'listings/success.html')
@login_required
def livraisonstoday(request):
    recuperation = "oui"
    today = datetime.now().date()
    tomorrow = today + timedelta(1)
    matin = ['05h00', '05h15', '05h30', '05h45', '06h00', '06h15', '06h30', '06h45', '07h00', '07h15', '07h30', '07h45', '08h00','08h15', '08h30', '08h45', '09h00', '09h15', '09h30','09h45']
    midi = ['10h00', '10h15', '10h30', '10h45', '11h00', '11h15', '11h30', '11h45', '12h00', '12h15', '12h30', '12h45']
    apresmidi = ['13h00', '13h15', '13h30', '13h45', '14h00', '14h15', '14h30', '14h45', '15h00', '15h15', '15h30', '15h45', '16h00', '16h15', '16h30', '16h45', '17h00', '17h15', '17h30', '17h45', '18h00', '18h15', '18h30', '18h45', '19h00']
    livraisons = Livraison.objects.order_by('statut', 'position').filter(date=today)
    livraisonsmatin =  Livraison.objects.order_by('statut', 'position').filter(heure_livraison__in= matin,date=today)
    livraisonsmidi =  Livraison.objects.order_by('statut', 'position').filter(heure_livraison__in=midi, date=today)
    livraisonsapresmidi =  Livraison.objects.order_by('statut', 'position').filter(heure_livraison__in=apresmidi, date=today)
    retourtraiteur = "oui"
    context = {'livraisons':livraisons,
               'recuperation': recuperation,
               'retourtraiteur': retourtraiteur,
               'livraisonsmatin': livraisonsmatin,
               'livraisonsmidi': livraisonsmidi,
               'livraisonsapresmidi': livraisonsapresmidi,
               }

    if not request.user.is_superuser:
        return redirect('unauthorized')

    return render(request, 'listings/livraisonstoday.html', context)
@login_required
def livraisonsresp(request):
    today = datetime.now().date()
    tomorrow = today + timedelta(1)
    livraisonstatusok = Livraison.objects.filter(status=True, date=today,recuperation=False)
    livraisonstatusko = Livraison.objects.filter(status=False, date=today,recuperation=False)
    recuperations = Livraison.objects.filter(recuperation=True, date=today)
    recuperationok = Livraison.objects.filter(recuperation=True, status=True, date=today)
    recuperationko = Livraison.objects.filter(status=False,recuperation=True, date=today)
    recuperation = "oui"
    retourtraiteur = "oui"
    loic = "Loic"
    maxime= "Maxime"
    livraison = Livraison.objects.all()
    livraisons = Livraison.objects.order_by('statut', 'position').filter(date=tomorrow)
    if not request.user.is_superuser:
        return redirect('unauthorized')
    return render(request, 'listings/livraisonsresp.html', context={'livraisons': livraisons,


                                                              'livraisonstatusok':livraisonstatusok,
                                                              'livraisonstatusko':livraisonstatusko,
                                                              'retourtraiteur':retourtraiteur,
                                                              'recuperationok':recuperationok,
                                                              'recuperationko':recuperationko,
                                                              'recuperation' : recuperation,
                                                              'recuperations' : recuperations,
                                                              'loic':loic,
                                                              'maxime': maxime,
                                                              'tomorrow': tomorrow,

                                                              })


from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages

@require_POST
def duplicate_selected_recups(request):
    selected_ids = request.POST.getlist('selected_livraisons')
    new_date_str = request.POST.get('new_date')
    new_journee_id = request.POST.get('new_journee')

    # Validate input
    try:
        new_date = datetime.strptime(new_date_str, '%Y-%m-%d').date()
    except ValueError:
        messages.error(request, "Format de date invalide.")
        return redirect('recuptoday')

    # Validate journee
    new_journee = get_object_or_404(Journee, id=new_journee_id)

    for liv_id in selected_ids:
        original = get_object_or_404(Livraison, pk=liv_id)

        # Duplicate the main Livraison
        new_livraison = Livraison(
            nom=original.nom,
            mode_envoi=original.mode_envoi,
            status=False,
            recuperation=True,
            commentaire=original.commentaire,
            commentairedispatch=original.commentairedispatch,
            adress=original.adress,
            infodetail=original.infodetail,
            zipcode=original.zipcode,
            app=original.app,
            ligne2=original.ligne2,
            convives=original.convives,
            num_commande=original.num_commande,
            nom_client=original.nom_client,
            contact_site=original.contact_site,
            date=new_date,
            heure_livraison="recup",
            heure_livraison_classement=original.heure_livraison_classement,
            date_livraison=original.date_livraison,
            statut=Route.objects.get(id=21),  # Confirm route correctness
            journee=new_journee,
            lat=original.lat,
            lng=original.lng,
            place_id=original.place_id
        )
        new_livraison.save()

        # Duplicate related photos
        for photo in original.livraison_photos.all():
            Photo.objects.create(
                livraison=new_livraison,
                image=photo.image,
                caption=photo.caption
            )

        # Duplicate checklists
        for checklist in original.checklist_set.all():
            new_checklist = Checklist(
                name=checklist.name,
                livraison=new_livraison,
                date=new_date,
                lieu=checklist.lieu,
                num_contrat=checklist.num_contrat,
                nb_convive=checklist.nb_convive,
                heure_livraison=checklist.heure_livraison,
                md=checklist.md,
                status=checklist.status,
                rapportmd=checklist.rapportmd,
                rapportrecup=checklist.rapportrecup,
                commentairevente=checklist.commentairevente,
                notechecklist=checklist.notechecklist,
                is_active=False
            )
            new_checklist.save()

            # Duplicate ChecklistItems
            for item in checklist.checklistitem_set.all():
                ChecklistItem.objects.create(
                    checklist=new_checklist,
                    product=item.product,
                    quantity=item.quantity,
                    status=item.status,
                    is_stock_updated=False
                )

            # Duplicate checklist photos
            for photo in checklist.checklistrecupphoto_set.all():
                ChecklistRecupPhoto.objects.create(
                    checklist=new_checklist,
                    image=photo.image
                )

    messages.success(request, f"{len(selected_ids)} récupérations dupliquées avec succès.")
    return redirect('recuptoday')



@login_required
def recuptoday(request):
    # Get date from request (if any), default to today
    date_str = request.GET.get('date')
    if date_str:
        try:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            selected_date = datetime.now().date()
    else:
        selected_date = datetime.now().date()

    # All recoveries on selected date
    recups = ["Porcelaine", "Chaud et porcelaine", "Porcelaine et bois", "Plateau de bois", "Froid et bois", "Chaud et jetable", "Froid et porcelaine", "Porcelaine / chaud en vrac", "Plateau JLT", "Plateaux à partager", "Chaud et Plateau JLT", "Assiette individuelle", "Assiette Salade-Repas", "En vrac", "Froid et Plateau JLT", "Plateau JLT / chaud en vrac", "Plateau JLT et bois"] 

    # Querysets with optional date filter
    recuperations = Livraison.objects.filter(recuperation=False, date=selected_date, mode_envoi__in=recups)
    recupsencours = Livraison.objects.filter(recuperation=True, status=False)
    recuperationstot_list = Livraison.objects.filter(recuperation=False, mode_envoi__in=recups, date=selected_date)

    journees = Journee.objects.all().order_by('-id')

    # Pagination remains the same
    paginator = Paginator(recuperationstot_list, 30)
    page_number = request.GET.get('page')
    recuperationstot = paginator.get_page(page_number)

    if not request.user.is_superuser:
        return redirect('unauthorized')

    return render(request, 'listings/recuptoday.html', context={
        'recuperations': recuperations,
        'recupsencours': recupsencours,
        'recuperationstot': recuperationstot,
        'journees': journees,
        'selected_date': selected_date.strftime('%Y-%m-%d'),  # pass for form value
    })

from django.shortcuts import get_object_or_404, redirect
from django.contrib.auth.decorators import login_required

from django.shortcuts import get_object_or_404, redirect

@login_required
def update_recup_date_journee(request, livraison_id):
    if request.method == 'POST':
        livraison = get_object_or_404(Livraison, pk=livraison_id)
        new_date_str = request.POST.get('date')
        new_journee_id = request.POST.get('journee')
        # Always assign statuc with id 21
        route = get_object_or_404(Route, pk=21)

        try:
            new_date = datetime.strptime(new_date_str, '%Y-%m-%d').date()
        except ValueError:
            # handle invalid date
            # show error or redirect
            return redirect('recuptoday')

        new_journee = get_object_or_404(Journee, pk=new_journee_id)

        # Update fields
        livraison.date = new_date
        livraison.journee = new_journee
        livraison.statut = route  # always set to ID 21
        livraison.save()

    return redirect('recuptoday')

@login_required
def faq(request):
    today = datetime.now().date()
    tomorrow = today + timedelta(1)

    return render(request, 'listings/faq.html', context={
                                                              })

@login_required
def livraisonrespdetail(request, ip):
    today = datetime.now().date()
    tomorrow = today + timedelta(1)
    livraison = Livraison.objects.get(id=ip)
    livraisons = Livraison.objects.order_by('statut', 'position').filter(date=tomorrow)
    livraisonstatusok = Livraison.objects.filter(status=True, date=today,recuperation=False)
    livraisonstatusko = Livraison.objects.filter(status=False, date=today,recuperation=False)
    recuperation = Livraison.objects.filter(recuperation=True, date=today)
    recuperationok = Livraison.objects.filter(recuperation=True, status=True, date=today)
    recuperationko = Livraison.objects.filter(status=False,recuperation=True, date=today)
    recuperation = "oui"
    retourtraiteur = "oui"
    formbis = LivraisonFeuilleForm(request.POST or None, instance=livraison)
    if formbis.is_valid():
       formbis.save()
       return redirect('livraisonstomorrow')
    if not request.user.is_superuser:
        return redirect('unauthorized')
    return render(request, 'listings/livraisonrespdetail.html', context={'livraisons': livraisons,
                                                              'livraison':livraison,
                                                              'recuperation' : recuperation,
                                                              'livraisonstatusok':livraisonstatusok,
                                                              'livraisonstatusko':livraisonstatusko,
                                                              'retourtraiteur':retourtraiteur,
                                                              'recuperationok':recuperationok,
                                                              'recuperationko':recuperationko,
                                                              'recuperation' : recuperation,
                                                              'tomorrow':tomorrow,
                                                              'formbis':formbis,
                                                              })
@login_required
def livraisonsventesdetail(request, pk):
    today = datetime.now().date()
    tomorrow = today + timedelta(1)
    livraison = Livraison.objects.get(id=pk)
    livraisons = Livraison.objects.order_by('statut', 'position').filter(date=tomorrow)
    livraisonstatusok = Livraison.objects.filter(status=True, date=today,recuperation=False)
    livraisonstatusko = Livraison.objects.filter(status=False, date=today,recuperation=False)
    recuperation = Livraison.objects.filter(recuperation=True, date=today)
    recuperationok = Livraison.objects.filter(recuperation=True, status=True, date=today)
    recuperationko = Livraison.objects.filter(status=False,recuperation=True, date=today)
    recuperation = "oui"
    retourtraiteur = "oui"
    formbis = LivraisonsVentesForm(request.POST or None, instance=livraison)
    if formbis.is_valid():
       formbis.save()
       return redirect('journees-list')

    return render(request, 'listings/livraisonsventes.html', context={'livraisons': livraisons,
                                                              'livraison':livraison,
                                                              'recuperation' : recuperation,
                                                              'livraisonstatusok':livraisonstatusok,
                                                              'livraisonstatusko':livraisonstatusko,
                                                              'retourtraiteur':retourtraiteur,
                                                              'recuperationok':recuperationok,
                                                              'recuperationko':recuperationko,
                                                              'recuperation' : recuperation,
                                                              'tomorrow':tomorrow,
                                                              'formbis':formbis,
                                                              })
@login_required
def livraisonshier(request):

    today = datetime.now().date()
    tomorrow = today + timedelta(1)
    yesterday = today - timedelta(1)
    livraisonstatusok = Livraison.objects.filter(status=True, date=yesterday,recuperation=False)
    livraisonstatusko = Livraison.objects.filter(status=False, date=yesterday,recuperation=False)
    recuperation = Livraison.objects.filter(recuperation=True, date=yesterday)
    recuperationok = Livraison.objects.filter(recuperation=True, status=True, date=yesterday)
    recuperationko = Livraison.objects.filter(status=False,recuperation=True, date=yesterday)
    recuperation = "oui"
    livraison = Livraison.objects.all()
    livraisons = Livraison.objects.order_by('statut', 'position').filter(date=yesterday)
    return render(request, 'listings/livraisonshier.html', context={'livraisons': livraisons,

                                                              'recuperation' : recuperation,
                                                              'livraisonstatusok':livraisonstatusok,
                                                              'livraisonstatusko':livraisonstatusko,

                                                              'recuperationok':recuperationok,
                                                              'recuperationko':recuperationko,
                                                              'recuperation' : recuperation,
                                                              })


class Livraisonsdrag(ListView):
    template_name = 'listings/livraisonsdrag.html'
    model = Livraison
    context_object_name = 'livraisons'

    def get_queryset(self):
        # Default queryset for the view
        today = datetime.now().date()
        tomorrow = today + timedelta(1)

        # Query for tomorrow's deliveries (default queryset)
        livraisons = Livraison.objects.order_by('statut', 'position').filter(date=tomorrow)

        return livraisons

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Get current date and the next few days
        today = datetime.now().date()
        tomorrow = today + timedelta(1)
        tomorrow1 = today + timedelta(2)
        tomorrow2 = today + timedelta(3)

        # Add queries for each day
        livraisonstoday = Livraison.objects.order_by('statut', 'position').filter(date=today)
        livraisonsdim = Livraison.objects.order_by('statut', 'position').filter(date=tomorrow1)
        livraisonslundi = Livraison.objects.order_by('statut', 'position').filter(date=tomorrow2)

        # Add these to the context so that the template can access them
        context['livraisonstoday'] = livraisonstoday
        context['livraisonsdim'] = livraisonsdim
        context['livraisonslundi'] = livraisonslundi

        return context

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json

@csrf_exempt  # Si vous n'avez pas mis la protection CSRF au bon endroit, sinon utilisez standard CSRF
def update_submission_status_detail(request, submission_id):
    if request.method == 'POST':
        data = json.loads(request.body)
        status = data.get('status')
        try:
            submission = Submission.objects.get(id=submission_id)
            submission.status = status
            submission.save()
            return JsonResponse({'success': True})
        except Submission.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Soumission non trouvée'})
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'}, status=405)

@login_required
def delete_livraison(request, pk):
    # remove the film from the user's list

    livraison = Livraison.objects.get(id=pk)

    livraison.delete()

    # return template fragment with all the user's films
    today = datetime.now().date()
    tomorrow = today + timedelta(1)
    livraisons = Livraison.objects.filter(date=tomorrow)

    return render(request, 'listings/partials/livraisonslist.html', {'livraisons': livraisons})
@login_required
def sort(request):
    film_pks_order = request.POST.getlist('film_order')
    films = []
    for idx, film_pk in enumerate(film_pks_order, start=1):
        userfilm = Livraison.objects.get(id=film_pk)
        userfilm.order = idx
        userfilm.save()
        films.append(userfilm)

    return render(request, '/listings/partials/livraisonslist.html', {'films': films})


from django.shortcuts import render, get_object_or_404
from .models import Md, Checklist
from .forms import DateFilterForm
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required

@login_required
def md_dashboard(request):
    # Get the Md instance related to the logged-in user
    md_instance = get_object_or_404(Md, user=request.user)  # This fetches the Md object for the logged-in user

    # Get related checklists for the Md instance
    checklists = Checklist.objects.filter(md=md_instance, is_active=True)

    # Handle date filtering if applicable
    if request.method == 'GET' and 'date' in request.GET:
        form = DateFilterForm(request.GET)
        if form.is_valid():
            date = form.cleaned_data['date']
            checklists = checklists.filter(date=date)
    else:
        form = DateFilterForm()

    # Paginate the checklists
    paginator = Paginator(checklists, 10)  # Show 10 checklists per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Prepare context data to pass to the template
    context = {
        'md_instance': md_instance,
        'checklists': checklists,
        'form': form,
        'page_obj': page_obj,
    }

    return render(request, 'listings/md_dashboard.html', context)

from django.shortcuts import get_object_or_404, redirect, render
from .models import Checklist, ChecklistMDPhoto, ChecklistRecupPhoto
from .forms import ChecklistMDPhotoFormSet, ChecklistRecupPhotoFormSet, RapportForm, RapportRecupForm

@login_required
def ChecklistmdDetailView(request, pk):
    checklist = get_object_or_404(Checklist, pk=pk)
    md_photos = ChecklistMDPhoto.objects.filter(checklist=checklist)
    recup_photos = ChecklistRecupPhoto.objects.filter(checklist=checklist)
    checklist_items = ChecklistItem.objects.filter(checklist=checklist, quantity__gt=0).select_related('product')
    checklist_products = checklist_items.values_list('product__id', flat=True)
    checklist_documents = ChecklistDocument.objects.filter(checklist=checklist)
    # Initialize the formsets and forms with unique prefixes
    formset = ChecklistMDPhotoFormSet(request.POST or None, request.FILES or None, prefix='formset', queryset=ChecklistMDPhoto.objects.filter(checklist=checklist))
    formset1 = ChecklistRecupPhotoFormSet(request.POST or None, request.FILES or None, prefix='formset1', queryset=ChecklistRecupPhoto.objects.filter(checklist=checklist))
    form = RapportForm(request.POST or None, prefix='form', instance=checklist)
    form1 = RapportRecupForm(request.POST or None, prefix='form1', instance=checklist)
    checklist_itemsbreuvage = ChecklistItem.objects.filter(checklist_id=checklist, product__category__name__in=["ALCOOL FORT", "SANS ALCOOL", "VINS",  "BIERES"], quantity__gt=0)

    products = Product.objects.filter(checklistitem__checklist=checklist, quantity__gt=0).prefetch_related('category')
    checklist_items_by_category = {}
    for item in checklist_items:
        product = item.product
        if product:
            for category in product.category.all():  # Iterate over each category of the product
                if category.name not in checklist_items_by_category:
                    checklist_items_by_category[category.name] = []
                checklist_items_by_category[category.name].append(item)

    # Check if checklist name contains 'CFCDN'
    show_cfcdn_category = "CFCDN" in checklist.name

    for item in checklist_itemsbreuvage:
        item.remaining_quantity = item.quantity - (item.consumed_quantity or 0)

    if request.method == 'POST':
    # Check which form or formset was submitted and process only that one
        if 'formset-TOTAL_FORMS' in request.POST and formset.is_valid():
            for form in formset:
                if form.cleaned_data:  # Ensure there's data to save
                    checklist_photo = form.save(commit=False)
                    checklist_photo.checklist = checklist  # Associate with the checklist
                    checklist_photo.save()
            return redirect('checklistmd_detail', pk=checklist.pk)

        elif 'photo' in request.POST:
            imagesrecup = request.FILES.getlist("imagerecup")
            images = request.FILES.getlist("image") # Fetching multiple uploaded files
            if imagesrecup:  # Ensure at least one file is uploaded
                for img in imagesrecup:
                    ChecklistRecupPhoto.objects.create(checklist=checklist, image=img)
            elif images:
                for img in images:
                    ChecklistMDPhoto.objects.create(checklist=checklist, image=img)
                return redirect('checklistmd_detail', pk=checklist.pk)

        elif 'formset1-TOTAL_FORMS' in request.POST and formset1.is_valid():
            for form in formset1:
                if form.cleaned_data:  # Ensure there's data to save
                    recup_photo = form.save(commit=False)
                    recup_photo.checklist = checklist  # Associate with the checklist
                    recup_photo.save()
            return redirect('checklistmd_detail', pk=checklist.pk)

        elif 'form-rapportmd' in request.POST and form.is_valid():
            form.save()
            return redirect('checklistmd_detail', pk=checklist.pk)

        elif 'form1-rapportrecup' in request.POST and form1.is_valid():
            form1.save()
        return redirect('checklistmd_detail', pk=checklist.pk)






    context = {
        'checklist': checklist,
        'livraison': checklist.livraison,
        'formset': formset,
        'formset1': formset1,
        'form': form,
        'checklist_items_by_category': checklist_items_by_category,
        'checklist_itemsbreuvage': checklist_itemsbreuvage,
        'form1': form1,
        'show_cfcdn_category': show_cfcdn_category,
        'checklist_items' : checklist_items,
        'md_photos' : md_photos,
        'recup_photos' : recup_photos,
        'checklist_documents' : checklist_documents,
    }
    return render(request, 'listings/checklistmd_detail.html', context)

from django.http import HttpResponse

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from .models import Submission
from .forms import SubmissionForm
@login_required
def submit_request(request):
    key = settings.GOOGLE_API_KEY
    clients = Client.objects.all()

    if request.method == 'POST':
        form = SubmissionForm(request.POST)
        if form.is_valid():
            submission = form.save(commit=False)  # Create the submission object but don't save immediately
            submission.user = request.user  # Link to the logged-in user
            
            # Get the selected client
            client_id = request.POST.get('client')
            if client_id:
                client = Client.objects.get(id=client_id)
                
                # Copy client details to submission
                submission.company_name = client.company_name
                submission.contact_person = client.contact_person
                submission.phone = client.phone
                submission.email = client.email
                submission.billing_address = client.billing_address
                submission.etage = client.etage
                submission.dock_livraison = client.dock_livraison
                submission.escalier = client.escalier
                submission.ascenseur = client.ascenseur
                submission.carte_dock = client.carte_dock
                submission.payment_mode = client.payment_mode
                submission.ordered_by = client.ordered_by
                submission.event_location = client.event_location


            messages.success(request, 'Votre soumission a été enregistrée avec succès.')

            submission.save()  # Save the submission to the database




            return redirect('submission_detail', submission_id=submission.id)  # Redirect after saving
    else:
        form = SubmissionForm()

    return render(request, 'listings/submit_request.html', {
        'form': form,
        'key': key,
        'clients': clients,
      
    })


from django.shortcuts import render
from django.db.models import Count, Q, F, Case, When, FloatField
from .models import Livraison
from .forms import LivraisonFilterForm
def livraison_stats_view(request):
    # initialiser par défaut
    percentage_with_vehicle = 0
    percentage_with_vehicle_photo = 0

    form = LivraisonFilterForm(request.GET or None)
    stats = {}

    if form.is_valid():
        livreur = form.cleaned_data['livreur']
        start_date = form.cleaned_data['start_date']
        end_date = form.cleaned_data['end_date']

        # Build the queryset
        livraison_qs = Livraison.objects.all()

        if livreur:
            livraison_qs = livraison_qs.filter(statut__livreur=livreur)
        if start_date:
            livraison_qs = livraison_qs.filter(date__gte=start_date)
        if end_date:
            livraison_qs = livraison_qs.filter(date__lte=end_date)

        total_livraisons = livraison_qs.count()

        # Livraisons avec recuperation=False
        livraisons_no_recup = livraison_qs.filter(recuperation=False)
        total_no_recup = livraisons_no_recup.count()

        # Livraisons avec recuperation=True
        livraisons_recup = livraison_qs.filter(recuperation=True).count()

        # Livraisons avec checklist actif (dans le filtre recuperation=False)
        livraisons_with_checklist = livraisons_no_recup.filter(checklist__isnull=False, checklist__is_active=True).count()

        livraisons_avec_photos = livraisons_no_recup.filter(livraison_photos__isnull=False).distinct().count()
        livraisons_avec_photos_recups = livraisons_no_recup.filter(livraison_photos_recups__isnull=False).distinct().count()

        # Calcul du pourcentage
        if total_no_recup > 0:
            pourcentage_photo = (livraisons_avec_photos / total_no_recup) * 100
            pourcentage_photo_recup = (livraisons_avec_photos_recups / total_no_recup) * 100
        else:
            pourcentage_photo = 0
            pourcentage_photo_recup = 0

        # Fonction pour calculer le pourcentage
        def percentage(numerator, denominator):
            return (numerator / denominator * 100) if denominator else 0

        # Calculs des pourcentages
        signature_pct = percentage(
            livraisons_no_recup.filter(signature__isnull=False).exclude(signature='').count(),
            total_no_recup
        )

      

        valide_pct = percentage(
            livraisons_no_recup.filter(status=True).count(),
            total_no_recup
        )

        nom_client_signature_pct = percentage(
            livraisons_no_recup.filter(nom_client_signature__isnull=False).exclude(nom_client_signature='').count(),
            total_no_recup
        )

        # Pourcentage de livraisons avec checklist par rapport à total_no_recup
        livraisons_with_checklist_pct = percentage(
            livraisons_with_checklist,
            total_no_recup
        )



        # Préparer le contexte
        stats = {
            'total_livraisons': total_livraisons,
            'total_no_recup': total_no_recup,
            'livraisons_with_checklist': livraisons_with_checklist,
            'livraisons_with_checklist_pct': livraisons_with_checklist_pct,  # Ajout ici
            'signature_pct': signature_pct,
            'pourcentage_photo': pourcentage_photo,
            'pourcentage_photo_recup':pourcentage_photo_recup,
            'valide_pct': valide_pct,
            'nom_client_signature_pct': nom_client_signature_pct,
            'livraisons_recup': livraisons_recup,
        }
        # Initialisation
        routes_queryset = Route.objects.all()

        # Filtrage par date
        if start_date:
            routes_queryset = routes_queryset.filter(date__gte=start_date)
        if end_date:
            routes_queryset = routes_queryset.filter(date__lte=end_date)

        total_routes = routes_queryset.count()

        # Filtrer routes liées au livrerur
        routes_for_livreur = routes_queryset.filter(livreur=livreur)

        # Routes avec véhicules liés
        routes_with_vehicles = routes_for_livreur.filter(vehicles__isnull=False).distinct()

        # Routes avec véhicules contenant des photos
        routes_with_vehicles_photo = routes_with_vehicles.annotate(
            nb_vehicles_with_photo=Count(
                'vehicles',
                filter=Q(vehicles__photos__isnull=False) & ~Q(vehicles__photos='')
            )
        ).filter(nb_vehicles_with_photo__gt=0)

        # Calcul du pourcentage si total > 0
        if total_routes > 0:
            percentage_with_vehicle = (routes_with_vehicles.count() / total_routes) * 100
            percentage_with_vehicle_photo = (routes_with_vehicles_photo.count() / total_routes) * 100
        else:
            percentage_with_vehicle = 0
            percentage_with_vehicle_photo = 0

        if not request.user.is_superuser:
            return redirect('unauthorized')




            # Ajouter dans le contexte
    context = {
        'form': form,
        'stats': stats,
        'percentage_with_vehicle': percentage_with_vehicle,
        'percentage_with_vehicle_photo': percentage_with_vehicle_photo,
    }
    return render(request, 'listings/dashboard_stats.html', context)



from django.http import JsonResponse
from .models import Client

def get_client_details(request, client_id):
    client = Client.objects.get(id=client_id)
    data = {
        'company_name': client.company_name,
        'contact_person': client.contact_person,
        'phone': client.phone,
        'email': client.email,
        'billing_address': client.billing_address,
        'etage': client.etage,
        'dock_livraison': client.dock_livraison,
        'escalier': client.escalier,
        'ascenseur': client.ascenseur,
        'carte_dock': client.carte_dock,
        'payment_mode': client.payment_mode,
    }
    return JsonResponse(data)


from django.db.models import Count, Q
from collections import defaultdict


def calendarsub_view(request):
    submissions = Submission.objects.all()
    current_year = date.today().year
    years = [year for year in range(current_year - 5, current_year + 1)]
    today = date.today()
    selected_day = int(request.GET.get('day', 1))  # Default to the first day of the month if none selected
    current_year = date.today().year
    months = [(month, calendar.month_name[month]) for month in range(1, 13)]
    selected_month = int(request.GET.get('month', today.month))
    # Get the number of days in the selected month
    _, num_days_in_month = calendar.monthrange(current_year, selected_month)
    days_in_month = [day for day in range(1, num_days_in_month + 1)]  # Adjust days in month

    # Get the starting day of the month (0 = Monday, 1 = Tuesday, ..., 6 = Sunday)
    first_day_of_month = date(current_year, selected_month, 1).weekday()  # Returns 0-6, Monday-Sunday
    first_day_of_month = (first_day_of_month + 1) % 7  # Adjust to make Sunday = 0

    # Pad the beginning of the month with empty days
    padded_days = [''] * first_day_of_month + days_in_month

        # Limit to a maximum of 7 rows (7 days x 6 rows if overflow, note that normally only one row will be needed to display)
    # Calculate how many rows we need (84 slots for a month max, but we are only showing 7 rows)
    num_rows = (len(padded_days) + 6) // 7  # To fill the calendar

    french_months = [
        "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
        "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"
    ]

    # Fetch checklists for the selected month
    submissions = Submission.objects.filter(date__month=int(selected_month))



    context = {
        'submissions': submissions,
        'today': today,
        'months': [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12],  # Months as numbers
        'selected_month': selected_month,
        'days': padded_days,
        'months': months,
        'french_months': french_months,
        'selected_day': selected_day,
        'selected_date': date(current_year, selected_month, selected_day).strftime('%d %B %Y'),
        'years': years,
        'selected_year': current_year,
        'is_chefcuisine': request.user.groups.filter(name='chefcuisine').exists(),
        'is_ventes': request.user.groups.filter(name='ventes').exists(),
        'is_admin': request.user.groups.filter(name='admin').exists(),
    }
    return render(request, 'listings/calendar_submission.html', context)


from django.shortcuts import render
from django.utils import timezone
from datetime import date, datetime, timedelta
from collections import defaultdict
import calendar

def calendarsubcreate_view(request):
    current_year = date.today().year
    years = [year for year in range(current_year - 5, current_year + 1)]
    today = date.today()
    selected_day = int(request.GET.get('day', 1))
    selected_month = int(request.GET.get('month', today.month))

    # Get the number of days in the selected month
    _, num_days_in_month = calendar.monthrange(current_year, selected_month)
    days_in_month = list(range(1, num_days_in_month + 1))

    # Get the starting day of the month (0 = Monday, ..., 6 = Sunday)
    first_day_of_month = (date(current_year, selected_month, 1).weekday() + 1) % 7

    # Pad the beginning of the month with empty days
    padded_days = [''] * first_day_of_month + days_in_month

    french_months = [
        "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
        "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"
    ]

    # Fetch all submissions for the selected month
    submissions = Submission.objects.filter(
        created_at__month=selected_month, 
        created_at__year=current_year
    )

    # NOUVEAU SYSTÈME DE NOTIFICATIONS
    now = timezone.now()
    
    # 1. Soumissions événement avec date événement dans moins de 48h et toujours en cours
    notifications_urgentes = Submission.objects.filter(
        submission_type='Soumission événement',
        status='en_cours',
        date__lte=now + timedelta(hours=48),
        date__gte=now  # Pour ne pas inclure les événements passés
    ).select_related('user').order_by('date')
    
    # 2. Soumissions événement du mois actuel, créées depuis plus de 48h et toujours en cours
    debut_mois_actuel = date.today().replace(day=1)
    fin_mois_actuel = (debut_mois_actuel.replace(month=debut_mois_actuel.month % 12 + 1, day=1) 
                       if debut_mois_actuel.month < 12 
                       else debut_mois_actuel.replace(year=debut_mois_actuel.year + 1, month=1, day=1))
    
    notifications_mois_actuel = Submission.objects.filter(
        submission_type='Soumission événement',
        status='en_cours',
        date__gte=debut_mois_actuel,
        date__lt=fin_mois_actuel,
        created_at__lt=now - timedelta(hours=48)
    ).exclude(
        # Exclure celles qui sont déjà dans notifications_urgentes
        id__in=notifications_urgentes.values_list('id', flat=True)
    ).select_related('user').order_by('date')

    # Initialize a dictionary to hold counts of statuses per day
    daily_counts = defaultdict(lambda: {'en_cours': 0})

    # Count statuses for each creation day for submissions that are "en cours"
    for submission in submissions:
        day = submission.created_at.day
        if submission.status == 'en_cours':
            daily_counts[day]['en_cours'] += 1

    notifications_envoyees_48h = Submission.objects.filter(
    submission_type='Soumission événement',
    status='envoyé',
    created_at__lt=now - timedelta(hours=48)  # Utiliser created_at si sent_at n'existe pas
    )

    context = {
        'submissions': submissions,
        'today': today,
        'months': list(enumerate(calendar.month_name[1:], start=1)),
        'selected_month': selected_month,
        'days': padded_days,
        'french_months': french_months,
        'selected_day': selected_day,
        'selected_date': date(current_year, selected_month, selected_day).strftime('%d %B %Y'),
        'years': years,
        'daily_counts': daily_counts,
        'selected_year': current_year,
        'is_chefcuisine': request.user.groups.filter(name='chefcuisine').exists(),
        'is_ventes': request.user.groups.filter(name='ventes').exists(),
        'is_admin': request.user.groups.filter(name='admin').exists(),
        # Nouvelles variables pour les notifications
        'notifications_urgentes': notifications_urgentes,
        'notifications_mois_actuel': notifications_mois_actuel,
        'notifications_envoyees_48h': notifications_envoyees_48h,
    }

    return render(request, 'listings/calendarcreate_submission.html', context)

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q
from datetime import datetime, timedelta
from .models import Submission, Conseiller

@login_required
def manage_submissions(request):
    # Récupérer tous les conseillers
    conseillers = Conseiller.objects.select_related('user').all()
    
    # Commencer avec toutes les soumissions
    submissions = Submission.objects.all().order_by('-created_at')
    
    # Filtre par Conseiller
    selected_conseiller = request.GET.get('conseiller')
    if selected_conseiller:
        submissions = submissions.filter(user__conseiller__id=selected_conseiller)

    # Filtre par Type de Soumission
    submission_type = request.GET.get('submission_type')
    if submission_type:
        submissions = submissions.filter(submission_type=submission_type)

    # Filtre par Statut
    status_filter = request.GET.get('status')
    if status_filter:
        submissions = submissions.filter(status=status_filter)

    # NOUVEAU : Filtres de date de création
    creation_start_date = request.GET.get('creation_start_date')
    creation_end_date = request.GET.get('creation_end_date')
    
    if creation_start_date and creation_end_date:
        start_datetime = datetime.strptime(creation_start_date, "%Y-%m-%d")
        end_datetime = datetime.strptime(creation_end_date, "%Y-%m-%d") + timedelta(days=1)
        submissions = submissions.filter(created_at__range=[start_datetime, end_datetime])
    elif creation_start_date:
        start_datetime = datetime.strptime(creation_start_date, "%Y-%m-%d")
        submissions = submissions.filter(created_at__gte=start_datetime)
    elif creation_end_date:
        end_datetime = datetime.strptime(creation_end_date, "%Y-%m-%d") + timedelta(days=1)
        submissions = submissions.filter(created_at__lt=end_datetime)

    # NOUVEAU : Filtres de date d'événement
    event_start_date = request.GET.get('event_start_date')
    event_end_date = request.GET.get('event_end_date')
    
    if event_start_date and event_end_date:
        submissions = submissions.filter(
            date__gte=event_start_date,
            date__lte=event_end_date
        )
    elif event_start_date:
        submissions = submissions.filter(date__gte=event_start_date)
    elif event_end_date:
        submissions = submissions.filter(date__lte=event_end_date)

    # NOUVEAU : Filtre pour activer/désactiver Soumissions et Commandes
    show_soumissions = request.GET.get('show_soumissions', 'on') == 'on'
    show_commandes = request.GET.get('show_commandes', 'on') == 'on'
    
    type_filter = Q()
    if show_soumissions and not show_commandes:
        type_filter = Q(submission_type__icontains='Soumission')
    elif show_commandes and not show_soumissions:
        type_filter = Q(submission_type__icontains='Commande')
    elif not show_soumissions and not show_commandes:
        # Si aucun n'est sélectionné, ne rien afficher
        type_filter = Q(pk__in=[])
    # Si les deux sont sélectionnés, on ne filtre pas (affiche tout)
    
    if type_filter:
        submissions = submissions.filter(type_filter)

    # Recherche
    search_query = request.GET.get('search', '')
    if search_query:
        submissions = submissions.filter(
            Q(company_name__icontains=search_query) |
            Q(user__username__icontains=search_query) |
            Q(ordered_by__icontains=search_query) |
            Q(contact_person__icontains=search_query) |
            Q(phone__icontains=search_query) |
            Q(email__icontains=search_query)
        )

    # Statistiques (avant pagination)
    submission_count = submissions.count()
    
    # Calcul des statistiques par type
    count_validated = submissions.filter(status='valide', submission_type__icontains='Soumission').count()
    count_rejected = submissions.filter(status='refuse', submission_type__icontains='Soumission').count()
    count_in_progress = submissions.filter(status='en_cours', submission_type__icontains='Soumission').count()
    
    count_validatedcontrat = submissions.filter(status='valide', submission_type__icontains='Commande').count()
    count_rejectedontrat = submissions.filter(status='refuse', submission_type__icontains='Commande').count()
    count_in_progressontrat = submissions.filter(status='en_cours', submission_type__icontains='Commande').count()

    # Pagination
    paginator = Paginator(submissions, 15)  # 15 soumissions par page
    page_number = request.GET.get('page')
    
    try:
        submissions_page = paginator.page(page_number)
    except PageNotAnInteger:
        submissions_page = paginator.page(1)
    except EmptyPage:
        submissions_page = paginator.page(paginator.num_pages)

    context = {
        'submissions': submissions_page,
        'conseillers': conseillers,
        'selected_conseiller': selected_conseiller,
        'submission_type': submission_type,
        'status_filter': status_filter,
        'submission_count': submission_count,
        'count_validated': count_validated,
        'count_rejected': count_rejected,
        'count_in_progress': count_in_progress,
        'count_validatedcontrat': count_validatedcontrat,
        'count_rejectedontrat': count_rejectedontrat,
        'count_in_progressontrat': count_in_progressontrat,
        # Dates de création
        'creation_start_date': creation_start_date,
        'creation_end_date': creation_end_date,
        # Dates d'événement  
        'event_start_date': event_start_date,
        'event_end_date': event_end_date,
        # Toggles
        'show_soumissions': show_soumissions,
        'show_commandes': show_commandes,
        # Recherche
        'search_query': search_query,
    }

    return render(request, 'listings/manage_submissions.html', context)

def get_conseiller_username(request):
    # Retrieve the conseiller id from the GET request
    conseiller_id = request.GET.get('id')

    # Check if the id is provided and not empty
    if not conseiller_id:
        return JsonResponse({'username': ''}, status=400)  # Return an empty username with a bad request status

    try:
        # Try to get the conseiller with the provided id
        conseiller = Conseiller.objects.get(id=conseiller_id)
        return JsonResponse({'username': conseiller.user.username})  # Return the username in the response
    except Conseiller.DoesNotExist:
        return JsonResponse({'username': ''}, status=404)  # Handle cases where conseiller does not exist


from django.http import JsonResponse
from django.views.decorators.http import require_POST

@require_POST
def save_credit_card_info(request):
    card_number = request.POST.get('card_number')
    card_holder = request.POST.get('card_holder')
    expiry_date = request.POST.get('expiry_date')
    cvv = request.POST.get('cvv')

    # Here you can implement your business logic to store the details
    # e.g., saving to the database or processing the payment

    return JsonResponse({'success': True, 'message': 'Détails de la carte enregistrés avec succès!'})


from django.shortcuts import render, get_object_or_404, redirect

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import Submission
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile

@csrf_exempt
def update_submission(request, submission_id):
    if request.method == 'POST':
        try:
            submission = get_object_or_404(Submission, id=submission_id)

            # Handle file upload
            if 'document' in request.FILES:
                file = request.FILES['document']
                if submission.document:
                    default_storage.delete(submission.document.path)
                path = default_storage.save(f'submission_documents/{file.name}', ContentFile(file.read()))
                submission.document = path

            # Parse date
            date_str = request.POST.get('date', '').strip()
            if date_str:
                from datetime import datetime
                submission.date = datetime.strptime(date_str, "%Y-%m-%d").date()

            # Update text fields
            submission.company_name = request.POST.get('company_name', '').strip()
            submission.event_postcode = request.POST.get('postal_code', '').strip()
            submission.event_location = request.POST.get('event_location', '').strip()
            submission.phone = request.POST.get('phone', '').strip()
            submission.email = request.POST.get('email', '').strip()
            
            # Handle ordered_by
            ordered_by = request.POST.get('ordered_by', '').strip()
            if ordered_by.lower() == 'none':
                ordered_by = ''
            submission.ordered_by = ordered_by

            # Handle boolean fields properly
            boolean_fields = ['avec_service_md', 'location_materiel', 'avec_alcool']
            
            for field in boolean_fields:
                if field in request.POST:
                    value = request.POST.get(field, '').strip()
                    # Remove non-breaking spaces and regular spaces
                    value = value.replace('\xa0', '').replace(' ', '')
                    
                    # Convert to boolean
                    if value in ['True', 'true', '1']:
                        setattr(submission, field, True)
                    elif value in ['False', 'false', '0', '']:
                        setattr(submission, field, False)
                    else:
                        # Default to False if unrecognized
                        setattr(submission, field, False)

            # Handle guest_count
            guest_str = request.POST.get('guest_count', '').replace("\xa0", "").strip()
            try:
                submission.guest_count = int(guest_str) if guest_str else None
            except ValueError:
                submission.guest_count = None

            # Save submission
            submission.save()

            return JsonResponse({'success': True})

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    else:
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée.'}, status=405)

def get_livraisons(request, journee_id):
    livraisons = Livraison.objects.filter(journee_id=journee_id, recuperation=False)
    livraison_list = [{
        'heure_depart': livraison.statut.heure_depart,
        'nom': livraison.nom,
        'heure_livraison': livraison.heure_livraison,
        'livreurs': [str(livreur) for livreur in livraison.statut.livreur.all()]
    } for livraison in livraisons]

    return JsonResponse({'livraisons': livraison_list})


from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
import json
from .models import PaymentMode

@csrf_exempt
def update_payment_mode(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            payment_mode_id = data.get('payment_mode_id')
            new_details = data.get('details')

            payment_mode_obj = PaymentMode.objects.get(id=payment_mode_id)
            payment_mode_obj.details = new_details
            payment_mode_obj.save()

            return JsonResponse({'success': True})

        except PaymentMode.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Mode de paiement non trouvé'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})


def get_livraisons_chaud(request, journee_id):
    livraisons = Livraison.objects.filter(
        journee_id=journee_id,
        recuperation=False,
        nom__icontains="part chaud"
    ).prefetch_related('statut__livreur')

    livraison_list = []
    for livraison in livraisons:
        livreurs = [{'nom': livreur.nom, 'phone': livreur.phone} for livreur in livraison.statut.livreur.all()]
        livraison_list.append({
            'heure_depart': livraison.statut.heure_depart,
            'nom': livraison.nom,
            'heure_livraison': livraison.heure_livraison,
            'livreurs': livreurs
        })

    return JsonResponse({'livraisons': livraison_list})

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json

from django.http import JsonResponse

@csrf_exempt
def create_payment_mode(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        name = data.get('name', '').strip()
        details = data.get('details', '').strip()
        submission_id = data.get('submission_id', None)

        if name:
            mode, created = PaymentMode.objects.get_or_create(
                name=name,
                defaults={'details': details}
            )
            if not created:
                mode.details = details
                mode.save()

            if submission_id:
                submission = Submission.objects.get(id=submission_id)
                submission.payment_mode = mode
                submission.save()

            return JsonResponse({'success': True, 'mode_id': mode.id})
        else:
            return JsonResponse({'success': False, 'error': 'Champs manquants'})
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})


def validate_time_format(time_str):
    """ Helper function to validate HH:MM format for time strings. """
    import re
    time_pattern = re.compile(r'^\d{1,2}:\d{2}$')  # Simple regex for HH:MM format
    return bool(time_pattern.match(time_str))

def validate_time_format(time_str):
    """ Helper function to validate HH:MM format for time strings. """
    import re
    time_pattern = re.compile(r'^\d{1,2}:\d{2}$')  # Simple regex for HH:MM format
    return bool(time_pattern.match(time_str))

def validate_time_format(time_str):
    """ Helper function to validate HH:MM format for time strings. """
    import re
    time_pattern = re.compile(r'^\d{1,2}:\d{2}$')  # Simple regex for HH:MM format
    return bool(time_pattern.match(time_str))

def validate_time_format(time_str):
    """ Helper function to validate HH:MM format for time strings. """
    import re
    time_pattern = re.compile(r'^\d{1,2}:\d{2}$')  # Simple regex for HH:MM format
    return bool(time_pattern.match(time_str))



@login_required
def update_submission_status(request, submission_id):
    submission = get_object_or_404(Submission, id=submission_id)

    if request.method == 'POST':
        new_status = request.POST.get('status')  # Expecting status from a form or AJAX
        if new_status in dict(Submission.STATUS_CHOICES):  # Ensure the new status is valid
            submission.status = new_status
            submission.save()
            messages.success(request, 'Statut de la soumission mis à jour avec succès.')
            return redirect('manage_submissions')

    return render(request, 'listings/update_submission_status.html', {'submission': submission})


from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
@login_required
@require_http_methods(["DELETE"])
def delete_photo_md(request, photo_id):
    # Try to delete from ChecklistMDPhoto first
    try:
        photo = ChecklistMDPhoto.objects.get(id=photo_id)
        photo.delete()
        return JsonResponse({'success': True, 'message': 'Photo MD deleted successfully.'})
    except ChecklistMDPhoto.DoesNotExist:
        pass  # If it doesn’t exist, we’ll try the recup photo

    # If the first attempt fails, try to delete from ChecklistRecupPhoto
    try:
        photo_recup = ChecklistRecupPhoto.objects.get(id=photo_id)
        photo_recup.delete()
        return JsonResponse({'success': True, 'message': 'Photo Récupération deleted successfully.'})
    except ChecklistRecupPhoto.DoesNotExist:
        # If both don't exist, return an error response
        return JsonResponse({'success': False, 'error': 'Photo not found in any category.'}, status=404)


from django.shortcuts import redirect
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Group
@login_required
def custom_login(request):
    print("Custom login view triggered!")  # Check if the view is being called
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            print(f"Logged in user: {user.username}")  # Confirm user is logged in

            # Debugging: print the user's groups
            print("User groups:", [group.name for group in user.groups.all()])

            # Check if user belongs to 'md' group
            if user.groups.filter(name='md').exists():
                print("User belongs to 'md' group, redirecting to faq...")
                return redirect('faq')
            else:
                print("User does not belong to 'md' group, redirecting elsewhere...")
                return redirect('journees-list')
        else:
            return render(request, 'registration/login.html', {'error': 'Invalid credentials'})

    return render(request, 'listings/login.html')





@login_required
def livraisonsdrag_detailtoday(request, pk):
    livraison = Livraison.objects.get(id=pk)
    context = {'livraison': livraison}
    if request.method == 'GET':
        return render(request, 'listings/livraisonsdragtoday.html', context)
    elif request.method == 'PUT':
        data = QueryDict(request.body).dict()
        form = LivraisonDragFormtoday(data, instance=livraison)
        if form.is_valid():
            form.save()
            return render(request, 'listings/partials/livraisonslisttoday.html', context)

    context['form'] = form
    if not request.user.is_superuser:
        return redirect('unauthorized')

    return render(request, 'listings/partials/edit-livraison-formtoday.html', context)
@login_required
def livraison_edit_formtoday(request, pk):
    livraison = Livraison.objects.get(id=pk)
    form = LivraisonDragFormtoday(instance=livraison)
    context = {'livraison': livraison, 'form': form}
    if not request.user.is_superuser:
        return redirect('unauthorized')
    return render(request, 'listings/partials/edit-livraison-formtoday.html', context)


class Livraisonsdragtoday(ListView):
    template_name = 'listings/livraisonsdragtoday.html'
    model = Livraison

    context_object_name = 'livraisons'


    def get_queryset(self):

        today = datetime.now().date()
        tomorrow = today + timedelta(1)
        livraisons = Livraison.objects.order_by('statut', 'position').filter(date=tomorrow)

        return livraisons


@login_required
def add_livraisontoday(request):
    nom = request.POST.get('filmname')
    date = request.POST.get('date')

    # add film
    livraison = Livraison.objects.create(nom=nom, date=date)

    today = datetime.now().date()
    tomorrow = today + timedelta(1)
    livraisons = Livraison.objects.filter(date=today)

    return render(request, 'listings/partials/livraisonslisttoday.html', {'livraisons': livraisons})
@login_required
def livraisonsdrag_detail(request, pk):
    livraison = Livraison.objects.get(id=pk)
    context = {'livraison': livraison}
    if request.method == 'GET':
        return render(request, 'listings/livraisonsdrag.html', context)
    elif request.method == 'PUT':
        data = QueryDict(request.body).dict()
        form = LivraisonDragForm(data, instance=livraison)
        if form.is_valid():
            form.save()
            return render(request, 'listings/partials/livraisonslist.html', context)

    context['form'] = form
    if not request.user.is_superuser:
        return redirect('unauthorized')
    return render(request, 'listings/partials/edit-livraison-form.html', context)
@login_required
def livraison_edit_form(request, livraison_id):
    livraison = get_object_or_404(Livraison, id=livraison_id)
    form = LivraisonDragForm(instance=livraison)

    if request.method == 'POST':
        form = LivraisonDragForm(request.POST, instance=livraison)
        if form.is_valid():
            form.save()
            return render(request, 'listings/partials/livraison_row.html', {'livraison': livraison})

    # Render the partial template with the form
    if not request.user.is_superuser:
        return redirect('unauthorized')
    return render(request, 'listings/partials/livraison_edit_form.html', {'form': form, 'livraison': livraison})

@login_required
def commentcamarche(request):
    context = {}
    if not request.user.is_superuser:
        return redirect('unauthorized')
    return render(request, 'listings/commentcamarche.html', context)
@login_required
def routesfrigo(request):
    today = timezone.now().date()
    routes = Route.objects.filter(date=today)
    journee = Journee.objects.filter(date=today)


    context = {
        'routes': routes,
        'journee':journee,
    }
    return render(request, 'listings/routesfrigo.html', context)


from datetime import datetime, timedelta
from django.shortcuts import redirect
@login_required
def duplicate_model(request, model_id):
    original_object = Livraison.objects.get(pk=model_id)
    today = datetime.now().date()
    tomorrow = today + timedelta(days=1)

    # Create a new Livraison object
    new_object = Livraison()
    new_object.nom = original_object.nom
    new_object.mode_envoi = original_object.mode_envoi
    new_object.status = False
    new_object.recuperation = True
    new_object.commentaire = original_object.commentaire
    new_object.commentairedispatch = original_object.commentairedispatch
    new_object.adress = original_object.adress
    new_object.infodetail = original_object.infodetail
    new_object.zipcode = original_object.zipcode
    new_object.app = original_object.app
    new_object.ligne2 = original_object.ligne2
    new_object.convives = original_object.convives
    new_object.num_commande = original_object.num_commande
    new_object.nom_client = original_object.nom_client
    new_object.contact_site = original_object.contact_site
    new_object.date = tomorrow
    new_object.heure_livraison = "recup"
    new_object.heure_livraison_classement = original_object.heure_livraison_classement
    new_object.date_livraison = original_object.date_livraison
    new_object.statut = Route.objects.get(id=21)  # Ensure this is the correct route
    new_journee = Journee.objects.get(id=original_object.journee.id + 1)
    new_object.journee = new_journee
    new_object.lat = original_object.lat
    new_object.lng = original_object.lng
    new_object.place_id = original_object.place_id

    # Save the new Livraison object first
    new_object.save()

    # Duplicate related Photo objects for the Livraison
    for photo_instance in original_object.livraison_photos.all():
        new_photo = Photo()
        new_photo.livraison = new_object  # Associate the new photo with the new Livraison
        new_photo.image = photo_instance.image  # Copy the image
        new_photo.caption = photo_instance.caption  # Copy the caption if available
        new_photo.save()



    print("Checklists to duplicate:", original_object.checklist_set.all())

    # Duplicate related Checklists
    for checklist in original_object.checklist_set.all():
        new_checklist = Checklist()
        new_checklist.name = checklist.name
        new_checklist.livraison = new_object  # Associate with the duplicated Livraison
        new_checklist.date = tomorrow  # Optionally set to tomorrow or keep original
        new_checklist.lieu = checklist.lieu
        new_checklist.num_contrat = checklist.num_contrat
        new_checklist.nb_convive = checklist.nb_convive
        new_checklist.heure_livraison = checklist.heure_livraison
        new_checklist.md = checklist.md
        new_checklist.status = checklist.status  # Keep the original status or reset if needed
        new_checklist.rapportmd = checklist.rapportmd
        new_checklist.rapportrecup = checklist.rapportrecup
        new_checklist.commentairevente = checklist.commentairevente
        new_checklist.notechecklist = checklist.notechecklist
        new_checklist.conseillere = None
        new_checklist.is_active = False
        new_checklist.save()


        # Duplicate ChecklistItems but do not update the inventory
        for item in checklist.checklistitem_set.all():
            new_item = ChecklistItem()
            new_item.checklist = new_checklist
            new_item.product = item.product
            new_item.quantity = item.quantity
            new_item.status = item.status
            new_item.skip_inventory_update = True  # Set the flag to skip inventory update
            new_item.save()


        for photo in checklist.checklistrecupphoto_set.all():
            new_photo = ChecklistRecupPhoto()
            new_photo.checklist = new_checklist
            new_photo.image = photo.image  # Copy the image
            new_photo.save()

    # Redirect back to a page or render a template
    return redirect('recuptoday')  # Redirect to a specific URL name

from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
@login_required
def duplicate_checklist(request, checklist_id):
    original_checklist = get_object_or_404(Checklist, pk=checklist_id)

    # Duplicate the checklist itself
    new_checklist = Checklist.objects.create(
        name=f"Copie de {original_checklist.name}",
        livraison=None,
        date=original_checklist.date,
        lieu=original_checklist.lieu,
        num_contrat=None,
        nb_convive=original_checklist.nb_convive,
        heure_livraison=original_checklist.heure_livraison,
        md=None,
        added_on=timezone.now() - timedelta(hours=5),
        status='en_cours',
        rapportmd=None,
        rapportrecup=None,
        commentairevente=None,
        notechecklist=None,
        conseillere=original_checklist.conseillere,
        is_active=True  # Assume the duplicate is active by default
    )

    # Duplicate the checklist items (if needed)
    for item in original_checklist.checklistitem_set.all():
        ChecklistItem.objects.create(
            checklist=new_checklist,
            product=item.product,
            quantity=item.quantity,
            is_completed=item.is_completed,
            status="en_cours",
            consumed_quantity=item.consumed_quantity,
            unconsumed_quantity=item.unconsumed_quantity,
            commentaire=item.commentaire,
        )


    # Redirect to the detail view of the new checklist
    return HttpResponseRedirect(reverse('checklist-detail', args=[new_checklist.id]))

    # view ejemplo
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
def delete_menu_submission(request, submission_id, submenu_id):
    if request.method == 'POST':
        try:
            submission = get_object_or_404(Submission, id=submission_id)
            submenu = get_object_or_404(MenuSubmission, id=submenu_id, submission=submission)
            submenu.delete()
            return JsonResponse({'success': True, 'message': 'Sous-menu supprimé.'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    else:
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})


# views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.contrib import messages
from django.http import JsonResponse
from django.db.models import Q, Sum, F, Prefetch
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.utils import timezone
from django.db import transaction
from datetime import datetime, date
from decimal import Decimal
from django.views.generic import DetailView
import json

from .models import (
    UniteMesure, Departement, Fournisseur, Ingredient, 
    CatalogueFournisseur, Recette, SousRecette, 
    RecetteIngredient, SousRecetteIngredient,
    Commande, LigneCommande, ProductionPrevue
)


# Dashboard
@login_required
def dashboard_commande(request):
    context = {
        'total_recettes': Recette.objects.count(),
        'total_ingredients': Ingredient.objects.count(),
        'total_fournisseurs': Fournisseur.objects.count(),
        'commandes_en_cours': Commande.objects.filter(statut='EN_COURS').count(),
        'ingredients_alerte': Ingredient.objects.filter(stock_reel__lte=F('stock_alerte')).count(),
        'commandes_recentes': Commande.objects.all()[:5],
    }
    return render(request, 'recipes/dashboard.html', context)

# Vues CRUD pour UniteMesure
class UniteMesureListView(LoginRequiredMixin, ListView):
    model = UniteMesure
    template_name = 'recipes/unite_mesure_list.html'
    context_object_name = 'unites'
    paginate_by = 20

class UniteMesureCreateView(LoginRequiredMixin, CreateView):
    model = UniteMesure
    form_class = UniteMesureForm
    template_name = 'recipes/unite_mesure_form.html'
    success_url = reverse_lazy('unite_mesure_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Unité de mesure créée avec succès.')
        return super().form_valid(form)

class UniteMesureUpdateView(LoginRequiredMixin, UpdateView):
    model = UniteMesure
    form_class = UniteMesureForm
    template_name = 'recipes/unite_mesure_form.html'
    success_url = reverse_lazy('unite_mesure_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Unité de mesure modifiée avec succès.')
        return super().form_valid(form)

class UniteMesureDeleteView(LoginRequiredMixin, DeleteView):
    model = UniteMesure
    template_name = 'recipes/unite_mesure_confirm_delete.html'
    success_url = reverse_lazy('unite_mesure_list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(self.request, 'Unité de mesure supprimée avec succès.')
        return super().delete(request, *args, **kwargs)

# Vues CRUD pour Ingredient
class IngredientListView(LoginRequiredMixin, ListView):
    model = Ingredient
    template_name = 'recipes/ingredient_list.html'
    context_object_name = 'ingredients'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = super().get_queryset()
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(nom__icontains=search) |
                Q(description__icontains=search)
            )
        return queryset.select_related('unite_mesure')

class IngredientCreateView(LoginRequiredMixin, CreateView):
    model = Ingredient
    form_class = IngredientForm
    template_name = 'recipes/ingredient_form.html'
    success_url = reverse_lazy('ingredient_list')

class IngredientUpdateView(LoginRequiredMixin, UpdateView):
    model = Ingredient
    form_class = IngredientForm
    template_name = 'recipes/ingredient_form.html'
    success_url = reverse_lazy('ingredient_list')

class IngredientDeleteView(LoginRequiredMixin, DeleteView):
    model = Ingredient
    template_name = 'recipes/ingredient_confirm_delete.html'
    success_url = reverse_lazy('ingredient_list')

# Vues CRUD pour Recette avec gestion des ingrédients
@login_required
def recette_create(request):
    if request.method == 'POST':
        form = RecetteForm(request.POST, request.FILES)
        formset = RecetteIngredientFormSet(request.POST, prefix='ingredients')
        
        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                recette = form.save()
                formset.instance = recette
                formset.save()
                recette.calculer_cout()
                messages.success(request, 'Recette créée avec succès.')
                return redirect('recette_detail', pk=recette.pk)
    else:
        form = RecetteForm()
        formset = RecetteIngredientFormSet(prefix='ingredients')
    
    return render(request, 'recipes/recette_form.html', {
        'form': form,
        'formset': formset,
        'title': 'Créer une recette'
    })

@login_required
def recette_update(request, pk):
    recette = get_object_or_404(Recette, pk=pk)
    
    if request.method == 'POST':
        form = RecetteForm(request.POST, request.FILES, instance=recette)
        formset = RecetteIngredientFormSet(request.POST, instance=recette, prefix='ingredients')
        
        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                recette = form.save()
                formset.save()
                recette.calculer_cout()
                messages.success(request, 'Recette modifiée avec succès.')
                return redirect('recette_detail', pk=recette.pk)
    else:
        form = RecetteForm(instance=recette)
        formset = RecetteIngredientFormSet(instance=recette, prefix='ingredients')
    
    return render(request, 'recipes/recette_form.html', {
        'form': form,
        'formset': formset,
        'recette': recette,
        'title': 'Modifier la recette'
    })

class RecetteListView(LoginRequiredMixin, ListView):
    model = Recette
    template_name = 'recipes/recette_list.html'
    context_object_name = 'recettes'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = super().get_queryset()
        search = self.request.GET.get('search')
        departement = self.request.GET.get('departement')
        
        if search:
            queryset = queryset.filter(
                Q(nom__icontains=search) |
                Q(description__icontains=search)
            )
        if departement:
            queryset = queryset.filter(departement_id=departement)
            
        return queryset.select_related('departement').prefetch_related('recette_ingredients__ingredient')

class RecetteDetailView(LoginRequiredMixin, DetailView):
    model = Recette
    template_name = 'recipes/recette_detail.html'
    context_object_name = 'recette'
    
    def get_queryset(self):
        return super().get_queryset().prefetch_related(
            'recette_ingredients__ingredient__unite_mesure',
            'sous_recettes__sousrecette_ingredients__ingredient__unite_mesure'
        )

from decimal import Decimal
from django.db import transaction
from django.db.models import Prefetch
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.shortcuts import render
from django.contrib import messages
import json
from .models import (
    Ingredient, Recette, Departement, Commande, LigneCommande,
    Fournisseur, CatalogueFournisseur
)

@login_required
def commande_create(request):
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        action = request.POST.get('action')

        # --------------------------------------------------
        # Récupérer les départements liés à des ingrédients
        # --------------------------------------------------
        if action == 'get_departements':
            ingredient_ids = json.loads(request.POST.get('ingredient_ids', '[]'))
            departements = Departement.objects.filter(
                recettes__recette_ingredients__ingredient_id__in=ingredient_ids
            ).distinct().values('id', 'nom')
            return JsonResponse(list(departements), safe=False)

        # ------------------------
        # Récupérer les recettes
        # ------------------------
        elif action == 'get_recettes':
            ingredient_id = request.POST.get('ingredient_id')
            departement_id = request.POST.get('departement_id')
            recettes = Recette.objects.filter(
                departement_id=departement_id,
                recette_ingredients__ingredient_id=ingredient_id
            ).distinct().values('id', 'nom')
            return JsonResponse(list(recettes), safe=False)

        # ---------------------------------------------
        # Calculer les besoins en quantités d'ingrédients
        # ---------------------------------------------
        elif action == 'calculer_quantites':
            productions = json.loads(request.POST.get('productions', '[]'))
            ingredients_besoins = {}

            for prod in productions:
                try:
                    recette = Recette.objects.get(id=prod['recette_id'])
                    quantite = Decimal(str(prod.get('quantite', '0')))

                    # Calcul des besoins principaux
                    for ri in recette.recette_ingredients.all():
                        key = str(ri.ingredient.id)
                        if key not in ingredients_besoins:
                            ingredients_besoins[key] = {
                                'nom': ri.ingredient.nom,
                                'unite': ri.ingredient.unite_mesure.symbole if ri.ingredient.unite_mesure else '',
                                'quantite': Decimal('0'),
                                'stock_actuel': Decimal(str(ri.ingredient.stock_reel)) if ri.ingredient.stock_reel else Decimal('0'),
                                'fournisseurs': list(ri.ingredient.catalogues.filter(
                                    actif=True
                                ).values('fournisseur__id', 'fournisseur__nom', 'prix'))
                            }
                        ingredients_besoins[key]['quantite'] += Decimal(str(ri.quantite)) * quantite

                    # Calcul pour les sous-recettes
                    if hasattr(recette, 'sous_recettes'):
                        for sr in recette.sous_recettes.all():
                            for sri in sr.sousrecette_ingredients.all():
                                key = str(sri.ingredient.id)
                                if key not in ingredients_besoins:
                                    ingredients_besoins[key] = {
                                        'nom': sri.ingredient.nom,
                                        'unite': sri.ingredient.unite_mesure.symbole if sri.ingredient.unite_mesure else '',
                                        'quantite': Decimal('0'),
                                        'stock_actuel': Decimal(str(sri.ingredient.stock_reel)) if sri.ingredient.stock_reel else Decimal('0'),
                                        'fournisseurs': list(sri.ingredient.catalogues.filter(
                                            actif=True
                                        ).values('fournisseur__id', 'fournisseur__nom', 'prix'))
                                    }
                                ingredients_besoins[key]['quantite'] += Decimal(str(sri.quantite)) * Decimal(str(sr.quantite)) * quantite

                except Recette.DoesNotExist:
                    return JsonResponse({'error': f'Recette {prod["recette_id"]} introuvable'}, status=404)
                except Exception as e:
                    return JsonResponse({'error': str(e)}, status=500)

            # Calcul de la quantité à commander (besoin - stock)
            for key in ingredients_besoins:
                besoin = ingredients_besoins[key]
                besoin['quantite_commander'] = max(Decimal('0'), besoin['quantite'] - besoin['stock_actuel'])

            return JsonResponse(ingredients_besoins)

        # -------------------------------------------------
        # Créer les commandes avec toutes les lignes et prix modifiés
        # -------------------------------------------------
        elif action == 'creer_commandes':
            try:
                commandes_data = json.loads(request.POST.get('commandes_data', '[]'))
                date_livraison = request.POST.get('date_livraison')
                notes_generales = request.POST.get('notes', '')

                if not date_livraison:
                    return JsonResponse({'error': 'Date de livraison requise'}, status=400)

                commandes_creees = []

                with transaction.atomic():
                    for cmd_data in commandes_data:
                        try:
                            fournisseur = Fournisseur.objects.get(id=cmd_data['fournisseur_id'])
                        except Fournisseur.DoesNotExist:
                            return JsonResponse({'error': f'Fournisseur {cmd_data["fournisseur_id"]} introuvable'}, status=404)

                        commande = Commande.objects.create(
                            fournisseur=fournisseur,
                            date_livraison_prevue=date_livraison,
                            notes=notes_generales,
                            statut='CONFIRMEE'
                        )

                        # Ajouter délai si fourni
                        delai = cmd_data.get('delai')
                        if delai:
                            commande.notes = (commande.notes or '') + f"\nDélai de livraison: {delai} jours"
                            commande.save()

                        for ligne_data in cmd_data.get('lignes', []):
                            try:
                                ingredient = Ingredient.objects.get(id=ligne_data['ingredient_id'])
                            except Ingredient.DoesNotExist:
                                return JsonResponse({'error': f'Ingrédient {ligne_data["ingredient_id"]} introuvable'}, status=404)

                            # Utiliser le prix fourni par le client (soit original, soit modifié)
                            prix_unitaire = ligne_data.get('prix_unitaire')
                            
                            # Si pas de prix fourni, chercher dans le catalogue
                            if prix_unitaire is None:
                                catalogue = CatalogueFournisseur.objects.filter(
                                    fournisseur=fournisseur,
                                    ingredient=ingredient,
                                    actif=True
                                ).first()
                                prix_unitaire = Decimal(str(catalogue.prix)) if catalogue and catalogue.prix is not None else Decimal('0')
                            else:
                                prix_unitaire = Decimal(str(prix_unitaire))
                                
                                # Optionnel: ajouter une note si le prix a été modifié
                                catalogue = CatalogueFournisseur.objects.filter(
                                    fournisseur=fournisseur,
                                    ingredient=ingredient,
                                    actif=True
                                ).first()
                                if catalogue and catalogue.prix != prix_unitaire:
                                    note_prix = f"Prix modifié (original: {catalogue.prix}€)"
                                    if ligne_data.get('note'):
                                        ligne_data['note'] = f"{ligne_data['note']}\n{note_prix}"
                                    else:
                                        ligne_data['note'] = note_prix

                            # Conversion sécurisée en Decimal
                            quantite_decimal = Decimal(str(ligne_data.get('quantite', '0')))

                            # Création de la ligne de commande
                            try:
                                ligne_commande = LigneCommande.objects.create(
                                    commande=commande,
                                    ingredient=ingredient,
                                    quantite=quantite_decimal,
                                    prix_unitaire=prix_unitaire,
                                    notes=ligne_data.get('note', '')
                                )
                            except Exception as e:
                                return JsonResponse({'error': f"Erreur création LigneCommande: {str(e)}"}, status=500)

                        # Calcul total commande
                        total = sum(l.quantite * l.prix_unitaire for l in commande.lignes.all())
                        commande.total = total
                        commande.save()

                        commandes_creees.append(commande.numero if hasattr(commande, 'numero') else str(commande.id))

                return JsonResponse({'success': True, 'commandes': commandes_creees})

            except Exception as e:
                return JsonResponse({'error': str(e)}, status=500)

    # -------------------------------------------------
    # GET request - afficher le formulaire
    # -------------------------------------------------
    ingredients = Ingredient.objects.select_related('unite_mesure').prefetch_related(
        Prefetch('catalogues', queryset=CatalogueFournisseur.objects.filter(actif=True).select_related('fournisseur'))
    )

    # Ajouter flag pour réapprovisionnement
    for ingredient in ingredients:
        ingredient.besoin_reappro = (
            ingredient.stock_reel is not None and
            ingredient.stock_alerte is not None and
            ingredient.stock_reel <= ingredient.stock_alerte
        )

    return render(request, 'recipes/commande_create.html', {
        'ingredients': ingredients
    })
from django.views.generic import ListView
from django.db.models import Count, Q
from .models import Commande, Fournisseur

from django.views.generic import ListView
from django.db.models import Count, Q, Sum, Prefetch
from datetime import date
from .models import Commande, Fournisseur, LigneCommande

from django.views.generic import ListView
from django.db.models import Count, Q, Sum, Prefetch
from datetime import date
from decimal import Decimal
from .models import Commande, Fournisseur, LigneCommande

from django.views.generic import ListView
from django.db.models import Count, Q, Sum, Prefetch
from datetime import date
from decimal import Decimal
from .models import Commande, Fournisseur, LigneCommande

from django.views.generic import ListView
from django.db.models import Count, Q, Sum
from datetime import date
from decimal import Decimal
from .models import Commande, Fournisseur, LigneCommande

from django.views.generic import ListView
from django.db.models import Count, Q, Sum
from datetime import date
from decimal import Decimal
from .models import Commande, Fournisseur

from django.views.generic import ListView
from django.db.models import Q
from datetime import date
from decimal import Decimal
from .models import Commande, Fournisseur

from django.views.generic import ListView
from django.db.models import Q
from datetime import date
from decimal import Decimal
from .models import Commande, Fournisseur

from django.views.generic import ListView
from django.db.models import Q, Sum, F, DecimalField
from decimal import Decimal
from datetime import date

class CommandeListView(ListView):
    model = Commande
    template_name = 'recipes/commande_list.html'
    context_object_name = 'commandes'
    paginate_by = 10
    
    def get_queryset(self):
        # Prefetch les lignes et ingrédients pour éviter les requêtes N+1
        # Annotate avec le calcul du montant total
        queryset = Commande.objects.select_related('fournisseur').prefetch_related(
            'lignes__ingredient'
        ).annotate(
            total_calcule=Sum(
                F('lignes__quantite') * F('lignes__prix_unitaire'),
                output_field=DecimalField(max_digits=10, decimal_places=2)
            )
        )
        
        # Recherche dans numero, fournisseur ET ingrédients
        search = self.request.GET.get('search', '').strip()
        if search:
            queryset = queryset.filter(
                Q(numero__icontains=search) |
                Q(fournisseur__nom__icontains=search) |
                Q(lignes__ingredient__nom__icontains=search)  # Recherche par ingrédient
            ).distinct()  # distinct() important pour éviter les doublons
        
        # Filtre fournisseur
        fournisseur_id = self.request.GET.get('fournisseur')
        if fournisseur_id:
            queryset = queryset.filter(fournisseur_id=fournisseur_id)
        
        # Filtre statut
        statut = self.request.GET.get('statut')
        if statut:
            queryset = queryset.filter(statut=statut)
        
        # Filtres par date
        date_debut = self.request.GET.get('date_debut')
        if date_debut:
            queryset = queryset.filter(date_creation__date__gte=date_debut)
        
        date_fin = self.request.GET.get('date_fin')
        if date_fin:
            queryset = queryset.filter(date_creation__date__lte=date_fin)
        
        return queryset.order_by('-date_creation')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Recréer le même queryset pour les stats (sans la pagination)
        stats_queryset = Commande.objects.select_related('fournisseur').prefetch_related('lignes__ingredient')
        
        # Appliquer les mêmes filtres pour les stats
        search = self.request.GET.get('search', '').strip()
        if search:
            stats_queryset = stats_queryset.filter(
                Q(numero__icontains=search) |
                Q(fournisseur__nom__icontains=search) |
                Q(lignes__ingredient__nom__icontains=search)
            ).distinct()
        
        fournisseur_id = self.request.GET.get('fournisseur')
        if fournisseur_id:
            stats_queryset = stats_queryset.filter(fournisseur_id=fournisseur_id)
        
        statut = self.request.GET.get('statut')
        if statut:
            stats_queryset = stats_queryset.filter(statut=statut)
        
        date_debut = self.request.GET.get('date_debut')
        if date_debut:
            stats_queryset = stats_queryset.filter(date_creation__date__gte=date_debut)
        
        date_fin = self.request.GET.get('date_fin')
        if date_fin:
            stats_queryset = stats_queryset.filter(date_creation__date__lte=date_fin)
        
        # Calculer les stats sur le queryset filtré
        context['total_count'] = stats_queryset.count()
        context['stats'] = {
            'en_cours': stats_queryset.filter(statut='EN_COURS').count(),
            'confirmees': stats_queryset.filter(statut='CONFIRMEE').count(),
            'livrees': stats_queryset.filter(statut='LIVREE').count(),
            'annulees': stats_queryset.filter(statut='ANNULEE').count(),
        }
        
        # Pour chaque commande, calculer les données supplémentaires
        for commande in context['commandes']:
            # Calculer le total à partir des lignes si nécessaire
            if commande.total_calcule:
                commande.total = commande.total_calcule
            elif commande.montant_total and commande.montant_total > 0:
                commande.total = commande.montant_total
            else:
                # Calculer le total manuellement si pas présent
                total = Decimal('0')
                for ligne in commande.lignes.all():
                    if hasattr(ligne, 'sous_total'):
                        total += ligne.sous_total
                    elif hasattr(ligne, 'prix_unitaire') and hasattr(ligne, 'quantite'):
                        total += ligne.prix_unitaire * ligne.quantite
                commande.total = total
            
            # Nombre de lignes
            commande.nb_lignes = commande.lignes.count()
            
            # Jours restants
            if commande.date_livraison_prevue:
                delta = commande.date_livraison_prevue - date.today()
                commande.jours_restants = delta.days
            else:
                commande.jours_restants = None
        
        # Fournisseurs pour le filtre
        context['fournisseurs'] = Fournisseur.objects.all().order_by('nom')
        
        return context




# Vue calendrier des livraisons
@login_required
def commande_calendar(request):
    selected_date = request.GET.get('date', date.today())
    if isinstance(selected_date, str):
        selected_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
    
    commandes = Commande.objects.filter(
        date_livraison_prevue=selected_date
    ).select_related('fournisseur').prefetch_related('lignes__ingredient')
    
    return render(request, 'recipes/commande_calendar.html', {
        'commandes': commandes,
        'selected_date': selected_date
    })

# Vue détail et validation de commande
@login_required
def commande_detail(request, pk):
    commande = get_object_or_404(Commande.objects.select_related('fournisseur').prefetch_related('lignes__ingredient'), pk=pk)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'valider':
            signature = request.POST.get('signature')
            commande.statut = 'LIVREE'
            commande.date_livraison_reelle = date.today()
            commande.signature = signature
            commande.validee_par = request.user
            commande.save()
            
            # Mettre à jour le stock
            for ligne in commande.lignes.all():
                ligne.ingredient.stock_reel += ligne.quantite
                ligne.ingredient.save()
            
            messages.success(request, 'Commande validée et stock mis à jour.')
            return redirect('commande_detail', pk=commande.pk)
    
    return render(request, 'recipes/commande_detail.html', {
        'commande': commande
    })

# API endpoints pour AJAX
@login_required
def api_ingredients_search(request):
    query = request.GET.get('q', '')
    ingredients = Ingredient.objects.filter(nom__icontains=query)[:20]
    data = [{'id': i.id, 'nom': i.nom, 'unite': i.unite_mesure.symbole} for i in ingredients]
    return JsonResponse(data, safe=False)

@login_required 
def api_fournisseurs_ingredient(request, ingredient_id):
    catalogues = CatalogueFournisseur.objects.filter(
        ingredient_id=ingredient_id,
        actif=True
    ).select_related('fournisseur')
    
    data = [{
        'fournisseur_id': c.fournisseur.id,
        'fournisseur_nom': c.fournisseur.nom,
        'prix': str(c.prix),
        'delai': c.delai_livraison
    } for c in catalogues]
    
    return JsonResponse(data, safe=False)


# views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from django.urls import reverse_lazy
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Sum, F, Prefetch
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.utils import timezone
from django.db import transaction
from datetime import datetime, date
from decimal import Decimal
import json
import csv
import pandas as pd
from io import BytesIO

from .models import (
    UniteMesure, Departement, Fournisseur, Ingredient, 
    CatalogueFournisseur, Recette, SousRecette, 
    RecetteIngredient, SousRecetteIngredient,
    Commande, LigneCommande, ProductionPrevue
)
from .forms import (
    UniteMesureForm, DepartementForm, FournisseurForm,
    IngredientForm, CatalogueFournisseurForm, RecetteForm,
    SousRecetteForm, CommandeForm, RecetteIngredientFormSet,
    SousRecetteIngredientFormSet, ImportCatalogueForm
)

# ===============================================
# Dashboard
# ===============================================


# ===============================================
# Vues CRUD pour UniteMesure
# ===============================================
class UniteMesureListView(LoginRequiredMixin, ListView):
    model = UniteMesure
    template_name = 'recipes/unite_mesure_list.html'
    context_object_name = 'unites'
    paginate_by = 20

class UniteMesureCreateView(LoginRequiredMixin, CreateView):
    model = UniteMesure
    form_class = UniteMesureForm
    template_name = 'recipes/unite_mesure_form.html'
    success_url = reverse_lazy('unite_mesure_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Unité de mesure créée avec succès.')
        return super().form_valid(form)

class UniteMesureUpdateView(LoginRequiredMixin, UpdateView):
    model = UniteMesure
    form_class = UniteMesureForm
    template_name = 'recipes/unite_mesure_form.html'
    success_url = reverse_lazy('unite_mesure_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Unité de mesure modifiée avec succès.')
        return super().form_valid(form)

class UniteMesureDeleteView(LoginRequiredMixin, DeleteView):
    model = UniteMesure
    template_name = 'recipes/unite_mesure_confirm_delete.html'
    success_url = reverse_lazy('unite_mesure_list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(self.request, 'Unité de mesure supprimée avec succès.')
        return super().delete(request, *args, **kwargs)

# ===============================================
# Vues CRUD pour Departement
# ===============================================
class DepartementListView(LoginRequiredMixin, ListView):
    model = Departement
    template_name = 'recipes/departement_list.html'
    context_object_name = 'departements'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = super().get_queryset()
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(nom__icontains=search) |
                Q(description__icontains=search)
            )
        return queryset

class DepartementCreateView(LoginRequiredMixin, CreateView):
    model = Departement
    form_class = DepartementForm
    template_name = 'recipes/departement_form.html'
    success_url = reverse_lazy('departement_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Département créé avec succès.')
        return super().form_valid(form)

class DepartementUpdateView(LoginRequiredMixin, UpdateView):
    model = Departement
    form_class = DepartementForm
    template_name = 'recipes/departement_form.html'
    success_url = reverse_lazy('departement_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Département modifié avec succès.')
        return super().form_valid(form)

class DepartementDeleteView(LoginRequiredMixin, DeleteView):
    model = Departement
    template_name = 'recipes/departement_confirm_delete.html'
    success_url = reverse_lazy('departement_list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(self.request, 'Département supprimé avec succès.')
        return super().delete(request, *args, **kwargs)

# ===============================================
# Vues CRUD pour Fournisseur
# ===============================================
class FournisseurListView(LoginRequiredMixin, ListView):
    model = Fournisseur
    template_name = 'recipes/fournisseur_list.html'
    context_object_name = 'fournisseurs'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = super().get_queryset()
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(nom__icontains=search) |
                Q(email__icontains=search) |
                Q(contact_principal__icontains=search)
            )
        return queryset

class FournisseurCreateView(LoginRequiredMixin, CreateView):
    model = Fournisseur
    form_class = FournisseurForm
    template_name = 'recipes/fournisseur_form.html'
    success_url = reverse_lazy('fournisseur_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Fournisseur créé avec succès.')
        return super().form_valid(form)

class FournisseurUpdateView(LoginRequiredMixin, UpdateView):
    model = Fournisseur
    form_class = FournisseurForm
    template_name = 'recipes/fournisseur_form.html'
    success_url = reverse_lazy('fournisseur_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Fournisseur modifié avec succès.')
        return super().form_valid(form)

class FournisseurDeleteView(LoginRequiredMixin, DeleteView):
    model = Fournisseur
    template_name = 'recipes/fournisseur_confirm_delete.html'
    success_url = reverse_lazy('fournisseur_list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(self.request, 'Fournisseur supprimé avec succès.')
        return super().delete(request, *args, **kwargs)

# ===============================================
# Vues CRUD pour Ingredient
# ===============================================
class IngredientListView(LoginRequiredMixin, ListView):
    model = Ingredient
    template_name = 'recipes/ingredient_list.html'
    context_object_name = 'ingredients'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = super().get_queryset()
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(nom__icontains=search) |
                Q(description__icontains=search)
            )
        return queryset.select_related('unite_mesure')

class IngredientCreateView(LoginRequiredMixin, CreateView):
    model = Ingredient
    form_class = IngredientForm
    template_name = 'recipes/ingredient_form.html'
    success_url = reverse_lazy('ingredient_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Ingrédient créé avec succès.')
        return super().form_valid(form)

class IngredientUpdateView(LoginRequiredMixin, UpdateView):
    model = Ingredient
    form_class = IngredientForm
    template_name = 'recipes/ingredient_form.html'
    success_url = reverse_lazy('ingredient_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Ingrédient modifié avec succès.')
        return super().form_valid(form)

class IngredientDeleteView(LoginRequiredMixin, DeleteView):
    model = Ingredient
    template_name = 'recipes/ingredient_confirm_delete.html'
    success_url = reverse_lazy('ingredient_list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(self.request, 'Ingrédient supprimé avec succès.')
        return super().delete(request, *args, **kwargs)

# ===============================================
# Vues CRUD pour CatalogueFournisseur
# ===============================================
class CatalogueFournisseurListView(LoginRequiredMixin, ListView):
    model = CatalogueFournisseur
    template_name = 'recipes/catalogue_list.html'
    context_object_name = 'catalogues'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = super().get_queryset()
        fournisseur = self.request.GET.get('fournisseur')
        ingredient = self.request.GET.get('ingredient')
        actif = self.request.GET.get('actif')
        
        if fournisseur:
            queryset = queryset.filter(fournisseur_id=fournisseur)
        if ingredient:
            queryset = queryset.filter(ingredient_id=ingredient)
        if actif:
            queryset = queryset.filter(actif=actif == 'true')
            
        return queryset.select_related('fournisseur', 'ingredient', 'ingredient__unite_mesure')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['fournisseurs'] = Fournisseur.objects.all()
        context['ingredients'] = Ingredient.objects.all()
        return context

class CatalogueFournisseurCreateView(LoginRequiredMixin, CreateView):
    model = CatalogueFournisseur
    form_class = CatalogueFournisseurForm
    template_name = 'recipes/catalogue_form.html'
    success_url = reverse_lazy('catalogue_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Catalogue fournisseur créé avec succès.')
        return super().form_valid(form)

class CatalogueFournisseurUpdateView(LoginRequiredMixin, UpdateView):
    model = CatalogueFournisseur
    form_class = CatalogueFournisseurForm
    template_name = 'recipes/catalogue_form.html'
    success_url = reverse_lazy('catalogue_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Catalogue fournisseur modifié avec succès.')
        return super().form_valid(form)

class CatalogueFournisseurDeleteView(LoginRequiredMixin, DeleteView):
    model = CatalogueFournisseur
    template_name = 'recipes/catalogue_confirm_delete.html'
    success_url = reverse_lazy('catalogue_list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(self.request, 'Catalogue fournisseur supprimé avec succès.')
        return super().delete(request, *args, **kwargs)

@login_required
def import_catalogue(request):
    """Import en masse du catalogue fournisseur depuis CSV/Excel"""
    if request.method == 'POST':
        form = ImportCatalogueForm(request.POST, request.FILES)
        if form.is_valid():
            fichier = request.FILES['fichier']
            fournisseur = form.cleaned_data['fournisseur']
            date_debut = form.cleaned_data['date_debut']
            
            try:
                # Lire le fichier selon son extension
                if fichier.name.endswith('.csv'):
                    df = pd.read_csv(fichier)
                else:  # Excel
                    df = pd.read_excel(fichier)
                
                # Colonnes attendues: nom_ingredient, prix, reference, conditionnement
                created_count = 0
                updated_count = 0
                
                with transaction.atomic():
                    for index, row in df.iterrows():
                        # Chercher ou créer l'ingrédient
                        ingredient, _ = Ingredient.objects.get_or_create(
                            nom=row['nom_ingredient'],
                            defaults={'unite_mesure_id': 1}  # Par défaut
                        )
                        
                        # Désactiver les anciens prix
                        CatalogueFournisseur.objects.filter(
                            fournisseur=fournisseur,
                            ingredient=ingredient,
                            actif=True
                        ).update(actif=False, date_fin=date_debut)
                        
                        # Créer le nouveau prix
                        CatalogueFournisseur.objects.create(
                            fournisseur=fournisseur,
                            ingredient=ingredient,
                            prix=row['prix'],
                            date_debut=date_debut,
                            reference_fournisseur=row.get('reference', ''),
                            conditionnement=row.get('conditionnement', ''),
                            actif=True
                        )
                        created_count += 1
                
                messages.success(request, f'{created_count} prix importés avec succès.')
                return redirect('catalogue_list')
                
            except Exception as e:
                messages.error(request, f'Erreur lors de l\'import: {str(e)}')
    else:
        form = ImportCatalogueForm()
    
    return render(request, 'recipes/catalogue_import.html', {'form': form})

# ===============================================
# Vues CRUD pour Recette avec gestion des ingrédients
# ===============================================
class RecetteListView(LoginRequiredMixin, ListView):
    model = Recette
    template_name = 'recipes/recette_list.html'
    context_object_name = 'recettes'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = super().get_queryset()
        search = self.request.GET.get('search')
        departement = self.request.GET.get('departement')
        
        if search:
            queryset = queryset.filter(
                Q(nom__icontains=search) |
                Q(description__icontains=search)
            )
        if departement:
            queryset = queryset.filter(departement_id=departement)
            
        return queryset.select_related('departement').prefetch_related('recette_ingredients__ingredient')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['departements'] = Departement.objects.all()
        return context

class RecetteDetailView(LoginRequiredMixin, DetailView):
    model = Recette
    template_name = 'recipes/recette_detail.html'
    context_object_name = 'recette'
    
    def get_queryset(self):
        return super().get_queryset().prefetch_related(
            'recette_ingredients__ingredient__unite_mesure',
            'sous_recettes__sousrecette_ingredients__ingredient__unite_mesure'
        )

@login_required
def recette_create(request):
    if request.method == 'POST':
        form = RecetteForm(request.POST, request.FILES)
        formset = RecetteIngredientFormSet(request.POST, prefix='ingredients')
        
        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                recette = form.save()
                formset.instance = recette
                formset.save()
                recette.calculer_cout()
                messages.success(request, 'Recette créée avec succès.')
                return redirect('recette_detail', pk=recette.pk)
    else:
        form = RecetteForm()
        formset = RecetteIngredientFormSet(prefix='ingredients')
    
    return render(request, 'recipes/recette_form.html', {
        'form': form,
        'formset': formset,
        'title': 'Créer une recette'
    })

@login_required
def recette_update(request, pk):
    recette = get_object_or_404(Recette, pk=pk)
    
    if request.method == 'POST':
        form = RecetteForm(request.POST, request.FILES, instance=recette)
        formset = RecetteIngredientFormSet(request.POST, instance=recette, prefix='ingredients')
        
        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                recette = form.save()
                formset.save()
                recette.calculer_cout()
                messages.success(request, 'Recette modifiée avec succès.')
                return redirect('recette_detail', pk=recette.pk)
    else:
        form = RecetteForm(instance=recette)
        formset = RecetteIngredientFormSet(instance=recette, prefix='ingredients')
    
    return render(request, 'recipes/recette_form.html', {
        'form': form,
        'formset': formset,
        'recette': recette,
        'title': 'Modifier la recette'
    })

class RecetteDeleteView(LoginRequiredMixin, DeleteView):
    model = Recette
    template_name = 'recipes/recette_confirm_delete.html'
    success_url = reverse_lazy('recette_list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(self.request, 'Recette supprimée avec succès.')
        return super().delete(request, *args, **kwargs)

@login_required
def recette_duplicate(request, pk):
    """Dupliquer une recette existante"""
    recette = get_object_or_404(Recette, pk=pk)
    
    with transaction.atomic():
        # Copier la recette
        new_recette = Recette.objects.create(
            nom=f"{recette.nom} (copie)",
            description=recette.description,
            explication_fabrication=recette.explication_fabrication,
            departement=recette.departement,
            temps_preparation=recette.temps_preparation,
            temps_cuisson=recette.temps_cuisson,
            portions=recette.portions
        )
        
        # Copier les ingrédients
        for ri in recette.recette_ingredients.all():
            RecetteIngredient.objects.create(
                recette=new_recette,
                ingredient=ri.ingredient,
                quantite=ri.quantite
            )
        
        # Copier les sous-recettes
        for sr in recette.sous_recettes.all():
            new_sr = SousRecette.objects.create(
                nom=sr.nom,
                recette_parent=new_recette,
                explication_fabrication=sr.explication_fabrication,
                quantite=sr.quantite
            )
            
            for sri in sr.sousrecette_ingredients.all():
                SousRecetteIngredient.objects.create(
                    sous_recette=new_sr,
                    ingredient=sri.ingredient,
                    quantite=sri.quantite
                )
        
        new_recette.calculer_cout()
    
    messages.success(request, f'Recette "{recette.nom}" dupliquée avec succès.')
    return redirect('recette_update', pk=new_recette.pk)

# ===============================================
# Vues pour Sous-Recettes
# ===============================================
@login_required
def sous_recette_create(request, recette_pk):
    recette = get_object_or_404(Recette, pk=recette_pk)
    
    if request.method == 'POST':
        form = SousRecetteForm(request.POST)
        formset = SousRecetteIngredientFormSet(request.POST, prefix='ingredients')
        
        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                sous_recette = form.save(commit=False)
                sous_recette.recette_parent = recette
                sous_recette.save()
                formset.instance = sous_recette
                formset.save()
                recette.calculer_cout()
                messages.success(request, 'Sous-recette créée avec succès.')
                return redirect('recette_detail', pk=recette.pk)
    else:
        form = SousRecetteForm()
        formset = SousRecetteIngredientFormSet(prefix='ingredients')
    
    return render(request, 'recipes/sous_recette_form.html', {
        'form': form,
        'formset': formset,
        'recette': recette,
        'title': 'Créer une sous-recette'
    })

@login_required
def sous_recette_update(request, pk):
    sous_recette = get_object_or_404(SousRecette, pk=pk)
    
    if request.method == 'POST':
        form = SousRecetteForm(request.POST, instance=sous_recette)
        formset = SousRecetteIngredientFormSet(request.POST, instance=sous_recette, prefix='ingredients')
        
        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                sous_recette = form.save()
                formset.save()
                sous_recette.recette_parent.calculer_cout()
                messages.success(request, 'Sous-recette modifiée avec succès.')
                return redirect('recette_detail', pk=sous_recette.recette_parent.pk)
    else:
        form = SousRecetteForm(instance=sous_recette)
        formset = SousRecetteIngredientFormSet(instance=sous_recette, prefix='ingredients')
    
    return render(request, 'recipes/sous_recette_form.html', {
        'form': form,
        'formset': formset,
        'sous_recette': sous_recette,
        'title': 'Modifier la sous-recette'
    })

@login_required
def sous_recette_delete(request, pk):
    sous_recette = get_object_or_404(SousRecette, pk=pk)
    recette_pk = sous_recette.recette_parent.pk
    
    if request.method == 'POST':
        sous_recette.delete()
        messages.success(request, 'Sous-recette supprimée avec succès.')
        return redirect('recette_detail', pk=recette_pk)
    
    return render(request, 'recipes/sous_recette_confirm_delete.html', {
        'sous_recette': sous_recette
    })

# ===============================================
# Vue de création de commande (déjà présente)
# ===============================================
@login_required
def commande_create(request):
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        action = request.POST.get('action')
        
        if action == 'get_departements':
            ingredient_ids = json.loads(request.POST.get('ingredient_ids', '[]'))
            departements = Departement.objects.filter(
                recettes__recette_ingredients__ingredient_id__in=ingredient_ids
            ).distinct().values('id', 'nom')
            return JsonResponse(list(departements), safe=False)
        
        elif action == 'get_recettes':
            ingredient_id = request.POST.get('ingredient_id')
            departement_id = request.POST.get('departement_id')
            recettes = Recette.objects.filter(
                departement_id=departement_id,
                recette_ingredients__ingredient_id=ingredient_id
            ).distinct().values('id', 'nom')
            return JsonResponse(list(recettes), safe=False)
        
        elif action == 'calculer_quantites':
            productions = json.loads(request.POST.get('productions', '[]'))
            ingredients_besoins = {}
            
            for prod in productions:
                recette = Recette.objects.get(id=prod['recette_id'])
                quantite = int(prod['quantite'])
                
                # Calculer les besoins en ingrédients
                for ri in recette.recette_ingredients.all():
                    key = ri.ingredient.id
                    if key not in ingredients_besoins:
                        ingredients_besoins[key] = {
                            'nom': ri.ingredient.nom,
                            'unite': ri.ingredient.unite_mesure.symbole,
                            'quantite': 0,
                            'stock_actuel': float(ri.ingredient.stock_reel),
                            'fournisseurs': list(ri.ingredient.catalogues.filter(
                                actif=True
                            ).values('fournisseur__id', 'fournisseur__nom', 'prix'))
                        }
                    ingredients_besoins[key]['quantite'] += float(ri.quantite) * quantite
                
                # Calculer pour les sous-recettes
                for sr in recette.sous_recettes.all():
                    for sri in sr.sousrecette_ingredients.all():
                        key = sri.ingredient.id
                        if key not in ingredients_besoins:
                            ingredients_besoins[key] = {
                                'nom': sri.ingredient.nom,
                                'unite': sri.ingredient.unite_mesure.symbole,
                                'quantite': 0,
                                'stock_actuel': float(sri.ingredient.stock_reel),
                                'fournisseurs': list(sri.ingredient.catalogues.filter(
                                    actif=True
                                ).values('fournisseur__id', 'fournisseur__nom', 'prix'))
                            }
                        ingredients_besoins[key]['quantite'] += float(sri.quantite) * float(sr.quantite) * quantite
            
            # Soustraire le stock actuel
            for key in ingredients_besoins:
                besoin = ingredients_besoins[key]
                besoin['quantite_commander'] = max(0, besoin['quantite'] - besoin['stock_actuel'])
            
            return JsonResponse(ingredients_besoins)
        
        elif action == 'creer_commandes':
            commandes_data = json.loads(request.POST.get('commandes_data', '[]'))
            date_livraison = request.POST.get('date_livraison')
            notes = request.POST.get('notes', '')
            
            commandes_creees = []
            
            with transaction.atomic():
                for cmd_data in commandes_data:
                    fournisseur = Fournisseur.objects.get(id=cmd_data['fournisseur_id'])
                    
                    commande = Commande.objects.create(
                        fournisseur=fournisseur,
                        date_livraison_prevue=date_livraison,
                        notes=notes,
                        statut='CONFIRMEE'
                    )
                    
                    for ligne_data in cmd_data['lignes']:
                        ingredient = Ingredient.objects.get(id=ligne_data['ingredient_id'])
                        catalogue = CatalogueFournisseur.objects.filter(
                            fournisseur=fournisseur,
                            ingredient=ingredient,
                            actif=True
                        ).first()
                        
                        LigneCommande.objects.create(
                            commande=commande,
                            ingredient=ingredient,
                            quantite=ligne_data['quantite'],
                            prix_unitaire=catalogue.prix if catalogue else 0
                        )
                    
                    commande.calculer_total()
                    commandes_creees.append(commande.numero)
            
            messages.success(request, f'Commandes créées: {", ".join(commandes_creees)}')
            return JsonResponse({'success': True, 'commandes': commandes_creees})
    
    ingredients = Ingredient.objects.select_related('unite_mesure').prefetch_related(
        Prefetch('catalogues', queryset=CatalogueFournisseur.objects.filter(actif=True).select_related('fournisseur'))
    )
    
    return render(request, 'recipes/commande_create.html', {
        'ingredients': ingredients
    })

# ===============================================
# Vue historique et gestion des commandes
# ===============================================

@login_required
def commande_calendar(request):
    selected_date = request.GET.get('date', date.today())
    if isinstance(selected_date, str):
        selected_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
    
    commandes = Commande.objects.filter(
        date_livraison_prevue=selected_date
    ).select_related('fournisseur').prefetch_related('lignes__ingredient')
    
    return render(request, 'recipes/commande_calendar.html', {
        'commandes': commandes,
        'selected_date': selected_date
    })

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from datetime import date
from decimal import Decimal
import json

@login_required
def commande_detail(request, pk):
    commande = get_object_or_404(
        Commande.objects.select_related('fournisseur', 'validee_par')
        .prefetch_related('lignes__ingredient__unite_mesure'), 
        pk=pk
    )
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'valider':
            try:
                # Récupération des données de base
                signature_data = request.POST.get('signature_data', '')
                nom_signataire = request.POST.get('nom_signataire', '')
                remarques = request.POST.get('remarques', '')
                
                # Vérification de la signature
                if not signature_data:
                    messages.error(request, 'La signature est obligatoire pour valider la commande.')
                    return redirect('commande_detail', pk=commande.pk)
                
                # Préparation du rapport de validation
                rapport_lines = []
                rapport_lines.append(f"RAPPORT DE VALIDATION - Commande {commande.numero}")
                rapport_lines.append(f"Date: {timezone.now().strftime('%d/%m/%Y %H:%M')}")
                rapport_lines.append(f"Validé par: {request.user.get_full_name() or request.user.username}")
                rapport_lines.append(f"Signataire: {nom_signataire}")
                rapport_lines.append("-" * 50)
                
                # Dictionnaire pour stocker les détails de validation
                validation_details = {
                    'date': timezone.now().isoformat(),
                    'validateur': request.user.username,
                    'signataire': nom_signataire,
                    'lignes': []
                }
                
                # Traitement de chaque ligne de commande
                total_problemes = 0
                total_differences = 0
                i = 0
                
                while f'ligne_id_{i}' in request.POST:
                    ligne_id = request.POST.get(f'ligne_id_{i}')
                    quantite_recue = request.POST.get(f'quantite_recue_{i}')
                    qualite = request.POST.get(f'qualite_{i}', 'OK')
                    
                    try:
                        ligne = commande.lignes.get(id=ligne_id)
                        
                        # Conversion en Decimal
                        quantite_recue_decimal = Decimal(quantite_recue) if quantite_recue else ligne.quantite
                        
                        # Mise à jour de la ligne
                        ligne.quantite_recue = quantite_recue_decimal
                        ligne.qualite_validation = qualite
                        ligne.date_validation = timezone.now()
                        ligne.save()
                        
                        # Calcul de la différence
                        difference = quantite_recue_decimal - ligne.quantite
                        
                        # Ajout au rapport
                        ligne_rapport = f"\n{ligne.ingredient.nom}:"
                        ligne_rapport += f"\n  - Commandé: {ligne.quantite} {ligne.ingredient.unite_mesure.symbole if ligne.ingredient.unite_mesure else ''}"
                        ligne_rapport += f"\n  - Reçu: {quantite_recue_decimal} {ligne.ingredient.unite_mesure.symbole if ligne.ingredient.unite_mesure else ''}"
                        
                        if difference != 0:
                            total_differences += 1
                            signe = "+" if difference > 0 else ""
                            ligne_rapport += f"\n  - Différence: {signe}{difference}"
                        
                        if qualite == 'PROBLEME':
                            total_problemes += 1
                            ligne_rapport += f"\n  - ⚠️ PROBLÈME DE QUALITÉ SIGNALÉ"
                        else:
                            ligne_rapport += f"\n  - ✓ Qualité OK"
                        
                        rapport_lines.append(ligne_rapport)
                        
                        # Mise à jour du stock avec la quantité réellement reçue
                        if ligne.ingredient.stock_reel is not None:
                            ligne.ingredient.stock_reel += quantite_recue_decimal
                        else:
                            ligne.ingredient.stock_reel = quantite_recue_decimal
                        ligne.ingredient.save()
                        
                        # Ajout aux détails de validation JSON
                        validation_details['lignes'].append({
                            'ingredient': ligne.ingredient.nom,
                            'quantite_commandee': float(ligne.quantite),
                            'quantite_recue': float(quantite_recue_decimal),
                            'difference': float(difference),
                            'qualite': qualite,
                            'unite': ligne.ingredient.unite_mesure.symbole if ligne.ingredient.unite_mesure else ''
                        })
                        
                    except Exception as e:
                        messages.warning(request, f"Erreur lors du traitement de la ligne {i}: {str(e)}")
                    
                    i += 1
                
                # Ajout du résumé au rapport
                rapport_lines.append("-" * 50)
                rapport_lines.append("\nRÉSUMÉ:")
                rapport_lines.append(f"- Total articles: {i}")
                rapport_lines.append(f"- Différences de quantité: {total_differences}")
                rapport_lines.append(f"- Problèmes de qualité: {total_problemes}")
                
                if remarques:
                    rapport_lines.append(f"\nREMARQUES:\n{remarques}")
                
                # Compilation du rapport final
                notes_validation = "\n".join(rapport_lines)
                
                # Mise à jour de la commande
                commande.statut = 'LIVREE'
                commande.date_livraison_reelle = date.today()
                commande.signature = signature_data
                commande.validee_par = request.user
                commande.notes_validation = notes_validation
                commande.validation_details = validation_details
                commande.save()
                
                # Message de succès détaillé
                if total_problemes > 0 or total_differences > 0:
                    messages.warning(
                        request, 
                        f'Commande validée avec {total_differences} différence(s) de quantité et {total_problemes} problème(s) de qualité signalé(s).'
                    )
                else:
                    messages.success(request, 'Commande validée avec succès. Stock mis à jour.')
                
                return redirect('commande_detail', pk=commande.pk)
                
            except Exception as e:
                messages.error(request, f'Erreur lors de la validation: {str(e)}')
                return redirect('commande_detail', pk=commande.pk)
    
    # Calcul des jours restants pour la livraison
    if commande.date_livraison_prevue and commande.statut not in ['LIVREE', 'ANNULEE']:
        delta = commande.date_livraison_prevue - date.today()
        commande.jours_restants = delta.days
    else:
        commande.jours_restants = None
    
    return render(request, 'recipes/commande_detail.html', {
        'commande': commande
    })

@login_required
def commande_print(request, pk):
    """Imprimer une commande"""
    commande = get_object_or_404(Commande, pk=pk)
    return render(request, 'recipes/commande_print.html', {
        'commande': commande
    })

@login_required
def commande_cancel(request, pk):
    """Annuler une commande"""
    commande = get_object_or_404(Commande, pk=pk)
    
    if commande.statut not in ['LIVREE', 'ANNULEE']:
        commande.statut = 'ANNULEE'
        commande.save()
        messages.success(request, f'Commande {commande.numero} annulée.')
    else:
        messages.error(request, 'Cette commande ne peut pas être annulée.')
    
    return redirect('commande_list')

# ===============================================
# API endpoints pour AJAX
# ===============================================
@login_required
def api_ingredients_search(request):
    query = request.GET.get('q', '')
    ingredients = Ingredient.objects.filter(nom__icontains=query)[:20]
    data = [{'id': i.id, 'nom': i.nom, 'unite': i.unite_mesure.symbole} for i in ingredients]
    return JsonResponse(data, safe=False)

@login_required 
def api_fournisseurs_ingredient(request, ingredient_id):
    catalogues = CatalogueFournisseur.objects.filter(
        ingredient_id=ingredient_id,
        actif=True
    ).select_related('fournisseur')
    
    data = [{
        'fournisseur_id': c.fournisseur.id,
        'fournisseur_nom': c.fournisseur.nom,
        'prix': str(c.prix),
        'delai': c.delai_livraison
    } for c in catalogues]
    
    return JsonResponse(data, safe=False)

@login_required
def api_recettes_departements(request):
    """Obtenir les recettes d'un département"""
    departement_id = request.GET.get('departement_id')
    recettes = Recette.objects.filter(departement_id=departement_id).values('id', 'nom')
    return JsonResponse(list(recettes), safe=False)

@login_required
def api_stock_update(request):
    """Mise à jour rapide du stock d'un ingrédient"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Validation et nettoyage des données
            ingredient_id = data.get('ingredient_id')
            nouveau_stock = data.get('nouveau_stock')
            
            # Convertir en nombre et valider
            try:
                nouveau_stock = float(str(nouveau_stock).replace(',', '.').strip())
            except (ValueError, TypeError):
                return JsonResponse({
                    'success': False, 
                    'error': 'Le stock doit être un nombre valide'
                }, status=400)
            
            # Vérifier que le stock n'est pas négatif
            if nouveau_stock < 0:
                return JsonResponse({
                    'success': False, 
                    'error': 'Le stock ne peut pas être négatif'
                }, status=400)
            
            # Récupérer et mettre à jour l'ingrédient
            ingredient = get_object_or_404(Ingredient, id=ingredient_id)
            ingredient.stock_reel = nouveau_stock
            ingredient.save()
            
            return JsonResponse({
                'success': True, 
                'nouveau_stock': float(ingredient.stock_reel)
            })
            
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False, 
                'error': 'Données JSON invalides'
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'success': False, 
                'error': str(e)
            }, status=500)
    
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'}, status=405)

# ===============================================
# Exports
# ===============================================
@login_required
def export_ingredients(request):
    """Export CSV des ingrédients"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="ingredients.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Nom', 'Unité', 'Stock Réel', 'Stock Alerte', 'Description'])
    
    for ingredient in Ingredient.objects.all():
        writer.writerow([
            ingredient.nom,
            ingredient.unite_mesure.symbole,
            ingredient.stock_reel,
            ingredient.stock_alerte,
            ingredient.description
        ])
    
    return response

@login_required
def export_recettes(request):
    """Export Excel des recettes avec leurs ingrédients"""
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    
    # Feuille des recettes
    recettes_data = []
    for recette in Recette.objects.all():
        recettes_data.append({
            'Nom': recette.nom,
            'Département': recette.departement.nom if recette.departement else '',
            'Temps Préparation': recette.temps_preparation,
            'Temps Cuisson': recette.temps_cuisson,
            'Portions': recette.portions,
            'Coût Calculé': float(recette.cout_calcule)
        })
    
    df_recettes = pd.DataFrame(recettes_data)
    df_recettes.to_excel(writer, sheet_name='Recettes', index=False)
    
    # Feuille des ingrédients par recette
    ingredients_data = []
    for recette in Recette.objects.all():
        for ri in recette.recette_ingredients.all():
            ingredients_data.append({
                'Recette': recette.nom,
                'Ingrédient': ri.ingredient.nom,
                'Quantité': float(ri.quantite),
                'Unité': ri.ingredient.unite_mesure.symbole
            })
    
    df_ingredients = pd.DataFrame(ingredients_data)
    df_ingredients.to_excel(writer, sheet_name='Ingrédients', index=False)
    
    writer.close()
    output.seek(0)
    
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="recettes.xlsx"'
    return response

@login_required
def export_commandes(request):
    """Export des commandes sur une période"""
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    
    queryset = Commande.objects.all()
    if date_debut:
        queryset = queryset.filter(date_creation__gte=date_debut)
    if date_fin:
        queryset = queryset.filter(date_creation__lte=date_fin)
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="commandes.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Numéro', 'Fournisseur', 'Date Création', 'Date Livraison Prévue', 'Statut', 'Montant Total'])
    
    for commande in queryset:
        writer.writerow([
            commande.numero,
            commande.fournisseur.nom,
            commande.date_creation.strftime('%Y-%m-%d'),
            commande.date_livraison_prevue,
            commande.get_statut_display(),
            float(commande.montant_total)
        ])
    
    return response



class CommandeDeleteView(DeleteView):
    model = Commande
    success_url = reverse_lazy('commande_list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Commande supprimée avec succès!')
        return super().delete(request, *args, **kwargs)
    

# Ajouter ces vues à votre views.py

from django.http import JsonResponse
from django.views import View

class FournisseurIngredientsAPIView(View):
    """API pour récupérer les ingrédients disponibles chez un fournisseur"""
    
    def get(self, request, fournisseur_id):
        try:
            fournisseur = Fournisseur.objects.get(id=fournisseur_id)
            
            # Récupérer tous les ingrédients de ce fournisseur via CatalogueFournisseur
            ingredients_data = []
            catalogues = CatalogueFournisseur.objects.filter(fournisseur=fournisseur, actif=True).select_related('ingredient__unite_mesure')
            
            for catalogue in catalogues:
                ingredient = catalogue.ingredient
                ingredients_data.append({
                    'ingredient_id': ingredient.id,
                    'ingredient_nom': ingredient.nom,
                    'prix': float(catalogue.prix),
                    'unite': ingredient.unite_mesure.symbole if ingredient.unite_mesure else '',
                    'stock_actuel': float(ingredient.stock_reel),
                    'delai_livraison': catalogue.delai_livraison if hasattr(catalogue, 'delai_livraison') else 2
                })
            
            return JsonResponse(ingredients_data, safe=False)
            
        except Fournisseur.DoesNotExist:
            return JsonResponse({'error': 'Fournisseur non trouvé'}, status=404)

# Mise à jour de CommandeUpdateView
# Le problème est à la ligne 9281 dans votre code
# Voici la correction pour CommandeUpdateView

from decimal import Decimal

from django.views.generic import UpdateView
from django.db.models import Sum
from django.shortcuts import redirect
from django.urls import reverse_lazy
from django.contrib import messages
from decimal import Decimal
from .models import (
    Commande, 
    Fournisseur, 
    LigneCommande,
    CatalogueFournisseur
)

class CommandeUpdateView(UpdateView):
    model = Commande
    template_name = 'recipes/commande_edit.html'
    fields = []
    success_url = reverse_lazy('commande_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['fournisseurs'] = Fournisseur.objects.all().order_by('nom')
        
        # Calculer le total pour l'affichage
        self.object.total = self.object.lignes.aggregate(
            total=Sum('sous_total')
        )['total'] or Decimal('0')
        
        return context
    
    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        
        # Mise à jour des informations générales
        self.object.fournisseur_id = request.POST.get('fournisseur')
        self.object.statut = request.POST.get('statut')
        self.object.notes = request.POST.get('notes')
        
        date_livraison = request.POST.get('date_livraison_prevue')
        if date_livraison:
            self.object.date_livraison_prevue = date_livraison
        
        self.object.save()
        
        # Traiter les lignes de commande
        processed_lignes = set()
        
        # Parcourir toutes les données POST
        for key in request.POST:
            if key.startswith('ingredient_'):
                ligne_id = key.replace('ingredient_', '')
                ingredient_id = request.POST.get(key)
                quantite_str = request.POST.get(f'quantite_{ligne_id}')
                
                if ingredient_id and quantite_str:
                    # Convertir la quantité en Decimal
                    try:
                        quantite = Decimal(str(quantite_str).strip())
                    except:
                        quantite = Decimal('0')
                    
                    # Récupérer le prix depuis CatalogueFournisseur
                    prix_unitaire = Decimal('0')
                    
                    # Chercher le prix actif
                    catalogue = CatalogueFournisseur.objects.filter(
                        fournisseur=self.object.fournisseur,
                        ingredient_id=ingredient_id,
                        actif=True
                    ).first()
                    
                    if not catalogue:
                        # Si pas de prix actif, prendre le plus récent
                        catalogue = CatalogueFournisseur.objects.filter(
                            fournisseur=self.object.fournisseur,
                            ingredient_id=ingredient_id
                        ).order_by('-date_debut').first()
                    
                    if catalogue:
                        # Le prix est déjà un Decimal grâce à DecimalField
                        prix_unitaire = catalogue.prix
                    
                    if ligne_id.startswith('new_'):
                        # Créer une nouvelle ligne
                        # Le modèle LigneCommande calculera sous_total automatiquement
                        LigneCommande.objects.create(
                            commande=self.object,
                            ingredient_id=ingredient_id,
                            quantite=quantite,
                            prix_unitaire=prix_unitaire
                        )
                    else:
                        # Mettre à jour une ligne existante
                        try:
                            ligne = LigneCommande.objects.get(
                                id=int(ligne_id),
                                commande=self.object
                            )
                            ligne.ingredient_id = ingredient_id
                            ligne.quantite = quantite
                            ligne.prix_unitaire = prix_unitaire
                            # save() recalculera sous_total automatiquement
                            ligne.save()
                            
                            processed_lignes.add(int(ligne_id))
                        except (LigneCommande.DoesNotExist, ValueError):
                            # Si la ligne n'existe pas, l'ignorer
                            pass
        
        # Supprimer les lignes non traitées
        self.object.lignes.exclude(id__in=processed_lignes).delete()
        
        # Recalculer le total de la commande
        self.object.calculer_total()
        
        messages.success(request, 'Commande mise à jour avec succès!')
        return redirect(self.success_url)
    model = Commande
    template_name = 'recipes/commande_edit.html'
    fields = []
    success_url = reverse_lazy('commande_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['fournisseurs'] = Fournisseur.objects.all().order_by('nom')
        
        # Calculer le total pour l'affichage
        self.object.total = self.object.lignes.aggregate(
            total=Sum('sous_total')
        )['total'] or Decimal('0')
        
        return context
    
    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        
        # Mise à jour des informations générales
        self.object.fournisseur_id = request.POST.get('fournisseur')
        self.object.statut = request.POST.get('statut')
        self.object.notes = request.POST.get('notes')
        
        date_livraison = request.POST.get('date_livraison_prevue')
        if date_livraison:
            self.object.date_livraison_prevue = date_livraison
        
        self.object.save()
        
        # Traiter les lignes de commande
        processed_lignes = set()
        
        # Parcourir toutes les données POST
        for key in request.POST:
            if key.startswith('ingredient_'):
                ligne_id = key.replace('ingredient_', '')
                ingredient_id = request.POST.get(key)
                quantite_str = request.POST.get(f'quantite_{ligne_id}')
                
                if ingredient_id and quantite_str:
                    # IMPORTANT: Convertir en Decimal, pas en float
                    quantite = Decimal(str(quantite_str))
                    
                    # Récupérer le prix depuis CatalogueFournisseur
                    prix_unitaire = Decimal('0')
                    try:
                        # Chercher le prix actif pour ce fournisseur et cet ingrédient
                        catalogue = CatalogueFournisseur.objects.filter(
                            fournisseur=self.object.fournisseur,
                            ingredient_id=ingredient_id,
                            actif=True
                        ).first()
                        
                        if catalogue:
                            # S'assurer que le prix est un Decimal
                            prix_unitaire = Decimal(str(catalogue.prix))
                        else:
                            # Si pas de prix actif, essayer de trouver n'importe quel prix
                            catalogue = CatalogueFournisseur.objects.filter(
                                fournisseur=self.object.fournisseur,
                                ingredient_id=ingredient_id
                            ).order_by('-date_debut').first()
                            
                            if catalogue:
                                prix_unitaire = Decimal(str(catalogue.prix))
                    except Exception as e:
                        print(f"Erreur récupération prix: {e}")
                        prix_unitaire = Decimal('0')
                    
                    if ligne_id.startswith('new_'):
                        # Nouvelle ligne
                        LigneCommande.objects.create(
                            commande=self.object,
                            ingredient_id=ingredient_id,
                            quantite=quantite,
                            prix_unitaire=prix_unitaire
                            # sous_total sera calculé automatiquement dans save()
                        )
                    else:
                        # Ligne existante
                        try:
                            ligne = LigneCommande.objects.get(
                                id=ligne_id,
                                commande=self.object
                            )
                            ligne.ingredient_id = ingredient_id
                            ligne.quantite = quantite
                            ligne.prix_unitaire = prix_unitaire
                            
                            # NE PAS calculer manuellement prix_total ou sous_total
                            # La méthode save() de LigneCommande le fait automatiquement
                            ligne.save()
                            
                            processed_lignes.add(int(ligne_id))
                        except (LigneCommande.DoesNotExist, ValueError) as e:
                            print(f"Erreur ligne existante: {e}")
                            pass
        
        # Supprimer les lignes qui ne sont plus dans le formulaire
        self.object.lignes.exclude(id__in=processed_lignes).delete()
        
        # Recalculer le montant total de la commande
        self.object.calculer_total()
        
        messages.success(request, 'Commande mise à jour avec succès!')
        return redirect(self.success_url)
    model = Commande
    template_name = 'recipes/commande_edit.html'
    fields = []  # On gère manuellement
    success_url = reverse_lazy('commande_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['fournisseurs'] = Fournisseur.objects.all().order_by('nom')
        
        # Calculer le total pour l'affichage
        self.object.total = self.object.lignes.aggregate(
            total=Sum('sous_total')
        )['total'] or 0
        
        return context
    
    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        
        # Mise à jour des informations générales
        self.object.fournisseur_id = request.POST.get('fournisseur')
        self.object.statut = request.POST.get('statut')
        self.object.notes = request.POST.get('notes')
        
        date_livraison = request.POST.get('date_livraison_prevue')
        if date_livraison:
            self.object.date_livraison_prevue = date_livraison
        
        self.object.save()
        
        # Traiter les lignes de commande
        processed_lignes = set()
        
        # Parcourir toutes les données POST
        for key in request.POST:
            if key.startswith('ingredient_'):
                ligne_id = key.replace('ingredient_', '')
                ingredient_id = request.POST.get(key)
                quantite = request.POST.get(f'quantite_{ligne_id}')
                
                if ingredient_id and quantite:
                    quantite = float(quantite)
                    
                    # Récupérer le prix depuis FournisseurIngredient
                    try:
                        fi = CatalogueFournisseur.objects.get(
                            fournisseur=self.object.fournisseur,
                            ingredient_id=ingredient_id
                        )
                        prix_unitaire = fi.prix
                    except CatalogueFournisseur.DoesNotExist:
                        # Prix par défaut si non trouvé
                        prix_unitaire = 0
                    
                    if ligne_id.startswith('new_'):
                        # Nouvelle ligne
                        LigneCommande.objects.create(
                            commande=self.object,
                            ingredient_id=ingredient_id,
                            quantite=quantite,
                            prix_unitaire=prix_unitaire,
                            prix_total=prix_unitaire * quantite
                        )
                    else:
                        # Ligne existante
                        try:
                            ligne = LigneCommande.objects.get(
                                id=ligne_id,
                                commande=self.object
                            )
                            ligne.ingredient_id = ingredient_id
                            ligne.quantite = quantite
                            ligne.prix_unitaire = prix_unitaire
                            ligne.save()
                            processed_lignes.add(int(ligne_id))
                        except LigneCommande.DoesNotExist:
                            pass
        
        # Supprimer les lignes qui ne sont plus dans le formulaire
        self.object.lignes.exclude(id__in=processed_lignes).delete()
        
        messages.success(request, 'Commande mise à jour avec succès!')
        return redirect(self.success_url)
    


# views.py - Import spécifique Hector Larivée

import pandas as pd
import re
from decimal import Decimal
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.shortcuts import render, redirect
from django.utils import timezone
from listings.models import (
    Ingredient, CatalogueFournisseur, UniteMesure, Fournisseur
)

@login_required
def import_hector_larivee(request):
    """Import spécifique pour les fichiers du fournisseur Hector Larivée"""
    
    if request.method == 'POST' and request.FILES.get('fichier'):
        fichier = request.FILES['fichier']
        date_debut = request.POST.get('date_debut', timezone.now().date())
        
        try:
            # Lire le fichier Excel (les données commencent à la ligne 3)
            df = pd.read_excel(fichier, header=2)
            df.columns = ['code_produit', 'description', 'format', 'provenance', 'prix']
            
            # Nettoyer les données
            df = df[df['code_produit'].notna()]
            df = df[~df['code_produit'].astype(str).str.contains('produit', case=False)]
            
            # Créer ou récupérer le fournisseur
            fournisseur, created = Fournisseur.objects.get_or_create(
                nom='Hector Larivée',
                defaults={
                    'email': 'commandes@hectorlarivee.com',
                    'telephone': '514-000-0000',
                    'adresse': 'Montréal, QC'
                }
            )
            
            if created:
                messages.info(request, "Fournisseur 'Hector Larivée' créé")
            
            # Créer les unités de mesure
            create_units_for_hector()
            
            # Statistiques
            stats = {
                'total_lignes': len(df),
                'ingredients_crees': 0,
                'ingredients_updates': 0,
                'prix_ajoutes': 0,
                'prix_desactives': 0,
                'erreurs': []
            }
            
            with transaction.atomic():
                for index, row in df.iterrows():
                    try:
                        # Traiter chaque ligne
                        result = process_hector_line(
                            row, fournisseur, date_debut
                        )
                        
                        if result['created']:
                            stats['ingredients_crees'] += 1
                        else:
                            stats['ingredients_updates'] += 1
                        
                        stats['prix_ajoutes'] += 1
                        stats['prix_desactives'] += result['prix_desactives']
                        
                    except Exception as e:
                        stats['erreurs'].append(
                            f"Ligne {index + 4}: {str(e)}"
                        )
            
            # Messages de succès
            messages.success(
                request,
                f"✅ Import terminé: {stats['prix_ajoutes']} prix importés"
            )
            
            if stats['ingredients_crees'] > 0:
                messages.info(
                    request,
                    f"🆕 {stats['ingredients_crees']} nouveaux ingrédients créés"
                )
            
            if stats['ingredients_updates'] > 0:
                messages.info(
                    request,
                    f"🔄 {stats['ingredients_updates']} ingrédients mis à jour"
                )
            
            if stats['erreurs']:
                for err in stats['erreurs'][:5]:
                    messages.warning(request, err)
            
            return redirect('catalogue_list')
            
        except Exception as e:
            messages.error(request, f"Erreur lors de l'import: {str(e)}")
    
    return render(request, 'recipes/import_hector.html')


def process_hector_line(row, fournisseur, date_debut):
    """Traiter une ligne du fichier Hector Larivée"""
    
    # Extraire les données
    code_produit = str(row['code_produit']).strip()
    description = str(row['description']).strip()
    format_produit = str(row['format']).strip() if pd.notna(row['format']) else ''
    prix = Decimal(str(row['prix']).replace(',', '.'))
    
    # Normaliser le nom
    nom_ingredient = normalise_nom_hector(description)
    
    # Déterminer l'unité
    unite = get_unite_from_format(format_produit)
    
    # Créer ou récupérer l'ingrédient
    ingredient, created = Ingredient.objects.get_or_create(
        nom__iexact=nom_ingredient,
        defaults={
            'nom': nom_ingredient,
            'unite_mesure': unite,
            'stock_reel': Decimal('0'),
            'stock_alerte': Decimal('10'),
            'description': f'Code fournisseur: {code_produit}'
        }
    )
    
    if not created and code_produit not in (ingredient.description or ''):
        # Ajouter le code à la description
        ingredient.description = f'{ingredient.description or ""} | Code: {code_produit}'.strip(' |')
        ingredient.save()
    
    # Désactiver les anciens prix
    anciens = CatalogueFournisseur.objects.filter(
        fournisseur=fournisseur,
        ingredient=ingredient,
        actif=True
    )
    nb_desactives = anciens.update(
        actif=False,
        date_fin=date_debut
    )
    
    # Créer le nouveau prix
    CatalogueFournisseur.objects.create(
        fournisseur=fournisseur,
        ingredient=ingredient,
        prix=prix,
        date_debut=date_debut,
        reference_fournisseur=code_produit,
        conditionnement=format_produit,
        actif=True,
        delai_livraison=2
    )
    
    return {
        'created': created,
        'prix_desactives': nb_desactives
    }


def normalise_nom_hector(description):
    """Normaliser le nom d'un produit Hector Larivée"""
    nom = description
    
    # Enlever les dimensions (ex: 19x13x10)
    nom = re.sub(r'\d+x\d+(?:x\d+)?', '', nom)
    
    # Enlever les codes comme 275C
    nom = re.sub(r'\b\d+[A-Z]\b', '', nom)
    
    # Enlever les nombres seuls au début ou à la fin
    nom = re.sub(r'^\d+\s+|\s+\d+$', '', nom)
    
    # Capitaliser correctement
    nom = nom.strip().title()
    
    # Corrections spécifiques pour le français
    corrections = {
        'Boite': 'Boîte',
        'Etiquette': 'Étiquette',
        'Etiqu': 'Étiquette',
        'Temperature': 'Température',
        'Temp': 'Température',
        'Enregistreur De': 'Enregistreur de',
        'Sachet': 'Sachet',
        'Transparant': 'Transparent',
        'A ': 'à ',
        ' De ': ' de ',
        ' Le ': ' le ',
        ' La ': ' la ',
        ' Les ': ' les ',
        ' Et ': ' et ',
    }
    
    for ancien, nouveau in corrections.items():
        nom = nom.replace(ancien, nouveau)
    
    return nom.strip()


def get_unite_from_format(format_str):
    """Extraire l'unité depuis le format Hector Larivée"""
    
    if not format_str:
        return UniteMesure.objects.get_or_create(
            symbole='UN',
            defaults={'nom': 'Unité'}
        )[0]
    
    format_upper = format_str.upper()
    
    # Mapping spécifique pour Hector Larivée
    mappings = [
        (['CS', 'CAISSE'], 'CS', 'Caisse'),
        (['KG', 'KILO'], 'KG', 'Kilogramme'),
        (['G', 'GRAMME'], 'G', 'Gramme'),
        (['L', 'LITRE'], 'L', 'Litre'),
        (['ML', 'MILLILITRE'], 'ML', 'Millilitre'),
        (['BTE', 'BOITE', 'BOX'], 'BTE', 'Boîte'),
        (['PQT', 'PAQUET', 'PACK'], 'PQT', 'Paquet'),
        (['SAC', 'SACK'], 'SAC', 'Sac'),
        (['UN', 'UNITE', 'UNIT', 'PCS'], 'UN', 'Unité'),
    ]
    
    for keywords, symbole, nom in mappings:
        if any(kw in format_upper for kw in keywords):
            unite, _ = UniteMesure.objects.get_or_create(
                symbole=symbole,
                defaults={'nom': nom}
            )
            return unite
    
    # Par défaut
    return UniteMesure.objects.get_or_create(
        symbole='UN',
        defaults={'nom': 'Unité'}
    )[0]


def create_units_for_hector():
    """Créer toutes les unités nécessaires pour Hector Larivée"""
    unites = [
        ('UN', 'Unité'),
        ('CS', 'Caisse'),
        ('BTE', 'Boîte'),
        ('PQT', 'Paquet'),
        ('SAC', 'Sac'),
        ('KG', 'Kilogramme'),
        ('G', 'Gramme'),
        ('L', 'Litre'),
        ('ML', 'Millilitre'),
    ]
    
    for symbole, nom in unites:
        UniteMesure.objects.get_or_create(
            symbole=symbole,
            defaults={'nom': nom}
        )


# views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.models import User
from django.contrib import messages
from django.db.models import Q, Sum, Count, Avg
from django.utils import timezone
from datetime import timedelta
import json

from .models import (
    Character, BallSkin, PlayerSkin, PowerUp, DifficultyLevel,
    UserProfile, Purchase, Score, Leaderboard
)

# ==================== Game Views ====================

def home_view(request):
    """Page d'accueil du jeu"""
    context = {
        'user': request.user if request.user.is_authenticated else None
    }
    return render(request, 'listings/homegame.html', context)


# views.py - Modifier la vue game_view pour vérifier le personnage
@login_required
def game_view(request):
    """Vue principale du jeu"""
    profile, created = UserProfileFoot.objects.get_or_create(user=request.user)
    
    # Vérifier si le joueur a un personnage sélectionné
    if not profile.selected_character:
        messages.warning(request, "Vous devez d'abord acheter et sélectionner un personnage!")
        return redirect('game:shop')
    
    context = {
        'profile': profile,
        'character': profile.selected_character,
    }
    return render(request, 'listings/game.html', context)


# ==================== Shop Views ====================

@login_required
def shop_view(request):
    """Vue de la boutique"""
    profile = UserProfileFoot.objects.get(user=request.user)
    
    # Récupérer tous les items disponibles
    characters = Character.objects.filter(is_active=True).order_by('order')
    ball_skins = BallSkin.objects.filter(is_active=True).order_by('order')
    player_skins = PlayerSkin.objects.filter(is_active=True).order_by('order')
    powerups = PowerUp.objects.filter(is_active=True).order_by('order')
    
    # Récupérer les achats de l'utilisateur
    user_purchases = Purchase.objects.filter(user=request.user)
    owned_characters = user_purchases.filter(character__isnull=False).values_list('character_id', flat=True)
    owned_balls = user_purchases.filter(ball_skin__isnull=False).values_list('ball_skin_id', flat=True)
    owned_skins = user_purchases.filter(player_skin__isnull=False).values_list('player_skin_id', flat=True)
    
    context = {
        'profile': profile,
        'characters': characters,
        'ball_skins': ball_skins,
        'player_skins': player_skins,
        'powerups': powerups,
        'owned_characters': list(owned_characters),
        'owned_balls': list(owned_balls),
        'owned_skins': list(owned_skins),
    }
    return render(request, 'listings/shop.html', context)


@login_required
@csrf_exempt
def purchase_item(request):
    """API pour acheter un item"""
    if request.method == 'POST':
        data = json.loads(request.body)
        item_type = data.get('type')
        item_id = data.get('id')
        
        profile = UserProfileFoot.objects.get(user=request.user)
        
        try:
            if item_type == 'character':
                item = Character.objects.get(id=item_id)
                # Vérifier si déjà possédé
                if Purchase.objects.filter(user=request.user, character=item).exists():
                    return JsonResponse({'success': False, 'message': 'Déjà possédé'})
                
                # Vérifier les fonds
                if profile.coins >= item.price:
                    profile.coins -= item.price
                    profile.save()
                    Purchase.objects.create(
                        user=request.user,
                        character=item,
                        price=item.price,
                        currency='coins'
                    )
                    return JsonResponse({
                        'success': True, 
                        'message': f'{item.name} acheté avec succès!',
                        'coins': profile.coins
                    })
                else:
                    return JsonResponse({'success': False, 'message': 'Fonds insuffisants'})
                    
            elif item_type == 'ball':
                item = BallSkin.objects.get(id=item_id)
                if Purchase.objects.filter(user=request.user, ball_skin=item).exists():
                    return JsonResponse({'success': False, 'message': 'Déjà possédé'})
                
                if profile.coins >= item.price:
                    profile.coins -= item.price
                    profile.save()
                    Purchase.objects.create(
                        user=request.user,
                        ball_skin=item,
                        price=item.price,
                        currency='coins'
                    )
                    return JsonResponse({
                        'success': True,
                        'message': f'{item.name} acheté avec succès!',
                        'coins': profile.coins
                    })
                else:
                    return JsonResponse({'success': False, 'message': 'Fonds insuffisants'})
                    
            elif item_type == 'skin':
                item = PlayerSkin.objects.get(id=item_id)
                if Purchase.objects.filter(user=request.user, player_skin=item).exists():
                    return JsonResponse({'success': False, 'message': 'Déjà possédé'})
                
                if profile.coins >= item.price:
                    profile.coins -= item.price
                    profile.save()
                    Purchase.objects.create(
                        user=request.user,
                        player_skin=item,
                        price=item.price,
                        currency='coins'
                    )
                    return JsonResponse({
                        'success': True,
                        'message': f'{item.name} acheté avec succès!',
                        'coins': profile.coins
                    })
                else:
                    return JsonResponse({'success': False, 'message': 'Fonds insuffisants'})
                    
            elif item_type == 'powerup':
                item = PowerUp.objects.get(id=item_id)
                if profile.coins >= item.price:
                    profile.coins -= item.price
                    profile.save()
                    Purchase.objects.create(
                        user=request.user,
                        powerup=item,
                        price=item.price,
                        currency='coins'
                    )
                    return JsonResponse({
                        'success': True,
                        'message': f'{item.name} acheté avec succès!',
                        'coins': profile.coins
                    })
                else:
                    return JsonResponse({'success': False, 'message': 'Fonds insuffisants'})
                    
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'})


@login_required
@csrf_exempt
def equip_item(request):
    """API pour équiper un item"""
    if request.method == 'POST':
        data = json.loads(request.body)
        item_type = data.get('type')
        item_id = data.get('id')
        
        profile = UserProfileFoot.objects.get(user=request.user)
        
        try:
            if item_type == 'character':
                character = Character.objects.get(id=item_id)
                # Vérifier la possession
                if not Purchase.objects.filter(user=request.user, character=character).exists():
                    if character.price > 0:  # Pas gratuit
                        return JsonResponse({'success': False, 'message': 'Non possédé'})
                
                # NOUVEAU : Désélectionner le skin s'il ne correspond pas au nouveau personnage
                if profile.selected_player_skin:
                    if profile.selected_player_skin.character and profile.selected_player_skin.character != character:
                        profile.selected_player_skin = None
                
                profile.selected_character = character
                profile.save()
                return JsonResponse({'success': True, 'message': f'{character.name} équipé!'})
                
            elif item_type == 'ball':
                ball = BallSkin.objects.get(id=item_id)
                if not Purchase.objects.filter(user=request.user, ball_skin=ball).exists():
                    if ball.price > 0:
                        return JsonResponse({'success': False, 'message': 'Non possédé'})
                
                profile.selected_ball_skin = ball
                profile.save()
                return JsonResponse({'success': True, 'message': f'{ball.name} équipé!'})
                
            elif item_type == 'skin':
                skin = PlayerSkin.objects.get(id=item_id)
                if not Purchase.objects.filter(user=request.user, player_skin=skin).exists():
                    if skin.price > 0:
                        return JsonResponse({'success': False, 'message': 'Non possédé'})
                
                # NOUVEAU : Vérifier que le skin correspond au personnage actuel
                if skin.character:
                    if not profile.selected_character:
                        return JsonResponse({'success': False, 'message': 'Sélectionnez d\'abord un personnage!'})
                    
                    if skin.character != profile.selected_character:
                        return JsonResponse({
                            'success': False, 
                            'message': f'Ce skin appartient à {skin.character.name}, pas à {profile.selected_character.name}!'
                        })
                
                profile.selected_player_skin = skin
                profile.save()
                return JsonResponse({'success': True, 'message': f'{skin.name} équipé!'})
                
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'})

# ==================== Character Views ====================

@login_required
def characters_view(request):
    """Vue de sélection des personnages"""
    profile = UserProfileFoot.objects.get(user=request.user)
    characters = Character.objects.filter(is_active=True).order_by('order')
    
    # Récupérer les personnages possédés
    owned = Purchase.objects.filter(
        user=request.user,
        character__isnull=False
    ).values_list('character_id', flat=True)
    
    # Ajouter les personnages gratuits
    free_characters = Character.objects.filter(price=0).values_list('id', flat=True)
    owned_characters = list(set(list(owned) + list(free_characters)))
    
    context = {
        'profile': profile,
        'characters': characters,
        'owned_characters': owned_characters,
        'selected_character': profile.selected_character,
    }
    return render(request, 'listings/characters.html', context)


# ==================== Leaderboard Views ====================

@login_required
def leaderboard_view(request):
    """Vue du classement"""
    period = request.GET.get('period', 'daily')
    
    # Calculer les dates selon la période
    now = timezone.now()
    if period == 'daily':
        start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif period == 'weekly':
        start_date = now - timedelta(days=now.weekday())
        start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
    elif period == 'monthly':
        start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    else:  # all-time
        start_date = None
    
    # Récupérer les scores
    scores_query = Score.objects.all()
    if start_date:
        scores_query = scores_query.filter(date__gte=start_date)
    
    # Agréger par joueur
    leaderboard = scores_query.values('player__username', 'player__id').annotate(
        total_score=Sum('score'),
        games_played=Count('id'),
        avg_accuracy=Avg('successful_shots') * 100 / Avg('total_shots')
    ).order_by('-total_score')[:100]
    
    # Position du joueur actuel
    user_rank = None
    user_stats = None
    for idx, entry in enumerate(leaderboard):
        if entry['player__id'] == request.user.id:
            user_rank = idx + 1
            user_stats = entry
            break
    
    context = {
        'leaderboard': leaderboard[:50],  # Top 50
        'period': period,
        'user_rank': user_rank,
        'user_stats': user_stats,
    }
    return render(request, 'listings/leaderboard.html', context)


# ==================== Score API ====================


@login_required
@csrf_exempt
def save_score(request):
    """Sauvegarde le score et met à jour le profil avec les gains"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            score_value = int(data.get('score', 0))
            exp_earned = int(data.get('exp_earned', 0))
            
            # Calculer les gains côté serveur pour la sécurité
            coins_to_add = score_value // 10
            gems_to_add = score_value // 500
            
            print(f"Score: {score_value}, Coins: {coins_to_add}, Gems: {gems_to_add}, Exp: {exp_earned}")
            
            # Créer le score
            score_obj = Score.objects.create(
                player=request.user,
                score=score_value,
                total_shots=data.get('total_shots', 0),
                successful_shots=data.get('successful_shots', 0),
                targets_hit=data.get('targets_hit', 0),
                perfect_shots=data.get('perfect_shots', 0),
                max_combo=data.get('max_combo', 1),
                coins_earned=coins_to_add,
                exp_earned=exp_earned
            )
            
            # Récupérer ou créer le profil avec 500 coins de départ
            profile, created = UserProfileFoot.objects.get_or_create(
                user=request.user,
                defaults={
                    'coins': 500,  # Nouveau joueur = 500 coins
                    'gems': 0,
                    'level': 1,
                    'experience': 0
                }
            )
            
            # Mettre à jour le profil
            profile.coins += coins_to_add
            profile.gems += gems_to_add
            profile.experience += exp_earned
            profile.total_games += 1
            profile.total_goals += score_obj.successful_shots
            
            # Si meilleur score
            if score_value > profile.best_score:
                profile.best_score = score_value
            
            # Calculer le nouveau niveau (500 XP = 1 niveau)
            new_level = (profile.experience // 500) + 1
            level_up = False
            
            if new_level > profile.level:
                profile.level = new_level
                level_up = True
                # Bonus de niveau
                profile.coins += 100 * new_level
                profile.gems += 5
                print(f"Level up! Nouveau niveau: {new_level}")
            
            profile.save()
            
            print(f"Profil mis à jour - Coins: {profile.coins}, Gems: {profile.gems}, Level: {profile.level}, Exp: {profile.experience}")
            
            # Retourner les nouvelles valeurs
            return JsonResponse({
                'success': True,
                'total_coins': profile.coins,
                'total_gems': profile.gems,
                'new_level': profile.level if level_up else None,
                'experience': profile.experience,
                'coins_earned': coins_to_add,
                'gems_earned': gems_to_add,
                'exp_earned': exp_earned,
                'level_up': level_up
            })
            
        except Exception as e:
            print(f"Erreur lors de la sauvegarde: {str(e)}")
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)
    
    return JsonResponse({
        'success': False,
        'error': 'Method not allowed'
    }, status=405)
# ==================== Profile API ====================

@login_required
def get_profile_data(request):
    try:
        profile, created = UserProfileFoot.objects.get_or_create(
            user=request.user,
            defaults={
                'coins': 500,
                'gems': 0,
                'level': 1,
                'experience': 0,
                'selected_character': None
            }
        )
        
        # Vérifier si le skin appartient au personnage
        skin_image = None
        if profile.selected_player_skin and profile.selected_character:
            # Si le skin appartient au personnage actuel
            if profile.selected_player_skin.character == profile.selected_character:
                skin_image = profile.selected_player_skin.image.url
            # Sinon, désélectionner le skin
            else:
                profile.selected_player_skin = None
                profile.save()
        
        return JsonResponse({
            'success': True,
            'username': request.user.username,
            'coins': profile.coins,
            'gems': profile.gems,
            'level': profile.level,
            'experience': profile.experience,
            'has_character': profile.selected_character is not None,
            'selected_character_id': profile.selected_character.id if profile.selected_character else None,
            'selected_character_image': profile.selected_character.image.url if profile.selected_character and profile.selected_character.image else None,
            'selected_player_skin_image': skin_image,  # None si pas de skin compatible
            'selected_ball_skin_image': profile.selected_ball_skin.image.url if profile.selected_ball_skin and profile.selected_ball_skin.image else None,
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
# ==================== Authentication Views ====================


from django.shortcuts import render, redirect
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.models import User
from django.contrib import messages
from django.views.decorators.cache import never_cache
from django.views.decorators.csrf import csrf_protect

@csrf_protect
@never_cache
def login_view(request):
    """Vue de connexion"""
    # Si l'utilisateur est déjà connecté, rediriger vers le jeu
    if request.user.is_authenticated:
        return redirect('homegame')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        remember_me = request.POST.get('remember')
        
        # Authentifier l'utilisateur
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            
            # Gérer "Se souvenir de moi"
            if not remember_me:
                # Session expire à la fermeture du navigateur
                request.session.set_expiry(0)
            else:
                # Session expire après 30 jours
                request.session.set_expiry(60 * 60 * 24 * 30)
            
            # Créer le profil de jeu si nécessaire
            from .models import UserProfileFoot
            profile, created = UserProfileFoot.objects.get_or_create(
                user=user,
                defaults={
                    'coins': 500,  # Nouveau joueur = 500 coins
                    'gems': 0,
                    'level': 1,
                    'experience': 0
                }
            )
            
            if created:
                messages.success(request, f'Bienvenue {user.username}! Votre compte a été créé avec 500 pièces de départ!')
            else:
                messages.success(request, f'Bon retour {user.username}!')
            
            # Redirection après connexion
            next_url = request.GET.get('next', 'homegame')
            return redirect(next_url)
        else:
            # Vérifier si l'utilisateur existe
            try:
                User.objects.get(username=username)
                messages.error(request, 'Mot de passe incorrect.')
            except User.DoesNotExist:
                messages.error(request, 'Nom d\'utilisateur ou mot de passe incorrect.')
    
    return render(request, 'listings/logingame.html')

def register_view(request):
    """Vue d'inscription"""
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')
        
        if password1 != password2:
            messages.error(request, 'Les mots de passe ne correspondent pas.')
            return render(request, 'listings/register.html')
        
        if User.objects.filter(username=username).exists():
            messages.error(request, 'Ce nom d\'utilisateur existe déjà.')
            return render(request, 'listings/register.html')
        
        if User.objects.filter(email=email).exists():
            messages.error(request, 'Cet email est déjà utilisé.')
            return render(request, 'listings/register.html')
        
        # Créer l'utilisateur
        user = User.objects.create_user(
            username=username,
            email=email,
            password=password1
        )
        
        # Créer le profil
        UserProfileFoot.objects.create(user=user)
        
        # Connecter l'utilisateur
        login(request, user)
        messages.success(request, 'Inscription réussie! Bienvenue dans Free Kick Master!')
        
        return redirect('game:home')
    
    return render(request, 'listings/registergame.html')


def logout_view(request):
    """Vue de déconnexion"""
    logout(request)
    messages.success(request, 'Vous êtes déconnecté.')
    return redirect('logingame')


# ==================== Daily Bonus ====================

@login_required
def claim_daily_bonus(request):
    """API pour réclamer le bonus quotidien"""
    profile = UserProfileFoot.objects.get(user=request.user)
    now = timezone.now()
    
    # Vérifier si le bonus peut être réclamé
    if profile.last_daily_bonus:
        time_since = now - profile.last_daily_bonus
        if time_since < timedelta(hours=24):
            remaining = timedelta(hours=24) - time_since
            hours = int(remaining.total_seconds() // 3600)
            minutes = int((remaining.total_seconds() % 3600) // 60)
            return JsonResponse({
                'success': False,
                'message': f'Prochain bonus dans {hours}h {minutes}m'
            })
    
    # Donner le bonus
    bonus_coins = 100
    profile.coins += bonus_coins
    profile.last_daily_bonus = now
    profile.save()
    
    return JsonResponse({
        'success': True,
        'coins_earned': bonus_coins,
        'total_coins': profile.coins,
        'message': f'Vous avez reçu {bonus_coins} pièces!'
    })

@login_required
def test_profile(request):
    """Vue de test pour vérifier le profil"""
    from django.http import HttpResponse
    
    profile_exists = UserProfileFoot.objects.filter(user=request.user).exists()
    scores_count = Score1.objects.filter(user=request.user).count()
    
    html = f"""
    <h1>Test Profil pour {request.user.username}</h1>
    <p>Profil existe: {profile_exists}</p>
    <p>Nombre de scores: {scores_count}</p>
    """
    
    if profile_exists:
        profile = UserProfileFoot.objects.get(user=request.user)
        html += f"""
        <p>Coins: {profile.coins}</p>
        <p>Gems: {profile.gems}</p>
        <p>Level: {profile.level}</p>
        <p>Experience: {profile.experience}</p>
        """
    
    return HttpResponse(html)

# views.py - Ajouter cette fonction

from django.db.models import Max

@login_required
def shop_data_api(request):
    """API pour récupérer toutes les données de la boutique"""
    try:
        profile, created = UserProfileFoot.objects.get_or_create(
            user=request.user,
            defaults={'coins': 100, 'gems': 0, 'level': 1}
        )
        
        # Récupérer tous les items
        characters = Character.objects.filter(is_active=True).order_by('order')
        ball_skins = BallSkin.objects.filter(is_active=True).order_by('order')
        player_skins = PlayerSkin.objects.filter(is_active=True).order_by('order')
        powerups = PowerUp.objects.filter(is_active=True).order_by('order')
        
        # Récupérer les achats de l'utilisateur
        user_purchases = Purchase.objects.filter(user=request.user)
        owned_characters = list(user_purchases.filter(
            character__isnull=False
        ).values_list('character_id', flat=True))
        owned_balls = list(user_purchases.filter(
            ball_skin__isnull=False
        ).values_list('ball_skin_id', flat=True))
        owned_skins = list(user_purchases.filter(
            player_skin__isnull=False
        ).values_list('player_skin_id', flat=True))
        
        # Ajouter les personnages gratuits aux possessions
        free_chars = Character.objects.filter(price=0, is_active=True).values_list('id', flat=True)
        owned_characters.extend(free_chars)
        owned_characters = list(set(owned_characters))
        
        # Préparer les données
        characters_data = []
        for char in characters:
            char_data = {
                'id': char.id,
                'name': char.name,
                'description': char.description,
                'emoji': char.emoji,
                'image': char.image.url if char.image else None,
                'power': char.power,
                'precision': char.precision,
                'luck': char.luck,
                'curve': char.curve,
                'price': char.price,
                'overall_rating': char.overall_rating
            }
            characters_data.append(char_data)
        
        ball_skins_data = []
        for ball in ball_skins:
            ball_data = {
                'id': ball.id,
                'name': ball.name,
                'description': ball.description,
                'emoji': ball.emoji,
                'image': ball.image.url if ball.image else None,
                'price': ball.price,
                'trail_color': ball.trail_color,
                'has_particles': ball.has_particles
            }
            ball_skins_data.append(ball_data)
        
        player_skins_data = []
        for skin in player_skins:
            skin_data = {
                'id': skin.id,
                'name': skin.name,
                'description': skin.description,
                'emoji': skin.emoji,
                'image': skin.image.url if skin.image else None,
                'price': skin.price,
                'primary_color': skin.primary_color,
                'secondary_color': skin.secondary_color
            }
            player_skins_data.append(skin_data)
        
        powerups_data = []
        for powerup in powerups:
            powerup_data = {
                'id': powerup.id,
                'name': powerup.name,
                'powerup_type': powerup.powerup_type,
                'description': powerup.description,
                'emoji': powerup.emoji,
                'image': powerup.image.url if powerup.image else None,
                'price': powerup.price,
                'duration': powerup.duration,
                'effect_value': powerup.effect_value
            }
            powerups_data.append(powerup_data)
        
        # Items équipés
        equipped_items = {
            'character': profile.selected_character.id if profile.selected_character else None,
            'ball': profile.selected_ball_skin.id if profile.selected_ball_skin else None,
            'skin': profile.selected_player_skin.id if profile.selected_player_skin else None
        }
        
        return JsonResponse({
            'success': True,
            'characters': characters_data,
            'ball_skins': ball_skins_data,
            'player_skins': player_skins_data,
            'powerups': powerups_data,
            'owned_characters': owned_characters,
            'owned_balls': owned_balls,
            'owned_skins': owned_skins,
            'equipped_items': equipped_items,
            'profile': {
                'coins': profile.coins,
                'gems': profile.gems,
                'level': profile.level
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
    
# views.py - Ajouter cette fonction

@login_required
def characters_data_api(request):
    """API pour récupérer les données des personnages"""
    try:
        profile, created = UserProfileFoot.objects.get_or_create(
            user=request.user,
            defaults={'coins': 100, 'gems': 0, 'level': 1}
        )
        
        # Récupérer tous les personnages
        characters = Character.objects.filter(is_active=True).order_by('order')
        
        # Récupérer les personnages possédés
        user_purchases = Purchase.objects.filter(user=request.user)
        owned_characters = list(user_purchases.filter(
            character__isnull=False
        ).values_list('character_id', flat=True))
        
        # Ajouter les personnages gratuits
        free_chars = Character.objects.filter(price=0, is_active=True).values_list('id', flat=True)
        owned_characters.extend(free_chars)
        owned_characters = list(set(owned_characters))
        
        # Personnage sélectionné
        selected_character_id = profile.selected_character.id if profile.selected_character else None
        
        # Si aucun personnage sélectionné, sélectionner le premier gratuit
        if not selected_character_id and len(owned_characters) > 0:
            first_owned = characters.filter(id__in=owned_characters).first()
            if first_owned:
                profile.selected_character = first_owned
                profile.save()
                selected_character_id = first_owned.id
        
        # Préparer les données des personnages
        characters_data = []
        for char in characters:
            char_data = {
                'id': char.id,
                'name': char.name,
                'description': char.description,
                'emoji': char.emoji,
                'image': char.image.url if char.image else None,
                'power': char.power,
                'precision': char.precision,
                'luck': char.luck,
                'curve': char.curve,
                'price': char.price,
                'overall_rating': char.overall_rating,
                'strengths': char.strengths,
                'weaknesses': char.weaknesses
            }
            characters_data.append(char_data)
        
        return JsonResponse({
            'success': True,
            'characters': characters_data,
            'owned_characters': owned_characters,
            'selected_character_id': selected_character_id
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
    
# views.py - Ajouter cette fonction améliorée

from django.db.models import Q, Sum, Count, Avg, F, Max
from datetime import datetime, timedelta

@login_required
def leaderboard_api(request):
    """API pour récupérer les données du classement"""
    period = request.GET.get('period', 'daily')
    
    try:
        # Calculer les dates selon la période
        now = timezone.now()
        if period == 'daily':
            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
        elif period == 'weekly':
            start_date = now - timedelta(days=now.weekday())
            start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
        elif period == 'monthly':
            start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        else:  # alltime
            start_date = None
        
        # Récupérer les scores
        scores_query = Score.objects.exclude(player__isnull=True)
        if start_date:
            scores_query = scores_query.filter(date__gte=start_date)
        
        # Agréger par joueur - ajuster selon votre modèle
        leaderboard = scores_query.values(
            'player__username', 
            'player__id'
        ).annotate(
            total_score=Sum('score'),
            games_played=Count('id'),
            avg_accuracy=Avg(
                F('successful_shots') * 100.0 / F('total_shots'),
                output_field=models.FloatField()
            )
        ).order_by('-total_score')[:100]
        
        # Formater les données du classement
        leaderboard_data = []
        user_rank = None
        user_stats = None
        
        for idx, entry in enumerate(leaderboard):
            player_data = {
                'rank': idx + 1,
                'user_id': entry['player__id'],
                'username': entry['player__username'],
                'total_score': entry['total_score'] or 0,
                'games_played': entry['games_played'] or 0,
                'avg_accuracy': round(entry['avg_accuracy'] or 0, 1)
            }
            leaderboard_data.append(player_data)
            
            # Vérifier si c'est le joueur actuel
            if entry['player__id'] == request.user.id:
                user_rank = idx + 1
                user_stats = {
                    'rank': user_rank,
                    'user_id': request.user.id,
                    'username': request.user.username,
                    'total_score': entry['total_score'] or 0,
                    'games_played': entry['games_played'] or 0,
                    'avg_accuracy': round(entry['avg_accuracy'] or 0, 1)
                }
        
        # Si le joueur n'est pas dans le top 100, récupérer ses stats
        if not user_stats:
            user_scores = scores_query.filter(player=request.user)
            if user_scores.exists():
                user_data = user_scores.aggregate(
                    total_score=Sum('score'),
                    games_played=Count('id'),
                    avg_accuracy=Avg(
                        F('successful_shots') * 100.0 / F('total_shots'),
                        output_field=models.FloatField()
                    )
                )
                
                # Calculer le rang
                better_scores = scores_query.values('player').annotate(
                    total=Sum('score')
                ).filter(total__gt=user_data['total_score']).count()
                
                user_stats = {
                    'rank': better_scores + 1,
                    'user_id': request.user.id,
                    'username': request.user.username,
                    'total_score': user_data['total_score'] or 0,
                    'games_played': user_data['games_played'] or 0,
                    'avg_accuracy': round(user_data['avg_accuracy'] or 0, 1)
                }
        
        return JsonResponse({
            'success': True,
            'leaderboard': leaderboard_data,
            'user_stats': user_stats,
            'period': period
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)